import contextlib
import copy
import csv
import hashlib
import json
import logging
import os
import re
import time
import traceback
import uuid
from datetime import date
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Optional

import requests
from celery import chord
from celery.exceptions import MaxRetriesExceededError, TimeoutError
from dateutil.rrule import DAILY, rrule
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.core.mail import mail_admins as django_mail_admins
from django.core.mail import mail_managers as django_mail_managers
from django.core.mail import send_mail as django_send_mail
from django.core.mail import send_mass_mail as django_send_mass_mail
from django.db import transaction
from django.utils import timezone
from django.utils.timezone import now
from django.utils.translation import gettext_lazy
from filtration import Expression
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from poms.accounts.models import Account
from poms.celery_tasks import finmars_task
from poms.celery_tasks.models import CeleryTask
from poms.expressions_engine import formula
from poms.common.middleware import activate
from poms.common.crypto.AESCipher import AESCipher
from poms.common.crypto.RSACipher import RSACipher
from poms.integrations.database_client import DatabaseService, get_backend_callback_url
from poms.expressions_engine.formula import ExpressionEvalError
from poms.common.jwt import encode_with_jwt
from poms.common.models import ProxyRequest, ProxyUser
from poms.integrations.monad import Monad, MonadStatus
from poms.common.storage import get_storage
from poms.counterparties.models import Counterparty, Responsible
from poms.counterparties.serializers import CounterpartySerializer
from poms.csv_import.handlers import handler_instrument_object
from poms.currencies.models import Currency, CurrencyHistory
from poms.file_reports.models import FileReport
from poms.instruments.models import (
    AccrualCalculationModel,
    DailyPricingModel,
    Instrument,
    InstrumentType,
    PaymentSizeDetail,
    Periodicity,
    PriceHistory,
    PricingCondition,
)
from poms.integrations.models import (
    AccountMapping,
    AccrualCalculationModelMapping,
    BloombergDataProviderCredential,
    ComplexTransactionImportScheme,
    CounterpartyMapping,
    CurrencyMapping,
    DailyPricingModelMapping,
    ImportConfig,
    InstrumentDownloadScheme,
    InstrumentMapping,
    InstrumentTypeMapping,
    PaymentSizeDetailMapping,
    PeriodicityMapping,
    PortfolioMapping,
    PriceDownloadScheme,
    PriceDownloadSchemeMapping,
    ResponsibleMapping,
    Strategy1Mapping,
    Strategy2Mapping,
    Strategy3Mapping,
    TransactionFileResult,
)
from poms.integrations.providers.base import (
    AbstractProvider,
    get_provider,
    parse_date_iso,
)
from poms.obj_attrs.models import GenericAttributeType
from poms.portfolios.models import Portfolio
from poms.procedures.models import RequestDataFileProcedureInstance
from poms.strategies.models import Strategy1, Strategy2, Strategy3
from poms.system_messages.handlers import send_system_message
from poms.transaction_import.tasks import transaction_import
from poms.transactions.handlers import TransactionTypeProcess
from poms.users.models import EcosystemDefault

_l = logging.getLogger("poms.integrations")

TYPE_PREFIX = "com.finmars.initial-instrument-type:"

storage = get_storage()


@finmars_task(name="integrations.health_check")
def health_check_async(*args, **kwargs):
    return True


def health_check():
    result = health_check_async.apply_async()
    with contextlib.suppress(TimeoutError):
        return result.get(timeout=0.5, interval=0.1)

    return False


@finmars_task(name="integrations.send_mail_async", ignore_result=True)
def send_mail_async(subject, message, from_email, recipient_list, html_message=None, *args, **kwargs):
    django_send_mail(
        subject,
        message,
        from_email,
        recipient_list,
        fail_silently=True,
        html_message=html_message,
    )


def send_mail(subject, message, from_email, recipient_list, html_message=None, space_code=None, realm_code=None):
    send_mail_async.apply_async(
        kwargs={
            "subject": subject,
            "message": message,
            "from_email": from_email,
            "recipient_list": recipient_list,
            "html_message": html_message, 'context': {
                'space_code': space_code,
                'realm_code': realm_code
            }
        }
    )


@finmars_task(name="integrations.send_mass_mail", ignore_result=True)
def send_mass_mail_async(messages, *args, **kwargs):
    django_send_mass_mail(messages, fail_silently=True)


def send_mass_mail(messages):
    send_mass_mail_async.apply_async(
        kwargs={
            "messages": messages,
        }
    )


@finmars_task(name="integrations.mail_admins", ignore_result=True)
def mail_admins_async(subject, message, *args, **kwargs):
    django_mail_admins(
        subject,
        message,
        fail_silently=True,
    )


def mail_admins(subject, message):
    mail_admins_async.apply_async(
        kwargs={
            "subject": subject,
            "message": message,
        }
    )


@finmars_task(name="integrations.mail_managers", ignore_result=True)
def mail_managers_async(subject, message, *args, **kwargs):
    django_mail_managers(
        subject,
        message,
        fail_silently=True,
    )


def mail_managers(subject, message):
    mail_managers_async.apply_async(
        kwargs={
            "subject": subject,
            "message": message,
        }
    )


def update_task_with_error(task: CeleryTask, err_msg: str):
    _l.error(err_msg)
    task.error_message = err_msg
    task.status = CeleryTask.STATUS_ERROR
    task.save()


def task_done_with_instrument_info(instrument: Instrument, task: CeleryTask):
    if not instrument or not task:
        _l.error(
            f"update_task_with_instrument error: missing task={task} or "
            f"instrument={instrument}!"
        )
        return

    result = task.result_object or {}
    result["result_id"] = instrument.pk
    result["name"] = instrument.name
    result["short_name"] = instrument.short_name
    result["user_code"] = instrument.user_code
    task.result_object = result
    task.status = CeleryTask.STATUS_DONE
    task.save()


@finmars_task(name="integrations.download_instrument", bind=True, ignore_result=False)
def download_instrument_async(self, task_id=None, *args, **kwargs):
    task = CeleryTask.objects.get(pk=task_id)
    _l.debug(
        "download_instrument_async: master_user_id=%s, task=%s",
        task.master_user_id,
        task,
    )

    task.celery_task_id = self.request.id

    provider_id = 1  # bloomberg

    try:
        provider = get_provider(task.master_user, provider_id)
    except Exception:
        _l.debug("provider load error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        raise

    if provider is None:
        _l.debug("provider not found")
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if task.status not in [CeleryTask.STATUS_PENDING, CeleryTask.STATUS_WAIT_RESPONSE]:
        _l.debug("invalid task status")
        return
    options = task.options_object

    try:
        result, is_ready = provider.download_instrument(options)
    except Exception:
        _l.error("provider processing error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
    else:
        if is_ready:
            _l.info(f"download_instrument_async.result {result}")

            task.status = CeleryTask.STATUS_DONE
            task.result_object = result
        else:
            task.status = CeleryTask.STATUS_WAIT_RESPONSE

    if response_id := options.get("response_id", None):
        task.response_id = response_id

    task.options_object = options
    task.save()

    if task.status == CeleryTask.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(
                countdown=provider.get_retry_delay(),
                max_retries=provider.get_max_retries(),
            )
            # self.retry(countdown=provider.get_retry_delay(), max_retries=
            # provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = CeleryTask.STATUS_TIMEOUT
            task.save()
        return

    return task_id


def download_instrument(
    instrument_code=None,
    instrument_download_scheme=None,
    master_user=None,
    member=None,
    task=None,
    value_overrides=None,
):
    _l.info(f"download_instrument value_overrides {value_overrides}")
    _l.info(
        "download_instrument: master_user_id=%s, task=%s, instrument_code=%s, instrument_download_scheme=%s",
        getattr(master_user, "id", None),
        getattr(task, "info", None),
        instrument_code,
        instrument_download_scheme,
    )

    if task is None:
        provider = get_provider(
            instrument_download_scheme.master_user,
            instrument_download_scheme.provider,
        )
        if not provider.is_valid_reference(instrument_code):
            raise ValueError("Invalid instrument_code value")

        options = {
            "instrument_download_scheme_id": instrument_download_scheme.id,
            "instrument_code": instrument_code,
        }
        with transaction.atomic():
            task = CeleryTask(
                master_user=master_user,
                member=member,
                status=CeleryTask.STATUS_PENDING,
                type="download_instrument",
            )
            task.options_object = options
            task.save()
            transaction.on_commit(
                lambda: download_instrument_async.apply_async(
                    kwargs={"task_id": task.id, 'context': {
                        'space_code': task.master_user.space_code,
                        'realm_code': task.master_user.realm_code
                    }}
                )
            )
    elif task.status == CeleryTask.STATUS_DONE:
        return create_instrument_with_updated_data(task, value_overrides)

    return task, None, None


def create_instrument_with_updated_data(task, value_overrides):
    provider = get_provider(task.master_user, 1)
    options = task.options_object
    values = task.result_object.copy()
    if value_overrides:
        values.update(value_overrides)

    instrument_download_scheme_id = options["instrument_download_scheme_id"]
    instrument_download_scheme = InstrumentDownloadScheme.objects.get(
        pk=instrument_download_scheme_id
    )
    instrument, errors = provider.create_instrument(
        instrument_download_scheme,
        values,
    )
    return task, instrument, errors


def create_instrument_from_finmars_database(data, master_user, member):
    from poms.instruments.serializers import InstrumentSerializer

    func = "create_instrument_from_finmars_database"
    _l.info(f"{func} data={data}")

    instrument_data = {
        key: None if value == "null" else value for key, value in data.items()
    }
    short_type = instrument_data["instrument_type"]["user_code"]
    try:
        if short_type in {"stocks", "stock"}:
            if (
                "default_exchange" in instrument_data
                and instrument_data["default_exchange"]
                and "default_currency_code" in instrument_data
                and instrument_data["default_currency_code"]
            ):
                if "." in instrument_data["user_code"]:
                    if ":" in instrument_data["user_code"]:
                        instrument_data["reference_for_pricing"] = instrument_data[
                            "user_code"
                        ]
                    else:
                        instrument_data["reference_for_pricing"] = (
                            instrument_data["user_code"]
                            + ":"
                            + instrument_data["default_currency_code"]
                        )
                else:
                    instrument_data["reference_for_pricing"] = (
                        instrument_data["user_code"]
                        + "."
                        + instrument_data["default_exchange"]
                        + ":"
                        + instrument_data["default_currency_code"]
                    )

                _l.info(
                    f"{func} Reference for pricing updated "
                    f"{instrument_data['reference_for_pricing']}"
                )

            _l.info(f"{func} Overwrite Pricing Currency for stock")

            if "default_currency_code" in instrument_data:
                instrument_data["pricing_currency"] = instrument_data[
                    "default_currency_code"
                ]

        instrument_type_user_code_full = f"{TYPE_PREFIX}{short_type}"
        try:
            instrument_type = InstrumentType.objects.get(
                master_user=master_user,
                user_code=instrument_type_user_code_full,
                # user_code__contains=short_type,  #TODO FOR DEBUG ONLY!
            )
        except InstrumentType.DoesNotExist as e:
            err_msg = f"{func} no such InstrumentType user_code={instrument_type_user_code_full}"
            _l.error(err_msg)
            raise RuntimeError(err_msg) from e

        ecosystem_defaults = EcosystemDefault.objects.get(master_user=master_user)
        content_type = ContentType.objects.get(
            model="instrument", app_label="instruments"
        )

        attribute_types = GenericAttributeType.objects.filter(
            master_user=master_user, content_type=content_type
        )
        object_data = handler_instrument_object(
            instrument_data,
            instrument_type,
            master_user,
            ecosystem_defaults,
            attribute_types,
        )
        object_data["short_name"] = (
            object_data["name"] + " (" + object_data["user_code"] + ")"
        )

        proxy_request = ProxyRequest(ProxyUser(member, master_user))
        activate(proxy_request)
        context = {
            "master_user": master_user,
            "request": proxy_request,
            "member": member,
        }

        try:
            instance = Instrument.objects.get(
                master_user=master_user, user_code=object_data["user_code"]
            )

            instance.is_active = True

            serializer = InstrumentSerializer(
                data=object_data,
                context=context,
                instance=instance,
            )
        except Instrument.DoesNotExist:
            serializer = InstrumentSerializer(data=object_data, context=context)

        if serializer.is_valid():
            instrument = serializer.save()

            _l.info(
                f"{func} Instrument {instrument.user_code} was imported successfully"
            )

            return instrument

        else:
            err_msg = f"{func} InstrumentSerializer error={serializer.errors}"
            _l.error(err_msg)
            raise RuntimeError(err_msg)

    except Exception as e:
        _l.info(f"{func} {repr(e)} {traceback.format_exc()}")
        raise e


def create_instrument_cbond(data, master_user, member):
    from poms.instruments.serializers import InstrumentSerializer

    try:
        ecosystem_defaults = EcosystemDefault.objects.get(master_user=master_user)
        content_type = ContentType.objects.get(
            model="instrument", app_label="instruments"
        )

        proxy_user = ProxyUser(member, master_user)
        proxy_request = ProxyRequest(proxy_user)

        context = {"master_user": master_user, "request": proxy_request}

        instrument_data = {
            key: None if value == "null" else value for key, value in data.items()
        }
        if instrument_data["instrument_type"] == "stock":
            if (
                "default_exchange" in instrument_data
                and instrument_data["default_exchange"]
                and "default_currency_code" in instrument_data
                and instrument_data["default_currency_code"]
            ):
                # isin.exchange:currency

                if "." in instrument_data["user_code"]:
                    if ":" in instrument_data["user_code"]:
                        instrument_data["reference_for_pricing"] = instrument_data[
                            "user_code"
                        ]
                    else:
                        instrument_data["reference_for_pricing"] = (
                            instrument_data["user_code"]
                            + ":"
                            + instrument_data["default_currency_code"]
                        )
                else:
                    instrument_data["reference_for_pricing"] = (
                        instrument_data["user_code"]
                        + "."
                        + instrument_data["default_exchange"]
                        + ":"
                        + instrument_data["default_currency_code"]
                    )

                _l.info(
                    f'Reference for pricing updated {instrument_data["reference_for_pricing"]}'
                )

            _l.info("Overwrite Pricing Currency for stock")
            if "default_currency_code" in instrument_data:
                instrument_data["pricing_currency"] = instrument_data[
                    "default_currency_code"
                ]

        attribute_types = GenericAttributeType.objects.filter(
            master_user=master_user, content_type=content_type
        )

        try:
            user_code = f"{TYPE_PREFIX}{instrument_data['instrument_type']}"

            instrument_type = InstrumentType.objects.get(
                master_user=master_user,
                user_code=user_code
                # user_code__contains=instrument_data["instrument_type"],
            )

        except Exception as e:
            err_msg = (
                f'InstrumentType user_code={instrument_data["instrument_type"]} '
                f"master_user={master_user.id} not found {e}"
            )
            _l.info(err_msg)
            raise RuntimeError(err_msg) from e

        object_data = handler_instrument_object(
            instrument_data,
            instrument_type,
            master_user,
            ecosystem_defaults,
            attribute_types,
        )

        object_data["short_name"] = (
            object_data["name"] + " (" + object_data["user_code"] + ")"
        )

        try:
            instance = Instrument.objects.get(
                master_user=master_user, user_code=object_data["user_code"]
            )

            instance.is_active = True

            serializer = InstrumentSerializer(
                data=object_data, context=context, instance=instance
            )
        except Instrument.DoesNotExist:
            serializer = InstrumentSerializer(data=object_data, context=context)

        if serializer.is_valid():
            instrument = serializer.save()

            _l.info("Instrument is imported successfully")

            return instrument
        else:
            _l.error(f"InstrumentExternalAPIViewSet error {serializer.errors}")
            raise RuntimeError(serializer.errors)

    except Exception as e:
        _l.error(
            f"InstrumentExternalAPIViewSet error {repr(e)} {traceback.format_exc()}"
        )
        raise e


def download_instrument_cbond(
    instrument_code=None,
    instrument_name=None,
    instrument_type_code=None,
    master_user=None,
    member=None,
):
    errors = []

    try:
        _l.debug(
            "download_instrument_cbond: master_user_id=%s, instrument_code=%s",
            getattr(master_user, "id", None),
            instrument_code,
        )
        options = {
            "isin": instrument_code,
        }
        with transaction.atomic():
            task = CeleryTask(
                master_user=master_user,
                member=member,
                status=CeleryTask.STATUS_PENDING,
                type="download_instrument",
            )
            task.options_object = options
            task.save()

        payload_jwt = {
            "sub": task.master_user.space_code,  # "user_id_or_name",
            "role": 0,  # 0 -- ordinary user, 1 -- admin (access to /loadfi and /loadeq)
        }

        token = encode_with_jwt(payload_jwt)

        headers = {
            "Content-type": "application/json",
            "Authorization": f"Bearer {token}",
        }
        options["request_id"] = task.pk
        options["base_api_url"] = master_user.space_code

        if (master_user.realm_code):
            options["callback_url"] = (
                f"https://{settings.DOMAIN_NAME}/{master_user.realm_code}/{master_user.space_code}"
                f"/api/instruments/fdb-create-from-callback/"
            )
        else:

            options["callback_url"] = (
                f"https://{settings.DOMAIN_NAME}/{master_user.space_code}"
                f"/api/instruments/fdb-create-from-callback/"
            )

        options["token"] = "fd09a190279e45a2bbb52fcabb7899bd"

        options["data"] = {}

        task.options_object = options
        task.save()

        response = None

        _l.info(f"options {options}")
        _l.info(f"settings.CBONDS_BROKER_URL {settings.CBONDS_BROKER_URL}")

        try:
            response = requests.post(
                url=f"{str(settings.CBONDS_BROKER_URL)}export/",
                data=json.dumps(options),
                headers=headers,
                timeout=25,
            )
            _l.info(f"response download_instrument_cbond {response}")
            _l.info(f"data response.text {response.text} ")
            _l.info(f"data response.status_code {response.status_code} ")

        except requests.exceptions.Timeout:
            _l.info(
                f"Finmars Database Timeout. Trying to create simple "
                f"instrument {instrument_code}"
            )

            try:
                Instrument.objects.get(
                    master_user=master_user,
                    user_code=instrument_code,
                )

                _l.info(
                    f"Finmars Database Timeout. Simple instrument {instrument_code}"
                    f" exist. Abort."
                )

            except Exception:
                itype = None

                if instrument_type_code == "equity":
                    instrument_type_code = "stock"

                _l.info(
                    f"Finmars Database Timeout. instrument_type_code "
                    f"{instrument_type_code} instrument_name {instrument_name}"
                )

                if instrument_type_code:
                    try:
                        itype = InstrumentType.objects.get(
                            master_user=master_user, user_code=instrument_type_code
                        )
                    except Exception:
                        itype = None

                if not instrument_name:
                    instrument_name = instrument_code

                ecosystem_defaults = EcosystemDefault.objects.get(
                    master_user=master_user
                )

                instrument = Instrument.objects.create(
                    master_user=master_user,
                    user_code=instrument_code,
                    name=instrument_name,
                    short_name=f"{instrument_name} ({instrument_code})",
                    instrument_type=ecosystem_defaults.instrument_type,
                    accrued_currency=ecosystem_defaults.currency,
                    pricing_currency=ecosystem_defaults.currency,
                    co_directional_exposure_currency=ecosystem_defaults.currency,
                    counter_directional_exposure_currency=ecosystem_defaults.currency,
                )

                instrument.is_active = False

                if itype:
                    instrument.instrument_type = itype

                    small_item = {
                        "user_code": instrument_code,
                        "instrument_type": instrument_type_code,
                    }

                    create_instrument_cbond(small_item, master_user, member)

                instrument.save()

                task_done_with_instrument_info(instrument, task)
            return task, errors

        except Exception as e:
            _l.debug(f"Can't send request to CBONDS BROKER. {e}")

            errors.append(f"Request to broker failed. {str(e)}")

        try:
            data = response.json()
            _l.info(f"Cbond response data {data}")
        except Exception as e:
            _l.info(f"Cbond response data Exception {e}")

            errors.append(f"Could not parse response from broker. {response.text}")
            return task, errors

        try:
            result_instrument = None

            if "instruments" in data:
                if "currencies" in data:
                    for item in data["currencies"]:
                        if item:
                            create_currency_from_callback_data(
                                item, master_user, member
                            )

                for item in data["instruments"]:
                    instrument = create_instrument_cbond(item, master_user, member)

                    if instrument.user_code == instrument_code:
                        result_instrument = instrument

            elif "items" in data["data"]:
                for item in data["data"]["items"]:
                    instrument = create_instrument_cbond(item, master_user, member)

                    if instrument.user_code == instrument_code:
                        result_instrument = instrument

            else:
                instrument = create_instrument_cbond(data["data"], master_user, member)
                result_instrument = instrument

            task_done_with_instrument_info(result_instrument, task)

        except Exception as e:
            errors.append(f"Could not create instrument. {str(e)}")
            return task, errors

        _l.info(f"data {data} ")

        return task, errors

    except Exception as e:
        _l.info(f"error {e} traceback.format_exc()")
        errors.append(f"Something went wrong. {str(e)}")

        return None, errors


def download_currency_cbond(currency_code=None, master_user=None, member=None):
    errors = []

    try:
        _l.debug(
            "download_currency_cbond: master_user_id=%s, currency_code=%s",
            getattr(master_user, "id", None),
            currency_code,
        )

        options = {
            "code": currency_code,
        }
        with transaction.atomic():
            task = CeleryTask(
                master_user=master_user,
                member=member,
                status=CeleryTask.STATUS_PENDING,
                type="download_currency",
            )
            task.options_object = options
            task.save()

            payload_jwt = {
                "sub": task.master_user.space_code,  # "user_id_or_name",
                "role": 0,  # 0 - ordinary user, 1 - admin (access to /loadfi & /loadeq)
            }

            token = encode_with_jwt(payload_jwt)

            headers = {
                "Content-type": "application/json",
                "Authorization": f"Bearer {token}",
            }
            options["request_id"] = task.pk
            options["base_api_url"] = master_user.space_code
            options["token"] = "fd09a190279e45a2bbb52fcabb7899bd"

            options["data"] = {}

            response = None

            _l.info(f"options {options}")
            _l.info(f"settings.CBONDS_BROKER_URL {settings.CBONDS_BROKER_URL}")

            try:
                # refactor to /export/currency when available
                response = requests.get(
                    url=f"{str(settings.CBONDS_BROKER_URL)}instr/currency/{currency_code}",
                    headers=headers,
                    verify=settings.VERIFY_SSL,
                )
                _l.info(f"response download_currency_cbond {response}")
                _l.info(f"data response.text {response.text} ")
            except Exception as e:
                _l.debug(f"Can't send request to CBONDS BROKER. {e}")

                errors.append(f"Request to broker failed. {str(e)}")

            try:
                data = response.json()
            except Exception as e:
                errors.append(f"Could not parse response from broker. {response.text}")
                return task, errors

            try:
                currency = create_currency_from_callback_data(data, master_user, member)

                result = {"currency_id": currency.pk}
                task.result_object = result
                task.save()

            except Exception as e:
                errors.append(f"Could not create currency. {str(e)}")
                return task, errors

            _l.info(f"data {data} ")

            return task, errors

    except Exception as e:
        _l.info(f"error {e} {traceback.format_exc()}")

        errors.append(f"Something went wrong. {str(e)}")

        return None, errors


def task_error(err_msg, task):
    _l.error(err_msg)
    task.error_message = err_msg
    return None


def create_simple_instrument(task: CeleryTask) -> Optional[Instrument]:
    from poms.instruments.handlers import InstrumentTypeProcess
    from poms.instruments.serializers import InstrumentSerializer

    func = f"create_simple_instrument task.id={task.id}"
    options_data = task.options_object["data"]
    _l.info(f"{func} started options_data={options_data}")

    type_user_type = options_data["type_user_code"]
    instrument_type_user_code_full = f"{TYPE_PREFIX}{type_user_type}"
    try:
        instrument_type = InstrumentType.objects.get(
            master_user=task.master_user,
            user_code=instrument_type_user_code_full,
            # user_code__contains=type_user_type,  #TODO FOR DEBUG ONLY!
        )
    except InstrumentType.DoesNotExist:
        err_msg = (
            f"{func} no such InstrumentType user_code={instrument_type_user_code_full}"
        )
        return task_error(err_msg, task)

    process = InstrumentTypeProcess(instrument_type=instrument_type)
    instrument_dict = process.instrument
    instrument_dict["name"] = options_data["name"]
    instrument_dict["user_code"] = options_data["user_code"]
    instrument_dict["is_active"] = False
    context = {
        "master_user": task.master_user,
        "request": ProxyRequest(ProxyUser(task.member, task.master_user)),
    }
    serializer = InstrumentSerializer(
        data=instrument_dict,
        context=context,
    )
    if not serializer.is_valid():
        err_msg = f"{func} instrument validation errors={serializer.errors}"
        return task_error(err_msg, task)

    instrument = serializer.save()
    _l.info(f"{func} created instrument.id={instrument.id}")
    return instrument


@finmars_task(
    name="integrations.download_instrument_cbond_task", bind=True, ignore_result=False
)
def download_instrument_cbond_task(self, task_id, *args, **kwargs):
    task = CeleryTask.objects.get(pk=task_id)

    instrument_type_code = None

    name = task.options["name"] if "name" in task.options else None
    if "instrument_type_code" in task.options:
        instrument_type_code = task.options["instrument_type_code"]

    download_instrument_cbond(
        task.options["user_code"],
        name,
        instrument_type_code,
        task.master_user,
        task.member,
    )


def download_unified_data(
    id=None,
    entity_type=None,
    master_user=None,
    member=None,
    task=None,
    value_overrides=None,
):
    errors = []

    try:
        with transaction.atomic():
            # DEPRECATED, REFACTOR SOON
            task = CeleryTask(
                master_user=master_user,
                member=member,
                status=CeleryTask.STATUS_PENDING,
                type="download_entity",
            )
            task.options_object = {"entity_type": entity_type, "id": id}
            task.save()

            headers = {"Content-type": "application/json"}

            response = None

            path = "company" if entity_type == "counterparty" else ""
            try:
                response = requests.get(
                    url=f"{str(settings.UNIFIED_DATA_PROVIDER_URL)}data/{path}/{id}/",
                    headers=headers,
                    verify=settings.VERIFY_SSL,
                )
                _l.info(f"response download_unified_data {response}")
                _l.info(f"data response.text {response.text} ")
            except Exception as e:
                _l.debug(f"Can't send request to Unified Data Provider. {e}")

                errors.append(f"Request to unified data provider. {str(e)}")

            try:
                data = response.json()
            except Exception:
                errors.append(
                    f"Could not parse response from unified data provider. {response.text}"
                )
                return task, errors
            try:
                obj_data = data

                proxy_user = ProxyUser(member, master_user)
                proxy_request = ProxyRequest(proxy_user)

                context = {"request": proxy_request}

                ecosystem_defaults = EcosystemDefault.objects.get(
                    master_user=master_user
                )

                record = None

                if entity_type == "counterparty":
                    obj_data["group"] = ecosystem_defaults.counterparty_group_id

                    serializer = CounterpartySerializer(data=obj_data, context=context)
                    serializer.is_valid(raise_exception=True)
                    record = serializer.save()

                result = {"id": record.pk}
                task.result_object = result
                task.save()

            except Exception as e:
                errors.append(f"Could not create record. {str(e)}")
                return task, errors

            return task, errors

    except Exception as e:
        _l.info(f"error {e} ")
        _l.info(traceback.format_exc())

        errors.append(f"Something went wrong. {str(e)}")

        return None, errors


@finmars_task(
    name="integrations.download_instrument_pricing_async",
    bind=True,
    ignore_result=False,
)
def download_instrument_pricing_async(self, task_id, *args, **kwargs):
    task = CeleryTask.objects.get(pk=task_id)
    _l.debug(
        "download_instrument_pricing_async: master_user_id=%s, task=%s",
        task.master_user_id,
        task,
    )

    task.celery_task_id = self.request.id

    try:
        provider_id = 1  # bloomberg
        provider = get_provider(task.master_user, provider_id)
    except Exception:
        _l.debug("provider load error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if provider is None:
        _l.debug("provider not found")
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if task.status not in [CeleryTask.STATUS_PENDING, CeleryTask.STATUS_WAIT_RESPONSE]:
        _l.warning("invalid task status")
        return

    options = task.options_object

    try:
        result, is_ready = provider.download_instrument_pricing(options)
    except Exception:
        _l.warning("provider processing error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
    else:
        if is_ready:
            task.status = CeleryTask.STATUS_DONE
            task.result_object = result
        else:
            task.status = CeleryTask.STATUS_WAIT_RESPONSE

    if response_id := options.get("response_id", None):
        task.response_id = response_id
    task.options_object = options
    task.save()

    if task.status == CeleryTask.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(
                countdown=provider.get_retry_delay(),
                max_retries=provider.get_max_retries(),
            )
            # self.retry(countdown=provider.get_retry_delay(), max_retries=
            # provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = CeleryTask.STATUS_TIMEOUT
            task.save()
        return

    return task_id


@finmars_task(
    name="integrations.test_certificate_async", bind=True, ignore_result=False
)
def test_certificate_async(self, task_id, *args, **kwargs):
    task = CeleryTask.objects.get(pk=task_id)
    _l.info(
        "handle_test_certificate_async: master_user_id=%s, task=%s",
        task.master_user_id,
        task,
    )

    task.celery_task_id = self.request.id

    try:
        provider_id = 1  # bloomberg

        provider = get_provider(task.master_user, provider_id)
    except Exception:
        _l.error("provider load error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if provider is None:
        _l.error("provider not found")
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if task.status not in [CeleryTask.STATUS_PENDING, CeleryTask.STATUS_WAIT_RESPONSE]:
        _l.error("invalid task status")
        return

    options = task.options_object

    try:
        result = provider.test_certificate(options)
    except Exception:
        _l.error("provider processing error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return
    else:
        _l.info(f"handle_test_certificate_async task: result {result}")
        _l.info(
            f"handle_test_certificate_async task: result is authorized "
            f'{result["is_authorized"]}'
        )

        task.status = CeleryTask.STATUS_DONE
        task.result_object = result

        task.options_object = options
        task.save()

        try:
            import_config = BloombergDataProviderCredential.objects.get(
                master_user=task.master_user
            )

            _l.debug("handle_test_certificate_async get actual bloomberg credential")

        except (BloombergDataProviderCredential.DoesNotExist, Exception) as e:
            _l.debug("handle_test_certificate_async get config error", e)

            import_config = ImportConfig.objects.get(
                master_user=task.master_user, provider=1
            )

        import_config.is_valid = result["is_authorized"]
        import_config.save()

        _l.debug(
            "handle_test_certificate_async import_config: import_config id",
            import_config.id,
        )
        _l.debug(
            "handle_test_certificate_async import_config: import_config=%s, is_valid=%s",
            import_config,
            import_config.is_valid,
        )
        _l.debug(
            "handle_test_certificate_async task: master_user_id=%s, task=%s",
            task.master_user_id,
            task.result,
        )
        _l.debug("handle_test_certificate_async task.status: ", task.status)

    if task.status == CeleryTask.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(
                countdown=provider.get_retry_delay(),
                max_retries=provider.get_max_retries(),
            )
            # self.retry(countdown=provider.get_retry_delay(),
            # max_retries=provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = CeleryTask.STATUS_TIMEOUT
            task.save()
        return

    return task_id


# DEPRECATED SINCE 22.09.2020 DELETE SOON
@finmars_task(
    name="integrations.download_currency_pricing_async", bind=True, ignore_result=False
)
def download_currency_pricing_async(self, task_id, *args, **kwargs):
    task = CeleryTask.objects.get(pk=task_id)
    _l.debug(
        "download_currency_pricing_async: master_user_id=%s, task=%s",
        task.master_user_id,
        task,
    )

    task.celery_task_id = self.request.id

    try:
        provider_id = 1  # bloomberg

        provider = get_provider(task.master_user, provider_id)
    except Exception:
        _l.debug("provider load error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if provider is None:
        _l.debug("provider not found")
        task.status = CeleryTask.STATUS_ERROR
        task.save()
        return

    if task.status not in [CeleryTask.STATUS_PENDING, CeleryTask.STATUS_WAIT_RESPONSE]:
        _l.warning("invalid task status")
        return

    options = task.options_object

    try:
        result, is_ready = provider.download_currency_pricing(options)
    except Exception:
        _l.warning("provider processing error", exc_info=True)
        task.status = CeleryTask.STATUS_ERROR
    else:
        if is_ready:
            task.status = CeleryTask.STATUS_DONE
            task.result_object = result
        else:
            task.status = CeleryTask.STATUS_WAIT_RESPONSE

    if response_id := options.get("response_id", None):
        task.response_id = response_id

    task.options_object = options
    task.save()

    if task.status == CeleryTask.STATUS_WAIT_RESPONSE:
        if self.request.is_eager:
            time.sleep(provider.get_retry_delay())
        try:
            self.retry(
                countdown=provider.get_retry_delay(),
                max_retries=provider.get_max_retries(),
            )
            # self.retry(countdown=provider.get_retry_delay(),
            # max_retries=provider.get_max_retries(), throw=False)
        except MaxRetriesExceededError:
            task.status = CeleryTask.STATUS_TIMEOUT
            task.save()
        return

    return task_id


# DEPRECATED SINCE 22.09.2020 DELETE SOON
def _create_currency_default_prices(options, currencies, pricing_policies):
    _l.debug("create_currency_default_prices: currencies=%s", currencies)

    errors = {}
    prices = []

    date_from = parse_date_iso(options["date_from"])
    date_to = parse_date_iso(options["date_to"])

    days = (date_to - date_from).days + 1

    for c in Currency.objects.filter(pk__in=currencies):
        for pp in pricing_policies:
            for d in rrule(freq=DAILY, count=days, dtstart=date_from):
                price = CurrencyHistory(
                    currency=c,
                    pricing_policy=pp,
                    date=d.date(),
                    fx_rate=c.default_fx_rate,
                )

                prices.append(price)

    return prices, errors


# DEPRECATED SINCE 22.09.2020 DELETE SOON
def _create_instrument_default_prices(options, instruments, pricing_policies):
    _l.debug("create_instrument_default_prices: instruments=%s", instruments)

    date_from = parse_date_iso(options["date_from"])
    date_to = parse_date_iso(options["date_to"])

    errors = {}
    prices = []

    days = (date_to - date_from).days + 1

    for i in Instrument.objects.filter(pk__in=instruments):
        for pp in pricing_policies:
            for dt in rrule(freq=DAILY, count=days, dtstart=date_from):
                d = dt.date()
                price = PriceHistory(
                    instrument=i,
                    pricing_policy=pp,
                    date=d,
                    principal_price=i.default_price,
                )

                prices.append(price)

    return prices, errors


# DEPRECATED SINCE 22.09.2020 DELETE SOON
def _create_instrument_manual_prices(options, instruments):
    _l.debug("create_instrument_manual_prices: instruments=%s", instruments)

    date_from = parse_date_iso(options["date_from"])
    date_to = parse_date_iso(options["date_to"])
    is_yesterday = options["is_yesterday"]
    fill_days = options["fill_days"]

    errors = {}
    prices = []

    if is_yesterday:
        for i in Instrument.objects.filter(pk__in=instruments):
            for mf in i.manual_pricing_formulas.all():
                if mf.expr:
                    values = {"d": date_to}
                    try:
                        principal_price = formula.safe_eval(mf.expr, names=values)
                    except formula.InvalidExpression:
                        AbstractProvider.fail_manual_pricing_formula(errors, mf, values)
                        continue
                    price = PriceHistory(
                        instrument=i,
                        pricing_policy=mf.pricing_policy,
                        date=date_to,
                        principal_price=principal_price,
                    )
                    prices.append(price)
    else:
        days = (date_to - date_from).days + 1

        for i in Instrument.objects.filter(pk__in=instruments):
            safe_instrument = {
                "id": i.id,
            }
            for mf in i.manual_pricing_formulas.all():
                if mf.expr:
                    for dt in rrule(freq=DAILY, count=days, dtstart=date_from):
                        d = dt.date()
                        values = {
                            "d": d,
                            "instrument": safe_instrument,
                        }
                        try:
                            principal_price = formula.safe_eval(mf.expr, names=values)
                        except formula.InvalidExpression:
                            AbstractProvider.fail_manual_pricing_formula(
                                errors, mf, values
                            )
                            continue
                        price = PriceHistory(
                            instrument=i,
                            pricing_policy=mf.pricing_policy,
                            date=d,
                            principal_price=principal_price,
                        )
                        prices.append(price)

    return prices, errors


def test_certificate(master_user=None, member=None, task=None):
    _l.debug(
        "test_certificate: master_user_id=%s, task=%s",
        getattr(master_user, "id", None),
        getattr(task, "info", None),
    )

    try:
        if task is None:
            with transaction.atomic():
                # DEPRECATED, REFACTOR SOON
                task = CeleryTask(
                    master_user=master_user,
                    member=member,
                    status=CeleryTask.STATUS_PENDING,
                    type="test_certificate",
                )

                options = {}
                task.options_object = options
                task.save()

                transaction.on_commit(
                    lambda: test_certificate_async.apply_async(
                        kwargs={"task_id": task.id, 'context': {
                            'space_code': task.master_user.space_code,
                            'realm_code': task.master_user.realm_code
                        }}
                    )
                )

        elif task.status == CeleryTask.STATUS_DONE:
            return task, True

        return task, False
    except Exception as e:
        _l.info(f"test_certificate error {e} ")
        _l.info(traceback.print_exc())

        return task, False


def generate_file_report(result_object, master_user, scheme, type, name, context=None):
    def get_unique_columns(res_object):
        unique_columns = []

        for item in res_object["error_rows"]:
            for item_column in item["error_data"]["columns"][
                "executed_input_expressions"
            ]:
                column = (
                    item_column
                    + ":"
                    + item["error_data"]["data"]["transaction_type_selector"][0]
                )

                if column not in unique_columns:
                    unique_columns.append(column)

        return unique_columns

    def generate_columns_for_file(res_object):
        columns = ["Row number"]

        # _l.debug('res_object %s' % res_object)

        if len(res_object["error_rows"]):
            columns += res_object["error_rows"][0]["error_data"]["columns"][
                "imported_columns"
            ]
            columns = (
                columns
                + res_object["error_rows"][0]["error_data"]["columns"][
                    "converted_imported_columns"
                ]
            )
            columns = (
                columns
                + res_object["error_rows"][0]["error_data"]["columns"][
                    "calculated_columns"
                ]
            )
            columns = (
                columns
                + res_object["error_rows"][0]["error_data"]["columns"][
                    "transaction_type_selector"
                ]
            )

            unique_columns = get_unique_columns(res_object)

            for unique_column in unique_columns:
                columns.append(unique_column)

        columns.append("Error Message")
        columns.append("Reaction")

        return columns

    def generate_columns_data_for_file(instance, error_row):
        result = []
        unique_columns = get_unique_columns(instance)

        index = 0

        for unique_column in unique_columns:
            result.append("")  # result[index] = ''

            item_column_index = 0

            for item_column in error_row["error_data"]["columns"][
                "executed_input_expressions"
            ]:
                column = (
                    item_column
                    + ":"
                    + error_row["error_data"]["data"]["transaction_type_selector"][0]
                )

                if (
                    column == unique_column
                    and error_row["error_data"]["data"]["executed_input_expressions"][
                        item_column_index
                    ]
                ):
                    result[index] = error_row["error_data"]["data"][
                        "executed_input_expressions"
                    ][item_column_index]

                item_column_index = item_column_index + 1

            index = index + 1

        return result

    _l.info(f"generate_file_report error_handler {scheme.error_handler}")
    _l.info(f"generate_file_report missing_data_handler {scheme.missing_data_handler}")

    result = []
    error_rows = []

    for item in result_object["error_rows"]:
        if item["level"] == "error":
            error_rows.append(item)

    result.append(f"Type, {name}")
    result.append(f"Scheme, {scheme.user_code}")
    result.append(f"Error handle, {scheme.error_handler}")
    # result.append('Filename, ' + instance.file.name)
    result.append(
        f"Import Rules - if object is not found, {scheme.missing_data_handler}"
    )

    rowsSuccessCount = 0

    if scheme.error_handler == "break":
        if "error_row_index" in result_object and result_object["error_row_index"]:
            rowsSuccessCount = result_object["error_row_index"] - 1
        else:
            rowsSuccessCount = result_object["total_rows"] - len(error_rows)
    else:
        rowsSuccessCount = result_object["total_rows"] - len(error_rows)

    result.append("Rows total, " + str(result_object["total_rows"]))
    result.append(f"Rows success import, {rowsSuccessCount}")
    result.append(f"Rows fail import, {len(error_rows)}")

    columns = generate_columns_for_file(result_object)

    column_row_list = []

    for item in columns:
        column_row_list.append(f'"{str(item)}"')

    column_row = ",".join(column_row_list)

    result.append(column_row)

    for error_row in result_object["error_rows"]:
        content = [error_row["original_row_index"]]
        content += error_row["error_data"]["data"]["imported_columns"]
        content = (
            content + error_row["error_data"]["data"]["converted_imported_columns"]
        )
        content = content + error_row["error_data"]["data"]["calculated_columns"]
        content = content + error_row["error_data"]["data"]["transaction_type_selector"]
        content = content + generate_columns_data_for_file(result_object, error_row)

        content.append(error_row["error_message"])
        content.append(error_row["error_reaction"])

        content_row_list = []

        for item in content:
            content_row_list.append(f'"{str(item)}"')

        content_row = ",".join(content_row_list)

        result.append(content_row)

    result = "\n".join(result)

    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

    file_name = f"file_report_{current_date_time}.csv"

    file_report = FileReport()

    _l.debug("generate_file_report uploading file ")

    file_report.upload_file(file_name=file_name, text=result, master_user=master_user)
    file_report.master_user = master_user
    file_report.name = f"{name} {current_date_time}.csv"
    file_report.file_name = file_name
    file_report.type = type
    file_report.notes = "System File"
    file_report.content_type = "text/csv"

    file_report.save()

    _l.debug(f"file_report {file_report} {file_report.file_url}")

    return file_report.pk


def generate_file_report_old(instance, master_user, type, name, context=None):
    def get_unique_columns(instance):
        unique_columns = []

        for item in instance.error_rows:
            for item_column in item["error_data"]["columns"][
                "executed_input_expressions"
            ]:
                column = (
                    item_column
                    + ":"
                    + item["error_data"]["data"]["transaction_type_selector"][0]
                )

                if column not in unique_columns:
                    unique_columns.append(column)

        return unique_columns

    def generate_columns_for_file(instance):
        columns = ["Row number"]

        _l.debug(f"instance {instance}")

        if len(instance.error_rows):
            columns += instance.error_rows[0]["error_data"]["columns"][
                "imported_columns"
            ]
            columns = (
                columns
                + instance.error_rows[0]["error_data"]["columns"][
                    "converted_imported_columns"
                ]
            )
            columns = (
                columns
                + instance.error_rows[0]["error_data"]["columns"]["calculated_columns"]
            )
            columns = (
                columns
                + instance.error_rows[0]["error_data"]["columns"][
                    "transaction_type_selector"
                ]
            )

            unique_columns = get_unique_columns(instance)

            for unique_column in unique_columns:
                columns.append(unique_column)

        columns.append("Error Message")
        columns.append("Reaction")

        return columns

    def generate_columns_data_for_file(instance, error_row):
        result = []
        unique_columns = get_unique_columns(instance)

        index = 0

        for unique_column in unique_columns:
            result.append("")  # result[index] = ''

            item_column_index = 0

            for item_column in error_row["error_data"]["columns"][
                "executed_input_expressions"
            ]:
                column = (
                    item_column
                    + ":"
                    + error_row["error_data"]["data"]["transaction_type_selector"][0]
                )

                if (
                    column == unique_column
                    and error_row["error_data"]["data"]["executed_input_expressions"][
                        item_column_index
                    ]
                ):
                    result[index] = error_row["error_data"]["data"][
                        "executed_input_expressions"
                    ][item_column_index]

                item_column_index = item_column_index + 1

            index = index + 1

        return result

    result = []
    error_rows = []

    for item in instance.error_rows:
        if item["level"] == "error":
            error_rows.append(item)

    result.append("Type, Transaction Import")
    result.append(f"Error handle, {instance.error_handling}")
    # result.append('Filename, ' + instance.file.name)
    result.append(
        f"Import Rules - if object is not found, {instance.missing_data_handler}"
    )

    if instance.error_handling == "break":
        if instance.error_row_index:
            rowsSuccessCount = instance.error_row_index - 1
        else:
            rowsSuccessCount = instance.total_rows - len(error_rows)
    else:
        rowsSuccessCount = instance.total_rows - len(error_rows)

    result.append(f"Rows total, {str(instance.total_rows)}")
    result.append(f"Rows success import, {rowsSuccessCount}")
    result.append(f"Rows fail import, {len(error_rows)}")

    columns = generate_columns_for_file(instance)

    column_row_list = []

    for item in columns:
        column_row_list.append(f'"{str(item)}"')

    column_row = ",".join(column_row_list)

    result.append(column_row)

    for error_row in instance.error_rows:
        content = [error_row["original_row_index"]]

        content += error_row["error_data"]["data"]["imported_columns"]
        content = (
            content + error_row["error_data"]["data"]["converted_imported_columns"]
        )
        content = content + error_row["error_data"]["data"]["calculated_columns"]
        content = content + error_row["error_data"]["data"]["transaction_type_selector"]
        content = content + generate_columns_data_for_file(instance, error_row)

        content.append(error_row["error_message"])
        content.append(error_row["error_reaction"])

        content_row_list = []

        for item in content:
            content_row_list.append(f'"{str(item)}"')

        content_row = ",".join(content_row_list)

        result.append(content_row)

    result = "\n".join(result)

    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

    file_name = f"file_report_{current_date_time}.csv"

    file_report = FileReport()

    _l.debug("generate_file_report uploading file ")

    file_report.upload_file(file_name=file_name, text=result, master_user=master_user)
    file_report.master_user = master_user
    file_report.name = f"{name} {current_date_time}"
    file_report.file_name = file_name
    file_report.type = type
    file_report.notes = "System File"
    file_report.content_type = "text/csv"

    file_report.save()

    _l.debug(f"file_report {file_report}")
    _l.debug(f"file_report {file_report.file_url}")

    return file_report.pk


@finmars_task(
    name="integrations.complex_transaction_csv_file_import_parallel_finish", bind=True
)
def complex_transaction_csv_file_import_parallel_finish(self, task_id, *args, **kwargs):
    try:
        _l.info(
            f"complex_transaction_csv_file_import_parallel_finish task_id {task_id} "
        )

        celery_task = CeleryTask.objects.get(pk=task_id)
        celery_task.celery_task_id = self.requst.id
        celery_task.save()

        scheme = ComplexTransactionImportScheme.objects.get(
            pk=celery_task.options_object["scheme_id"]
        )

        master_user = celery_task.master_user
        member = celery_task.member

        result_object = {
            "error_rows": [],
            "total_rows": celery_task.options_object["total_rows"],
            "processed_rows": 0,
        }

        _l.info(
            f"complex_transaction_csv_file_import_parallel_finish iterating over "
            f"{len(celery_task.children.all())} childs"
        )

        for sub_task in celery_task.children.all():
            if sub_task.result_object:
                if "error_rows" in sub_task.result_object:
                    result_object["error_rows"] = (
                        result_object["error_rows"]
                        + sub_task.result_object["error_rows"]
                    )

                if "processed_rows" in sub_task.result_object:
                    result_object["processed_rows"] = (
                        result_object["processed_rows"]
                        + sub_task.result_object["processed_rows"]
                    )

        result_object["stats_file_report"] = generate_file_report(
            result_object,
            master_user,
            scheme,
            "transaction_import.import",
            "Transaction Import",
            celery_task.options_object["execution_context"],
        )

        if (
            celery_task.options_object["execution_context"]
            and celery_task.options_object["execution_context"]["started_by"]
            == "procedure"
        ):
            _l.info(
                "complex_transaction_csv_file_import_parallel_finish send "
                "final import message"
            )

            send_system_message(
                master_user=celery_task.master_user,
                performed_by="System",
                description="Import Finished",
                attachments=[result_object["stats_file_report"]],
            )
        else:
            send_system_message(
                master_user=celery_task.master_user,
                performed_by="System",
                description=f"User {celery_task.member.username} Transaction Import Finished",
                attachments=[result_object["stats_file_report"]],
            )

        celery_task.result_object = result_object

        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

    except Exception as e:
        _l.info(f"Exception occurred {e}")
        _l.info(traceback.format_exc())


# DEPRECATED
@finmars_task(name="integrations.complex_transaction_csv_file_import", bind=True)
def complex_transaction_csv_file_import(self, task_id, procedure_instance_id=None, *args, **kwargs):
    try:
        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.transactions.models import TransactionTypeInput

        celery_task = CeleryTask.objects.get(pk=task_id)
        parent_celery_task = celery_task.parent

        celery_task.celery_task_id = self.request.id
        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        procedure_instance = None
        if procedure_instance_id:
            procedure_instance = RequestDataFileProcedureInstance.objects.get(
                id=procedure_instance_id
            )

        master_user = celery_task.master_user
        member = celery_task.member

        proxy_user = ProxyUser(member, master_user)
        proxy_request = ProxyRequest(proxy_user)

        instance = ComplexTransactionCsvFileImport(
            task_id=task_id,
            master_user=master_user,
            member=member,
            skip_first_line=True,
        )

        scheme = ComplexTransactionImportScheme.objects.get(
            pk=celery_task.options_object["scheme_id"]
        )

        instance.scheme = scheme
        instance.error_handling = scheme.error_handler
        instance.delimiter = scheme.delimiter
        instance.missing_data_handler = scheme.missing_data_handler
        instance.file_path = celery_task.options_object["file_path"]
        execution_context = celery_task.options_object["execution_context"]

        instance.processed_rows = 0

        scheme = instance.scheme
        scheme_inputs = list(scheme.inputs.all())
        scheme_calculated_inputs = list(scheme.calculated_inputs.all())
        scheme_selector_values = list(scheme.selector_values.all())

        master_user = instance.master_user
        member = instance.member

        rule_scenarios = scheme.rule_scenarios.prefetch_related(
            "transaction_type", "fields", "fields__transaction_type_input"
        ).all()

        default_rule_scenario = None

        for scenario in rule_scenarios:
            if scenario.is_default_rule_scenario:
                default_rule_scenario = scenario

        mapping_map = {
            Account: AccountMapping,
            Currency: CurrencyMapping,
            Instrument: InstrumentMapping,
            InstrumentType: InstrumentTypeMapping,
            Counterparty: CounterpartyMapping,
            Responsible: ResponsibleMapping,
            Strategy1: Strategy1Mapping,
            Strategy2: Strategy2Mapping,
            Strategy3: Strategy3Mapping,
            DailyPricingModel: DailyPricingModelMapping,
            PaymentSizeDetail: PaymentSizeDetailMapping,
            Portfolio: PortfolioMapping,
            PriceDownloadScheme: PriceDownloadSchemeMapping,
            Periodicity: PeriodicityMapping,
            AccrualCalculationModel: AccrualCalculationModelMapping,
        }

        props_map = {
            Account: "account",
            Currency: "currency",
            Instrument: "instrument",
            InstrumentType: "instrument_type",
            Counterparty: "counterparty",
            Responsible: "responsible",
            Strategy1: "strategy1",
            Strategy2: "strategy2",
            Strategy3: "strategy3",
            DailyPricingModel: "daily_pricing_model",
            PaymentSizeDetail: "payment_size_detail",
            Portfolio: "portfolio",
            PriceDownloadScheme: "price_download_scheme",
            Periodicity: "periodicity",
            AccrualCalculationModel: "accrual_calculation_model",
        }

        def _get_default_relation(field):
            i = field.transaction_type_input

            model_class = i.content_type.model_class()
            model_map_class = mapping_map[model_class]

            key = props_map[model_class]

            v = None

            ecosystem_default = EcosystemDefault.objects.get(
                master_user=instance.master_user
            )

            # _l.info('key %s' % key)
            # _l.info('value %s' % value)

            if hasattr(ecosystem_default, key):
                v = getattr(ecosystem_default, key)
            else:
                v = model_map_class.objects.get(
                    master_user=instance.master_user, value="-"
                ).content_object

            return v

        def _convert_value(field, value, error_rows):
            i = field.transaction_type_input

            if i.value_type == TransactionTypeInput.STRING:
                return str(value)

            if i.value_type == TransactionTypeInput.SELECTOR:
                return str(value)

            elif i.value_type == TransactionTypeInput.NUMBER:
                return float(value)

            elif i.value_type == TransactionTypeInput.DATE:
                if not isinstance(value, date):
                    return formula._parse_date(value)
                else:
                    return value

            elif i.value_type == TransactionTypeInput.RELATION:
                model_class = i.content_type.model_class()
                model_map_class = mapping_map[model_class]

                v = None

                try:
                    v = model_class.objects.get(
                        master_user=instance.master_user, user_code=value
                    )

                except Exception:
                    try:
                        v = model_map_class.objects.get(
                            master_user=instance.master_user, value=value
                        ).content_object

                    except Exception:
                        v = None

                        _l.info(
                            "User code %s not found for %s "
                            % (value, field.transaction_type_input.name)
                        )

                if not v:
                    if instance.missing_data_handler == "set_defaults":
                        v = _get_default_relation(field)

                    else:
                        error_rows["error_message"] = (
                            error_rows["error_message"]
                            + " Can't find relation of "
                            + "["
                            + field.transaction_type_input.name
                            + "]"
                            + "(value:"
                            + value
                            + ")"
                        )

                return v

        def update_row_with_calculated_data(row, inputs, error_rows):
            for i in scheme_calculated_inputs:
                try:
                    value = formula.safe_eval(
                        i.name_expr,
                        names=inputs,
                        context={"master_user": master_user, "member": member},
                    )
                    inputs[i.name] = value

                except Exception:
                    msg = "can't process calculated input: %s|%s", i.name, i.column
                    _l.debug(msg, exc_info=True)
                    error_rows["error_message"] = error_rows["error_message"] + str(msg)
                    # row.append("Invalid Expression")

            # return row

        def _process_rule_scenario(
            processed_scenarios, scheme_rule, inputs, error_rows, row_index
        ):
            _l.info("_process_rule_scenario %s %s " % (row_index, scheme_rule))

            result = None

            processed_scenarios = processed_scenarios + 1

            rule = scheme_rule

            fields = {}
            fields_error = []
            for field in rule.fields.all():
                try:
                    field_value = formula.safe_eval(
                        field.value_expr,
                        names=inputs,
                        context={"master_user": master_user, "member": member},
                    )
                    field_value = _convert_value(field, field_value, error_rows)
                    fields[field.transaction_type_input.name] = field_value

                except (Exception, ValueError, ExpressionEvalError):
                    _l.debug(
                        "can't process field: %s|%s",
                        field.transaction_type_input.name,
                        field.transaction_type_input.pk,
                    )
                    fields_error.append(field)

            _l.debug("fields (step 1): error=%s, values=%s", fields_error, fields)

            if fields_error:
                error_rows["level"] = "error"

                error_rows["error_message"] = (
                    error_rows["error_message"]
                    + "\n"
                    + str(
                        gettext_lazy("Can't process fields: %(fields)s")
                        % {
                            "fields": ", ".join(
                                "[" + f.transaction_type_input.name + "] "
                                for f in fields_error
                            )
                        }
                    )
                )

                if instance.break_on_error:
                    instance.error_row_index = row_index
                    error_rows["error_reaction"] = "Break"
                    instance.error_rows.append(error_rows)

                    result = "break"

                    return result, processed_scenarios
                else:
                    error_rows["error_reaction"] = "Continue import"

                    result = "continue"

                    return result, processed_scenarios

            with transaction.atomic():
                try:
                    tt_process = TransactionTypeProcess(
                        transaction_type=rule.transaction_type,
                        default_values=fields,
                        context={
                            "master_user": instance.master_user,
                            "member": instance.member,
                            "request": proxy_request,
                        },
                        uniqueness_reaction=instance.scheme.book_uniqueness_settings,
                        member=instance.member,
                        linked_import_task=celery_task,
                    )
                    tt_process.process()

                    _l.debug("tt_process %s" % tt_process)

                    if tt_process.uniqueness_status == "skip":
                        error_rows["level"] = "skip"
                        error_rows["error_message"] = error_rows["error_message"] + str(
                            gettext_lazy("Unique code already exist. Skip")
                        )

                    if tt_process.uniqueness_status == "error":
                        error_rows["level"] = "error"
                        error_rows["error_message"] = error_rows["error_message"] + str(
                            gettext_lazy("Unique code already exist. Error")
                        )

                    processed_scenarios = processed_scenarios + 1

                except Exception as e:
                    error_rows["level"] = "error"

                    _l.debug("can't process transaction type", exc_info=True)

                    _l.debug("error %s" % e)

                    transaction.set_rollback(True)
                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows["error_reaction"] = "Break"
                        result = "break"
                        return result, processed_scenarios
                    else:
                        result = "continue"
                        return result, processed_scenarios
                finally:
                    _l.debug("final")

            return result, processed_scenarios

        def _process_csv_file(file, original_file, original_file_path):
            _l.info("_process_csv_file %s " % instance.file_path)

            instance.processed_rows = 0

            reader = []

            if ".csv" in instance.file_path or (
                execution_context and execution_context["started_by"] == "procedure"
            ):
                delimiter = instance.delimiter.encode("utf-8").decode("unicode_escape")

                reader = csv.reader(
                    file,
                    delimiter=delimiter,
                    quotechar=instance.quotechar,
                    strict=False,
                    skipinitialspace=True,
                )

            elif ".xlsx" in instance.file_path:
                _l.info("trying to parse excel %s " % instance.file_path)

                wb = load_workbook(filename=original_file_path)

                if (
                    instance.scheme.spreadsheet_active_tab_name
                    and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames
                ):
                    ws = wb[instance.scheme.spreadsheet_active_tab_name]
                else:
                    ws = wb.active

                _l.info("ws %s" % ws)
                _l.info(
                    "task_instance.scheme.spreadsheet_start_cell %s"
                    % instance.scheme.spreadsheet_start_cell
                )

                reader = []

                if instance.scheme.spreadsheet_start_cell == "A1":
                    for r in ws.rows:
                        reader.append([cell.value for cell in r])

                else:
                    start_cell_row_number = int(
                        re.search(r"\d+", instance.scheme.spreadsheet_start_cell)[0]
                    )
                    start_cell_letter = instance.scheme.spreadsheet_start_cell.split(
                        str(start_cell_row_number)
                    )[0]

                    start_cell_column_number = column_index_from_string(
                        start_cell_letter
                    )

                    row_number = 1

                    for r in ws.rows:
                        row_values = []

                        if row_number >= start_cell_row_number:
                            for cell in r:
                                if cell.column >= start_cell_column_number:
                                    row_values.append(cell.value)

                            reader.append(row_values)

                        row_number = row_number + 1

            _process_list_of_items(reader)

        def _process_list_of_items(items):
            input_column_name_map = {}

            for row_index, row in enumerate(items):
                _l.info("process row_index %s " % row_index)
                _l.info("process row %s " % row)

                if row_index == 0:
                    first_row = row

                    _local_index = 0
                    for item in first_row:
                        input_column_name_map[item] = _local_index
                        _local_index = _local_index + 1

                # _l.debug('process row: %s -> %s', row_index, row)
                if (
                    row_index == 0
                    and instance.skip_first_line
                    and not scheme.has_header_row
                ) or not row:
                    _l.debug("skip first row")
                    continue

                inputs_raw = {}
                inputs = {}
                inputs_error = []
                inputs_conversion_error = []
                calculated_columns_error = []

                error_rows = {
                    "level": "info",
                    "error_message": "",
                    "inputs": inputs_raw,
                    "original_row_index": row_index,
                    "original_row": row,
                    "error_data": {
                        "columns": {
                            "imported_columns": [],
                            "calculated_columns": [],
                            "converted_imported_columns": [],
                            "transaction_type_selector": [],
                            "executed_input_expressions": [],
                        },
                        "data": {
                            "imported_columns": [],
                            "calculated_columns": [],
                            "converted_imported_columns": [],
                            "transaction_type_selector": [],
                            "executed_input_expressions": [],
                        },
                    },
                    "error_reaction": "Success",
                }

                for i in scheme_inputs:
                    error_rows["error_data"]["columns"]["imported_columns"].append(
                        i.name
                    )

                    if instance.scheme.column_matcher == "index":
                        try:
                            inputs_raw[i.name] = row[i.column - 1]
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                row[i.column - 1]
                            )
                        except Exception:
                            _l.debug(
                                "can't process input: %s|%s",
                                i.name,
                                i.column,
                                exc_info=True,
                            )
                            _l.debug("can't process inputs_raw: %s|%s", inputs_raw)
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                gettext_lazy("Invalid expression")
                            )
                            inputs_error.append(i)

                    if instance.scheme.column_matcher == "name":
                        try:
                            if type(row) is dict:
                                if i.column_name in row:
                                    inputs_raw[i.name] = row[i.column_name]
                                else:
                                    inputs_raw[i.name] = None
                            else:
                                _col_index = input_column_name_map[i.name]
                                inputs_raw[i.name] = row[_col_index]
                                error_rows["error_data"]["data"][
                                    "imported_columns"
                                ].append(row[_col_index])
                        except Exception:
                            _l.debug(
                                "can't process input: %s|%s",
                                i.name,
                                i.column,
                                exc_info=True,
                            )
                            _l.debug("can't process inputs_raw: %s|%s", inputs_raw)
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                gettext_lazy("Invalid expression")
                            )
                            inputs_error.append(i)

                # _l.debug('Row %s inputs_raw: %s' % (row_index, inputs_raw))

                if scheme.filter_expression:
                    # expr = Expression.parseString("a == 1 and b == 2")
                    expr = Expression.parseString(scheme.filter_expression)

                    if expr(inputs_raw):
                        # filter passed
                        pass
                    else:
                        _l.info("Row skipped due filter %s" % row_index)
                        continue

                original_columns_count = len(row)

                if inputs_error:
                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        gettext_lazy("Can't process fields: %(inputs)s")
                        % {
                            "inputs": ", ".join(
                                "[" + i.name + "] (Can't find input)"
                                for i in inputs_error
                            )
                        }
                    )
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        error_rows["error_reaction"] = "Break"
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"
                        continue

                for i in scheme_inputs:
                    error_rows["error_data"]["columns"][
                        "converted_imported_columns"
                    ].append(
                        i.name + ": Conversion Expression " + "(" + i.name_expr + ")"
                    )

                    try:
                        inputs[i.name] = formula.safe_eval(
                            i.name_expr,
                            names=inputs_raw,
                            context={"master_user": master_user, "member": member},
                        )
                        error_rows["error_data"]["data"][
                            "converted_imported_columns"
                        ].append(inputs_raw[i.name])
                    except Exception:
                        _l.debug(
                            "can't process conversion input: %s|%s",
                            i.name,
                            i.column,
                            exc_info=True,
                        )
                        error_rows["error_data"]["data"][
                            "converted_imported_columns"
                        ].append(gettext_lazy("Invalid expression"))
                        inputs_conversion_error.append(i)

                # _l.debug('Row %s inputs_conversion: %s' % (row_index, inputs))

                if inputs_conversion_error:
                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        gettext_lazy("Can't process fields: %(inputs)s")
                        % {
                            "inputs": ", ".join(
                                "["
                                + str(i.name)
                                + '] (Imported column conversion expression, value; "'
                                + str(i.name_expr)
                                + '")'
                                for i in inputs_conversion_error
                            )
                        }
                    )
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        error_rows["error_reaction"] = "Break"
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"
                        continue

                update_row_with_calculated_data(row, inputs, error_rows)

                # for i in scheme_calculated_inputs:
                #
                #     error_rows['error_data']['columns']['calculated_columns'].append(i.name)
                #
                #     try:
                #
                #         index = original_columns_count + i.column - 1
                #
                #         # print('index %s ' % index)
                #         # print('i.name %s ' % i.name)
                #
                #         inputs[i.name] = row[index]
                #
                #         error_rows['error_data']['data']['calculated_columns'].append(row[index])
                #     except Exception:
                #         _l.debug('can\'t process input: %s|%s', i.name, i.column, exc_info=True)
                #         error_rows['error_data']['data']['calculated_columns'].append(gettext_lazy('Invalid expression'))
                #         calculated_columns_error.append(i)

                # _l.debug('Row %s inputs_with_calculated: %s' % (row_index, inputs))

                try:
                    rule_value = formula.safe_eval(
                        scheme.rule_expr,
                        names=inputs,
                        context={"master_user": master_user, "member": member},
                    )
                except Exception:
                    error_rows["level"] = "error"

                    _l.debug("can't process rule expression", exc_info=True)
                    error_rows["error_message"] = (
                        error_rows["error_message"]
                        + "\n"
                        + str(gettext_lazy("Can't eval rule expression"))
                    )
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        continue

                if not rule_value:
                    _l.debug("no rule value: %s", rule_value)

                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        gettext_lazy("Rule expression is invalid")
                    )

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows["error_reaction"] = "Break"
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"
                        continue

                # else:
                #     _l.debug('rule value: %s', rule_value)

                matched_selector = False
                processed_scenarios = 0

                unknown_rule = True

                for selector in scheme_selector_values:
                    if selector.value == rule_value:
                        unknown_rule = False

                if unknown_rule and default_rule_scenario:
                    _l.info(
                        "Process rule %s with default rule scenario " % (rule_value)
                    )
                    res, processed_scenarios = _process_rule_scenario(
                        processed_scenarios,
                        default_rule_scenario,
                        inputs,
                        error_rows,
                        row_index,
                    )

                    # TODO refactor soon
                    if res == "break":
                        return
                    elif res == "continue":
                        continue

                else:
                    for scheme_rule in rule_scenarios:
                        matched_selector = False

                        selector_values = scheme_rule.selector_values.all()

                        for selector_value in selector_values:
                            if selector_value.value == rule_value:
                                matched_selector = True

                        if matched_selector:
                            res, processed_scenarios = _process_rule_scenario(
                                processed_scenarios,
                                scheme_rule,
                                inputs,
                                error_rows,
                                row_index,
                            )

                            # TODO refactor soon
                            if res == "break":
                                return
                            elif res == "continue":
                                continue

                if processed_scenarios == 0:
                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        "Selector does not match"
                    )

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows["error_reaction"] = "Break"
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"

                instance.error_rows.append(error_rows)

                instance.processed_rows = instance.processed_rows + 1

                total_rows = 0

                if parent_celery_task:
                    total_rows = (parent_celery_task.options_object["total_rows"],)
                else:
                    total_rows = instance.total_rows

                # DEPRECATED
                # send_websocket_message(data={
                #     'type': 'transaction_import_status',
                #     'payload': {
                #         'parent_task_id': celery_task.parent_id,
                #         'task_id': instance.task_id,
                #         'state': Task.STATUS_PENDING,
                #         'processed_rows': instance.processed_rows,
                #         'parent_total_rows': total_rows,
                #         'total_rows': instance.total_rows,
                #         'scheme_name': scheme.user_code,
                #         'file_name': instance.file_name}
                # }, level="member",
                #     context={"master_user": master_user, "member": member})

        def _row_count_csv(file):
            delimiter = instance.delimiter.encode("utf-8").decode("unicode_escape")

            reader = csv.reader(
                file,
                delimiter=delimiter,
                quotechar=instance.quotechar,
                strict=False,
                skipinitialspace=True,
            )

            row_index = 0

            for row_index, row in enumerate(reader):
                pass

            _l.info("Total rows in file: %s" % row_index)

            return row_index

        def _row_count_xlsx(filename):
            wb = load_workbook(filename=filename)

            if (
                instance.scheme.spreadsheet_active_tab_name
                and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames
            ):
                ws = wb[instance.scheme.spreadsheet_active_tab_name]
            else:
                ws = wb.active

            reader = []

            row_index = 0

            for r in ws.rows:
                row_index = row_index + 1

            return row_index

        instance.error_rows = []

        try:
            if celery_task.options_object and "items" in celery_task.options_object:
                _l.info("Parse json data")

                items = celery_task.options_object["items"]

                instance.total_rows = len(items)

                _process_list_of_items(items)

            else:
                _l.info("Open file %s" % instance.file_path)
                # with import_file_storage.open(instance.file_path, 'rb') as f:
                with storage.open(instance.file_path, "rb") as f:
                    with NamedTemporaryFile() as tmpf:
                        for chunk in f.chunks():
                            tmpf.write(chunk)
                        tmpf.flush()

                        os.link(tmpf.name, tmpf.name + ".xlsx")

                        if ".csv" in instance.file_path or (
                            execution_context
                            and execution_context["started_by"] == "procedure"
                        ):
                            with open(
                                tmpf.name,
                                mode="rt",
                                encoding=instance.encoding,
                                errors="ignore",
                            ) as cfr:
                                instance.total_rows = _row_count_csv(cfr)

                            with open(
                                tmpf.name,
                                mode="rt",
                                encoding=instance.encoding,
                                errors="ignore",
                            ) as cf:
                                _process_csv_file(cf, f, "")

                        elif ".xlsx" in instance.file_path:
                            instance.total_rows = _row_count_xlsx(tmpf.name + ".xlsx")

                            with open(
                                tmpf.name,
                                mode="rt",
                                encoding=instance.encoding,
                                errors="ignore",
                            ) as cf:
                                _process_csv_file(cf, f, tmpf.name + ".xlsx")

        except Exception:
            _l.debug("Can't process file", exc_info=True)
            instance.error_message = gettext_lazy(
                "Invalid file format or file already deleted."
            )

            if execution_context and execution_context["started_by"] == "procedure":
                send_system_message(
                    master_user=instance.master_user,
                    performed_by="System",
                    description="Can't process file. Possibly wrong format",
                )

        finally:
            # import_file_storage.delete(instance.file_path)

            if celery_task.options_object and "items" in celery_task.options_object:
                pass
            else:
                storage.delete(instance.file_path)

            instance.error = (
                bool(instance.error_message)
                or (instance.error_row_index is not None)
                or bool(instance.error_rows)
            )

            # instance.stats_file_report = generate_file_report(instance, master_user, 'transaction_import.import',
            #                                                   'Transaction Import', execution_context)

            _l.debug(
                "complex_transaction_file_import execution_context: %s",
                execution_context,
            )

            # _l.debug("Reached end instance.stats_file_report: %s " % instance.stats_file_report)

            # if execution_context and execution_context["started_by"] == 'procedure':
            #
            #     _l.debug('send final import message')
            #
            #     send_system_message(master_user=instance.master_user,
            #                         source="Transaction Import Service",
            #                         text="Import Finished",
            #                         file_report_id=instance.stats_file_report)

            total_rows = 0

            if parent_celery_task:
                total_rows = (parent_celery_task.options_object["total_rows"],)
            else:
                total_rows = instance.total_rows

            # DEPRECATED
            # send_websocket_message(data={
            #     'type': 'transaction_import_status',
            #     'payload': {
            #         'parent_task_id': celery_task.parent_id,
            #         'task_id': instance.task_id,
            #         'state': Task.STATUS_DONE,
            #         'processed_rows': instance.processed_rows,
            #         'parent_total_rows': total_rows,
            #         'total_rows': instance.total_rows,
            #         'file_name': instance.file_name,
            #         'error_rows': instance.error_rows,
            #         'stats_file_report': instance.stats_file_report,
            #         'scheme': scheme.id,
            #         'scheme_object': {
            #             'id': scheme.id,
            #             'scheme_name': scheme.user_code,
            #             'delimiter': scheme.delimiter,
            #             'error_handler': scheme.error_handler,
            #             'missing_data_handler': scheme.missing_data_handler
            #         }}
            # }, level="member",
            #     context={"master_user": master_user, "member": member})

            result_object = {
                "processed_rows": instance.processed_rows,
                "total_rows": instance.total_rows,
                "error_row_index": instance.error_row_index,
                "file_name": instance.file_name,
                "error_rows": instance.error_rows,
                "stats_file_report": instance.stats_file_report,
            }

            celery_task.result_object = result_object
            celery_task.status = CeleryTask.STATUS_DONE
            celery_task.save()

            # if JSON IMPORT
            if celery_task.options_object and "items" in celery_task.options_object:
                result_object["stats_file_report"] = generate_file_report(
                    result_object,
                    master_user,
                    scheme,
                    "transaction_import.import",
                    "Transaction Import",
                    celery_task.options_object["execution_context"],
                )

                if (
                    celery_task.options_object["execution_context"]
                    and celery_task.options_object["execution_context"]["started_by"]
                    == "procedure"
                ):
                    # from poms.portfolios.tasks import calculate_portfolio_register_record, \
                    #     calculate_portfolio_register_price_history
                    _l.info(
                        "complex_transaction_csv_file_import_parallel_finish send final import message"
                    )

                    send_system_message(
                        master_user=celery_task.master_user,
                        performed_by="System",
                        title="Import Finished. Prices Recalculation Required",
                        description="Please, run schedule or execute procedures to calculate portfolio prices and nav history",
                    )

                    send_system_message(
                        master_user=celery_task.master_user,
                        performed_by="System",
                        description="Import Finished",
                        attachments=[result_object["stats_file_report"]],
                    )

                    # if celery_task.options_object['execution_context']['date_from']:
                    #     calculate_portfolio_register_record.apply_async(link=[
                    #         calculate_portfolio_register_price_history.s(
                    #             date_from=celery_task.options_object['execution_context']['date_from'])
                    #     ])

                celery_task.result_object = result_object

                celery_task.status = CeleryTask.STATUS_DONE
                celery_task.save()

        if procedure_instance and procedure_instance.schedule_instance:
            procedure_instance.schedule_instance.run_next_procedure()

        return instance
    except Exception as e:
        _l.info("Exception occurred %s" % e)
        _l.info(traceback.format_exc())


# DEPRECATED
# @finmars_task(name='integrations.complex_transaction_csv_file_import_parallel', bind=True)
def complex_transaction_csv_file_import_parallel(task_id, *args, **kwargs):
    try:
        _l.info("complex_transaction_csv_file_import_parallel: task_id %s" % task_id)

        celery_task = CeleryTask.objects.get(pk=task_id)

        # celery_task.celery_task_id = self.request.id
        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        options_object = celery_task.options_object
        sub_tasks = []
        celery_sub_tasks = []

        if "items" in options_object:
            sub_task = CeleryTask.objects.create(
                master_user=celery_task.master_user,
                member=celery_task.member,
                parent=celery_task,
            )

            sub_task_options_object = copy.deepcopy(celery_task.options_object)

            sub_task.options_object = sub_task_options_object
            sub_task.save()

            sub_tasks.append(sub_task)

            for sub_task in sub_tasks:
                ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
                celery_sub_tasks.append(ct)

            transaction.on_commit(
                lambda: chord(
                    celery_sub_tasks,
                    complex_transaction_csv_file_import_parallel_finish.si(
                        task_id=task_id
                    ),
                ).apply_async()
            )

        else:
            sub_tasks = []

            lines_per_file = 300
            header_line = None

            def _get_path(master_user, file_name, ext):
                return "%s/public/%s.%s" % (master_user.space_code, file_name, ext)

            chunk = None

            with storage.open(celery_task.options_object["file_path"], "rb") as f:
                _l.info("Start reading file to split it into chunks")

                ext = celery_task.options_object["file_path"].split(".")[-1]

                for lineno, line in enumerate(f):
                    # _l.info('line %s' % lineno)

                    if lineno == 0:
                        header_line = line
                        # _l.info('set header line %s' % lineno)

                    if lineno % lines_per_file == 0:
                        if chunk is not None:
                            # _l.info("Saving chunk %s" % chunk)
                            storage.save(
                                chunk_path, chunk
                            )  # save working chunk before creating new one

                        chunk_filename = "%s_chunk_file_%s" % (
                            celery_task.id,
                            str(lineno) + "_" + str(lineno + lines_per_file),
                        )
                        chunk_path = _get_path(
                            celery_task.master_user, chunk_filename, ext
                        )

                        # _l.info('creating chunk file %s' % chunk_path)

                        chunk = BytesIO()
                        if lineno != 0:
                            chunk.write(header_line)

                        _l.info("creating chunk %s" % chunk_filename)

                        # _l.info('creating sub task for %s' % chunk_filename)

                        sub_task = CeleryTask.objects.create(
                            master_user=celery_task.master_user,
                            member=celery_task.member,
                            parent=celery_task,
                        )

                        sub_task_options_object = copy.deepcopy(
                            celery_task.options_object
                        )
                        sub_task_options_object["file_path"] = chunk_path

                        sub_task.options_object = sub_task_options_object
                        sub_task.save()

                        sub_tasks.append(sub_task)

                    chunk.write(line)

                _l.info("Saving last chunk")
                storage.save(
                    chunk_path, chunk
                )  # save working chunk before creating new one

            _l.info("sub_tasks created %s" % len(sub_tasks))
            _l.info("original file total rows %s" % lineno)

            options_object["total_rows"] = lineno

            celery_task.options_object = options_object
            celery_task.save()

            celery_sub_tasks = []

            # for sub_task in sub_tasks:
            #
            #     _l.info('initializing sub_task %s' % sub_task.options_object['file_path'])
            #
            #     ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
            #     celery_sub_tasks.append(ct)
            #
            # # chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id)).apply_async()
            # chord(celery_sub_tasks)(complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id))

            for sub_task in sub_tasks:
                _l.info(
                    "initializing sub_task %s" % sub_task.options_object["file_path"]
                )

                ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
                celery_sub_tasks.append(ct)

            _l.info("celery_sub_tasks len %s" % len(celery_sub_tasks))
            _l.info("celery_sub_tasks %s" % celery_sub_tasks)

            # chord(celery_sub_tasks, complex_transaction_csv_file_import_validate_parallel_finish.si(task_id=task_id)).apply_async()
            # chord(celery_sub_tasks)(complex_transaction_csv_file_import_parallel_finish.si(task_id=task_id))

            transaction.on_commit(
                lambda: chord(
                    celery_sub_tasks,
                    complex_transaction_csv_file_import_parallel_finish.si(
                        task_id=task_id
                    ),
                ).apply_async()
            )

    except Exception as e:
        _l.info("Exception occurred %s" % e)
        _l.info(traceback.format_exc())


# DEPRECATED
@finmars_task(
    name="integrations.complex_transaction_csv_file_import_validate_parallel_finish",
    bind=True,
)
def complex_transaction_csv_file_import_validate_parallel_finish(self, task_id, *args, **kwargs):
    try:
        _l.info(
            "complex_transaction_csv_file_import_validate_parallel_finish task_id %s "
            % task_id
        )

        celery_task = CeleryTask.objects.get(pk=task_id)

        scheme = ComplexTransactionImportScheme.objects.get(
            pk=celery_task.options_object["scheme_id"]
        )

        master_user = celery_task.master_user
        member = celery_task.member

        result_object = {
            "error_rows": [],
            "total_rows": celery_task.options_object["total_rows"],
            "processed_rows": 0,
        }

        _l.info(
            "complex_transaction_csv_file_import_validate_parallel_finish iterating over %s childs"
            % len(celery_task.children.all())
        )

        for sub_task in celery_task.children.all():
            if sub_task.result_object:
                if "error_rows" in sub_task.result_object:
                    result_object["error_rows"] = (
                        result_object["error_rows"]
                        + sub_task.result_object["error_rows"]
                    )

                if "processed_rows" in sub_task.result_object:
                    result_object["processed_rows"] = (
                        result_object["processed_rows"]
                        + sub_task.result_object["processed_rows"]
                    )

        result_object["stats_file_report"] = generate_file_report(
            result_object,
            master_user,
            scheme,
            "transaction_import.validate",
            "Transaction Import Validation",
            False,
        )

        # DEPRECATED
        # send_websocket_message(data={
        #     'type': 'transaction_import_status',
        #     'payload': {'task_id': task_id,
        #                 'state': Task.STATUS_DONE,
        #                 'error_rows': result_object['error_rows'],
        #                 'total_rows': result_object['total_rows'],
        #                 'processed_rows': result_object['processed_rows'],
        #                 'stats_file_report': result_object['stats_file_report'],
        #                 'scheme': scheme.id,
        #                 'scheme_object': {
        #                     'id': scheme.id,
        #                     'scheme_name': scheme.user_code,
        #                     'delimiter': scheme.delimiter,
        #                     'error_handler': scheme.error_handler,
        #                     'missing_data_handler': scheme.missing_data_handler
        #                 }}
        # }, level="member",
        #     context={"master_user": master_user, "member": member})

        celery_task.result_object = result_object

        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

    except Exception as e:
        _l.info("Exception occurred %s" % e)
        _l.info(traceback.format_exc())


# DEPRECATED
@finmars_task(
    name="integrations.complex_transaction_csv_file_import_validate", bind=True
)
def complex_transaction_csv_file_import_validate(self, task_id, *args, **kwargs):
    try:
        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.transactions.models import TransactionTypeInput

        celery_task = CeleryTask.objects.get(pk=task_id)
        parent_celery_task = celery_task.parent

        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        master_user = celery_task.master_user
        member = celery_task.member

        instance = ComplexTransactionCsvFileImport(
            task_id=task_id,
            master_user=master_user,
            member=member,
            skip_first_line=True,
        )

        scheme = ComplexTransactionImportScheme.objects.get(
            pk=celery_task.options_object["scheme_id"]
        )

        instance.scheme = scheme
        instance.error_handling = scheme.error_handler
        instance.delimiter = scheme.delimiter
        instance.missing_data_handler = scheme.missing_data_handler
        instance.file_path = celery_task.options_object["file_path"]

        _l.info("complex_transaction_csv_file_import_validate %s" % instance.file_path)

        instance.processed_rows = 0
        _l.info("complex_transaction_file_import: %s", instance)
        _l.info(
            "complex_transaction_file_import: instance.break_on_error %s",
            instance.break_on_error,
        )

        scheme_inputs = list(scheme.inputs.all())
        scheme_calculated_inputs = list(scheme.calculated_inputs.all())
        rule_scenarios = scheme.rule_scenarios.prefetch_related(
            "transaction_type", "fields", "fields__transaction_type_input"
        ).all()

        _l.info(
            "scheme %s - inputs=%s, rules=%s",
            scheme,
            [(i.name, i.column) for i in scheme_inputs],
            [(r.transaction_type.user_code) for r in rule_scenarios],
        )

        master_user = instance.master_user
        member = instance.member

        mapping_map = {
            Account: AccountMapping,
            Currency: CurrencyMapping,
            Instrument: InstrumentMapping,
            InstrumentType: InstrumentTypeMapping,
            Counterparty: CounterpartyMapping,
            Responsible: ResponsibleMapping,
            Strategy1: Strategy1Mapping,
            Strategy2: Strategy2Mapping,
            Strategy3: Strategy3Mapping,
            DailyPricingModel: DailyPricingModelMapping,
            PaymentSizeDetail: PaymentSizeDetailMapping,
            Portfolio: PortfolioMapping,
            PriceDownloadScheme: PriceDownloadSchemeMapping,
            Periodicity: PeriodicityMapping,
            AccrualCalculationModel: AccrualCalculationModelMapping,
        }

        props_map = {
            Account: "account",
            Currency: "currency",
            Instrument: "instrument",
            InstrumentType: "instrument_type",
            Counterparty: "counterparty",
            Responsible: "responsible",
            Strategy1: "strategy1",
            Strategy2: "strategy2",
            Strategy3: "strategy3",
            DailyPricingModel: "daily_pricing_model",
            PaymentSizeDetail: "payment_size_detail",
            Portfolio: "portfolio",
            PriceDownloadScheme: "price_download_scheme",
            Periodicity: "periodicity",
            AccrualCalculationModel: "accrual_calculation_model",
        }

        mapping_cache = {}

        def _get_default_relation(field):
            i = field.transaction_type_input

            model_class = i.content_type.model_class()
            model_map_class = mapping_map[model_class]

            key = props_map[model_class]

            v = None

            ecosystem_default = EcosystemDefault.objects.get(
                master_user=instance.master_user
            )

            # _l.info('key %s' % key)
            # _l.info('value %s' % value)

            if hasattr(ecosystem_default, key):
                v = getattr(ecosystem_default, key)
            else:
                v = model_map_class.objects.get(
                    master_user=instance.master_user, value="-"
                ).content_object

            return v

        def _convert_value(field, value, error_rows):
            i = field.transaction_type_input

            if i.value_type == TransactionTypeInput.STRING:
                return str(value)

            if i.value_type == TransactionTypeInput.SELECTOR:
                return str(value)

            elif i.value_type == TransactionTypeInput.NUMBER:
                return float(value)

            elif i.value_type == TransactionTypeInput.DATE:
                if not isinstance(value, date):
                    return formula._parse_date(value)
                else:
                    return value

            elif i.value_type == TransactionTypeInput.RELATION:
                model_class = i.content_type.model_class()
                model_map_class = mapping_map[model_class]

                v = None

                try:
                    v = model_map_class.objects.get(
                        master_user=instance.master_user, value=value
                    ).content_object
                except Exception:
                    try:
                        v = model_class.objects.get(
                            master_user=instance.master_user, user_code=value
                        )

                    except (model_class.DoesNotExist, KeyError):
                        v = None

                        _l.info(
                            "User code %s not found for %s "
                            % (value, field.transaction_type_input.name)
                        )

                if not v:
                    if instance.missing_data_handler == "set_defaults":
                        v = _get_default_relation(field)

                    else:
                        error_rows["error_message"] = (
                            error_rows["error_message"]
                            + " Can't find relation of "
                            + "["
                            + field.transaction_type_input.name
                            + "]"
                            + "(value:"
                            + value
                            + ")"
                        )

                return v

        def update_row_with_calculated_data(row, inputs):
            for i in scheme_calculated_inputs:
                # _l.info('update_row_with_calculated_data inputs %s' % inputs)

                try:
                    value = formula.safe_eval(i.name_expr, names=inputs)
                    row.append(value)

                except Exception:
                    _l.info(
                        "can't process calculated input: %s|%s",
                        i.name,
                        i.column,
                        exc_info=True,
                    )
                    row.append(None)

            return row

        def _validate_process_csv_file(file, orignal_file, original_file_name):
            delimiter = instance.delimiter.encode("utf-8").decode("unicode_escape")

            reader = csv.reader(
                file,
                delimiter=delimiter,
                quotechar=instance.quotechar,
                strict=False,
                skipinitialspace=True,
            )

            first_row = None
            input_column_name_map = {}

            for row_index, row in enumerate(reader):
                if row_index == 0:
                    first_row = row

                    _local_index = 0
                    for item in first_row:
                        input_column_name_map[item] = _local_index
                        _local_index = _local_index + 1

                # _l.info('_validate_process_csv_file row: %s -> %s', row_index, row)
                if (row_index == 0 and instance.skip_first_line) or not row:
                    _l.info("skip first row")
                    continue

                inputs_raw = {}
                inputs = {}
                inputs_error = []
                calculated_columns_error = []

                error_rows = {
                    "level": "info",
                    "error_message": "",
                    "inputs": inputs_raw,
                    "original_row_index": row_index,
                    "original_row": row,
                    "error_data": {
                        "columns": {
                            "imported_columns": [],
                            "calculated_columns": [],
                            "converted_imported_columns": [],
                            "transaction_type_selector": [],
                            "executed_input_expressions": [],
                        },
                        "data": {
                            "imported_columns": [],
                            "calculated_columns": [],
                            "converted_imported_columns": [],
                            "transaction_type_selector": [],
                            "executed_input_expressions": [],
                        },
                    },
                    "error_reaction": "Success",
                }

                for i in scheme_inputs:
                    error_rows["error_data"]["columns"]["imported_columns"].append(
                        i.name
                    )

                    if instance.scheme.column_matcher == "index":
                        try:
                            inputs_raw[i.name] = row[i.column - 1]
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                row[i.column - 1]
                            )
                        except Exception:
                            _l.debug(
                                "can't process input: %s|%s",
                                i.name,
                                i.column,
                                exc_info=True,
                            )
                            _l.debug("can't process inputs_raw: %s|%s", inputs_raw)
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                gettext_lazy("Invalid expression")
                            )
                            inputs_error.append(i)

                    if instance.scheme.column_matcher == "name":
                        try:
                            _col_index = input_column_name_map[i.name]

                            inputs_raw[i.name] = row[_col_index]
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                row[_col_index]
                            )
                        except Exception:
                            _l.debug(
                                "can't process input: %s|%s",
                                i.name,
                                i.column,
                                exc_info=True,
                            )
                            _l.debug("can't process inputs_raw: %s|%s", inputs_raw)
                            error_rows["error_data"]["data"]["imported_columns"].append(
                                gettext_lazy("Invalid expression")
                            )
                            inputs_error.append(i)

                # _l.info('Row %s inputs_raw: %s' % (row_index, inputs_raw))

                if scheme.filter_expression:
                    # expr = Expression.parseString("a == 1 and b == 2")
                    expr = Expression.parseString(scheme.filter_expression)

                    _l.info("scheme.filter_expression %s " % scheme.filter_expression)
                    _l.info("scheme.inputs_raw %s " % inputs_raw)

                    if expr(inputs_raw):
                        # filter passed

                        pass
                    else:
                        _l.info("Row skipped due filter %s" % row_index)
                        continue

                for i in scheme_inputs:
                    error_rows["error_data"]["columns"][
                        "converted_imported_columns"
                    ].append(
                        i.name + ": Conversion Expression " + "(" + i.name_expr + ")"
                    )

                    try:
                        inputs[i.name] = formula.safe_eval(
                            i.name_expr, names=inputs_raw
                        )
                        error_rows["error_data"]["data"][
                            "converted_imported_columns"
                        ].append(row[i.column - 1])
                    except Exception:
                        _l.info(
                            "can't process input: %s|%s",
                            i.name,
                            i.column,
                            exc_info=True,
                        )
                        error_rows["error_data"]["data"][
                            "converted_imported_columns"
                        ].append(gettext_lazy("Invalid expression"))
                        inputs_error.append(i)

                # _l.info('Row %s inputs_converted: %s' % (row_index, inputs))

                original_columns_count = len(row)

                row = update_row_with_calculated_data(row, inputs)

                # _l.info('Row %s inputs_with_calculated: %s' % (row_index, inputs))

                for i in scheme_calculated_inputs:
                    error_rows["error_data"]["columns"]["calculated_columns"].append(
                        i.name
                    )

                    try:
                        index = original_columns_count + i.column - 1

                        # _l.info('original_columns_count %s' % original_columns_count)
                        # _l.info('i.column %s' % i.column)
                        # _l.info('row %s' % row)

                        inputs[i.name] = row[index]

                        error_rows["error_data"]["data"]["calculated_columns"].append(
                            row[index]
                        )
                    except Exception:
                        _l.info(
                            "can't process input: %s|%s",
                            i.name,
                            i.column,
                            exc_info=True,
                        )
                        error_rows["error_data"]["data"]["calculated_columns"].append(
                            gettext_lazy("Invalid expression")
                        )
                        calculated_columns_error.append(i)

                if inputs_error:
                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        gettext_lazy("Can't process inputs: %(inputs)s")
                        % {
                            "inputs": ", ".join(
                                "[" + i.name + "]" for i in inputs_error
                            )
                        }
                    )
                    instance.error_rows.append(error_rows)

                    if instance.break_on_error:
                        error_rows["error_reaction"] = "Break"
                        instance.error_row_index = row_index
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"
                        continue

                try:
                    rule_value = formula.safe_eval(scheme.rule_expr, names=inputs)
                except Exception as e:
                    error_rows["level"] = "error"

                    _l.info("can't process rule expression", exc_info=True)
                    _l.info("error %s" % e)
                    error_rows["error_message"] = error_rows["error_message"] + str(
                        gettext_lazy("Can't eval rule expression")
                    )
                    instance.error_rows.append(error_rows)
                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows["error_reaction"] = "Break"
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"
                        continue

                if not rule_value:
                    _l.info("no rule value: %s", rule_value)

                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        gettext_lazy("Rule expression is invalid")
                    )

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows["error_reaction"] = "Break"
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"
                        continue

                # else:
                #     _l.info('rule value: %s', rule_value)

                processed_scenarios = 0
                matched_rule = False

                for scheme_rule in rule_scenarios:
                    matched_selector = False

                    selector_values = scheme_rule.selector_values.all()

                    for selector_value in selector_values:
                        if selector_value.value == rule_value:
                            matched_selector = True

                    if matched_selector:
                        processed_scenarios = processed_scenarios + 1

                        error_rows["error_data"]["columns"][
                            "transaction_type_selector"
                        ].append("TType Selector")

                        try:
                            rule = scheme_rule

                            error_rows["error_data"]["data"][
                                "transaction_type_selector"
                            ].append(rule_value)

                        except Exception:
                            error_rows["level"] = "error"

                            _l.info("rule does not find: %s", rule_value, exc_info=True)
                            error_rows["error_message"] = error_rows[
                                "error_message"
                            ] + str(
                                gettext_lazy(
                                    'Can\'t find transaction type by "%(value)s"'
                                )
                                % {"value": rule_value}
                            )
                            instance.error_rows.append(error_rows)

                            error_rows["error_data"]["data"][
                                "transaction_type_selector"
                            ].append(gettext_lazy("Invalid expression"))

                            if instance.break_on_error:
                                instance.error_row_index = row_index
                                error_rows["error_reaction"] = "Break"
                                instance.error_rows.append(error_rows)
                                return
                            else:
                                error_rows["error_reaction"] = "Continue import"
                                continue

                        _l.info("founded rule: %s -> %s", rule, rule.transaction_type)

                        fields = {}
                        fields_error = []

                        for field in rule.fields.all():
                            error_rows["error_data"]["columns"][
                                "executed_input_expressions"
                            ].append(field.transaction_type_input.name)

                            try:
                                field_value = formula.safe_eval(
                                    field.value_expr, names=inputs
                                )

                                field_value = _convert_value(
                                    field, field_value, error_rows
                                )

                                fields[field.transaction_type_input.name] = field_value

                                if hasattr(field_value, "name"):
                                    error_rows["error_data"]["data"][
                                        "executed_input_expressions"
                                    ].append(field_value.name)
                                else:
                                    error_rows["error_data"]["data"][
                                        "executed_input_expressions"
                                    ].append(field_value)

                            except (Exception, ValueError, formula.InvalidExpression):
                                _l.info(
                                    "can't process field: %s|%s",
                                    field.transaction_type_input.name,
                                    field.transaction_type_input.pk,
                                    exc_info=True,
                                )
                                fields_error.append(field)

                                error_rows["error_data"]["data"][
                                    "executed_input_expressions"
                                ].append(gettext_lazy("Invalid expression"))

                        if len(fields_error):
                            _l.info(
                                "fields (step 1): error=%s, values=%s",
                                fields_error,
                                fields,
                            )

                            _l.info(error_rows["error_message"])

                            inputs_messages = []

                            for field_error in fields_error:
                                message = (
                                    "["
                                    + field_error.transaction_type_input.name
                                    + "] "
                                    + "( TType Input, TType "
                                    + rule.transaction_type.name
                                    + " ["
                                    + rule.transaction_type.user_code
                                    + "] )"
                                )

                                inputs_messages.append(message)

                            error_rows["error_message"] = error_rows[
                                "error_message"
                            ] + str(
                                gettext_lazy("Can't process fields: %(messages)s")
                                % {
                                    "messages": ", ".join(
                                        str(m) for m in inputs_messages
                                    )
                                }
                            )

                            error_rows["level"] = "error"

                            if instance.break_on_error:
                                error_rows["error_reaction"] = "Break"
                                instance.error_row_index = row_index
                                instance.error_rows.append(error_rows)
                                return
                            else:
                                error_rows["error_reaction"] = "Continue import"
                                continue

                # print('matched_rule %s' % matched_rule)

                if processed_scenarios == 0:
                    error_rows["level"] = "error"

                    error_rows["error_message"] = error_rows["error_message"] + str(
                        "Selector does not match"
                    )

                    if instance.break_on_error:
                        instance.error_row_index = row_index
                        error_rows["error_reaction"] = "Break"
                        instance.error_rows.append(error_rows)
                        return
                    else:
                        error_rows["error_reaction"] = "Continue import"

                instance.error_rows.append(error_rows)

                # if fields_error:
                #
                #     if instance.break_on_error:
                #         error_rows['error_reaction'] = 'Break'
                #         instance.error_row_index = row_index
                #         return
                #     else:
                #         error_rows['error_reaction'] = 'Continue import'
                #         continue

                instance.processed_rows = instance.processed_rows + 1
                # instance.save()

                # DEPRECATED
                # send_websocket_message(data={
                #     'type': 'transaction_import_status',
                #     'payload': {
                #         'parent_task_id': celery_task.parent_id,
                #         'task_id': instance.task_id,
                #         'state': Task.STATUS_PENDING,
                #         'processed_rows': instance.processed_rows,
                #         'parent_total_rows': parent_celery_task.options_object['total_rows'],
                #         'total_rows': instance.total_rows,
                #         'scheme_name': scheme.user_code,
                #         'file_name': instance.file_name}
                # }, level="member",
                #     context={"master_user": master_user, "member": member})

        def _row_count_xlsx(file):
            wb = load_workbook(filename=file)

            if (
                instance.scheme.spreadsheet_active_tab_name
                and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames
            ):
                ws = wb[instance.scheme.spreadsheet_active_tab_name]
            else:
                ws = wb.active

            reader = []

            row_index = 0

            for r in ws.rows:
                row_index = row_index + 1

            return row_index

        def _row_count(file):
            delimiter = instance.delimiter.encode("utf-8").decode("unicode_escape")

            reader = csv.reader(
                file,
                delimiter=delimiter,
                quotechar=instance.quotechar,
                strict=False,
                skipinitialspace=True,
            )

            row_index = 0

            for row_index, row in enumerate(reader):
                pass
            return row_index

        instance.error_rows = []

        try:
            # with import_file_storage.open(instance.file_path, 'rb') as f:

            _l.info("Trying to open %s" % instance.file_path)
            with storage.open(instance.file_path, "rb") as f:
                with NamedTemporaryFile() as tmpf:
                    for chunk in f.chunks():
                        tmpf.write(chunk)
                    tmpf.flush()

                    os.link(tmpf.name, tmpf.name + ".xlsx")

                    if ".csv" in instance.file_path:
                        with open(
                            tmpf.name,
                            mode="rt",
                            encoding=instance.encoding,
                            errors="ignore",
                        ) as cfr:
                            instance.total_rows = _row_count(cfr)

                        with open(
                            tmpf.name,
                            mode="rt",
                            encoding=instance.encoding,
                            errors="ignore",
                        ) as cf:
                            _validate_process_csv_file(cf, f, "")

                    elif ".xlsx" in instance.file_path:
                        instance.total_rows = _row_count_xlsx(tmpf.name + ".xlsx")

                        with open(
                            tmpf.name,
                            mode="rt",
                            encoding=instance.encoding,
                            errors="ignore",
                        ) as cf:
                            _validate_process_csv_file(cf, f, tmpf.name + ".xlsx")

        except Exception:
            _l.info("Can't process file", exc_info=True)
            instance.error_message = gettext_lazy(
                "Invalid file format or file already deleted."
            )
        finally:
            # import_file_storage.delete(instance.file_path)
            storage.delete(instance.file_path)

        _l.info("transaction import validation completed")

        instance.error = (
            bool(instance.error_message)
            or (instance.error_row_index is not None)
            or bool(instance.error_rows)
        )

        # instance.stats_file_report = generate_file_report(instance, master_user, 'transaction_import.validate',
        #                                                   'Transaction Import Validation')

        # DEPRECATED
        # send_websocket_message(data={
        #     'type': 'transaction_import_status',
        #     'payload': {
        #         'parent_task_id': celery_task.parent_id,
        #         'task_id': instance.task_id,
        #         'state': Task.STATUS_DONE,
        #         'processed_rows': instance.processed_rows,
        #         'parent_total_rows': parent_celery_task.options_object['total_rows'],
        #         'total_rows': instance.total_rows,
        #         'file_name': instance.file_name,
        #         'error_rows': instance.error_rows,
        #         'stats_file_report': instance.stats_file_report,
        #         'scheme': scheme.id,
        #         'scheme_object': {
        #             'id': scheme.id,
        #             'scheme_name': scheme.user_code,
        #             'delimiter': scheme.delimiter,
        #             'error_handler': scheme.error_handler,
        #             'missing_data_handler': scheme.missing_data_handler
        #         }}
        # }, level="member",
        #     context={"master_user": master_user, "member": member})

        result_object = {
            "processed_rows": instance.processed_rows,
            "total_rows": instance.total_rows,
            "file_name": instance.file_name,
            "error_rows": instance.error_rows,
            "stats_file_report": instance.stats_file_report,
        }

        celery_task.result_object = result_object
        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

        return instance

    except Exception as e:
        _l.info("Exception occurred %s" % e)
        _l.info(traceback.format_exc())


# DEPRECATED
# @finmars_task(name='integrations.complex_transaction_csv_file_import_validate_parallel', bind=True)
def complex_transaction_csv_file_import_validate_parallel(task_id, *args, **kwargs):
    try:
        _l.info(
            "complex_transaction_csv_file_import_validate_parallel: task_id %s"
            % task_id
        )

        celery_task = CeleryTask.objects.get(pk=task_id)

        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.save()

        options_object = celery_task.options_object

        sub_tasks = []

        lines_per_file = 300
        header_line = None

        def _get_path(master_user, file_name, ext):
            return "%s/public/%s.%s" % (master_user.space_code, file_name, ext)

        chunk = None

        with storage.open(celery_task.options_object["file_path"], "rb") as f:
            _l.info("Start reading file to split it into chunks")
            _l.debug(
                "Start reading file to split it into chunks options %s"
                % celery_task.options_object
            )

            ext = celery_task.options_object["file_path"].split(".")[-1]

            for lineno, line in enumerate(f):
                # _l.info('line %s' % lineno)

                if lineno == 0:
                    header_line = line
                    # _l.info('set header line %s' % lineno)

                if lineno % lines_per_file == 0:
                    if chunk is not None:
                        # _l.info("Saving chunk %s" % chunk)
                        storage.save(
                            chunk_path, chunk
                        )  # save working chunk before creating new one

                    chunk_filename = "%s_chunk_file_%s" % (
                        celery_task.id,
                        str(lineno) + "_" + str(lineno + lines_per_file),
                    )
                    chunk_path = _get_path(celery_task.master_user, chunk_filename, ext)

                    # _l.info('creating chunk file %s' % chunk_path)

                    chunk = BytesIO()
                    if lineno != 0:
                        chunk.write(header_line)

                    _l.info("creating chunk %s" % chunk_filename)

                    # _l.info('creating sub task for %s' % chunk_filename)

                    sub_task = CeleryTask.objects.create(
                        master_user=celery_task.master_user,
                        member=celery_task.member,
                        parent=celery_task,
                    )

                    sub_task_options_object = copy.deepcopy(celery_task.options_object)
                    sub_task_options_object["file_path"] = chunk_path

                    sub_task.options_object = sub_task_options_object
                    sub_task.save()

                    sub_tasks.append(sub_task)

                chunk.write(line)

            _l.info("Saving last chunk")
            storage.save(
                chunk_path, chunk
            )  # save working chunk before creating new one

        _l.info("sub_tasks created %s" % len(sub_tasks))
        _l.info("original file total rows %s" % lineno)

        options_object["total_rows"] = lineno

        celery_task.options_object = options_object
        celery_task.save()

        celery_sub_tasks = []

        for sub_task in sub_tasks:
            _l.info("initializing sub_task %s" % sub_task.options_object["file_path"])

            ct = complex_transaction_csv_file_import_validate.s(task_id=sub_task.id)
            celery_sub_tasks.append(ct)

        _l.info("celery_sub_tasks len %s" % len(celery_sub_tasks))
        _l.info("celery_sub_tasks %s" % celery_sub_tasks)

        # chord(celery_sub_tasks, complex_transaction_csv_file_import_validate_parallel_finish.si(task_id=task_id)).apply_async()
        chord(celery_sub_tasks)(
            complex_transaction_csv_file_import_validate_parallel_finish.si(
                task_id=task_id
            )
        )

    except Exception as e:
        _l.info("Exception occurred %s" % e)
        _l.info(traceback.format_exc())


@finmars_task(
    name="integrations.complex_transaction_csv_file_import_by_procedure", bind=True
)
def complex_transaction_csv_file_import_by_procedure(
    self, procedure_instance_id, transaction_file_result_id, *args, **kwargs
):
    with transaction.atomic():
        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.procedures.models import RequestDataFileProcedureInstance

        procedure_instance = RequestDataFileProcedureInstance.objects.get(
            id=procedure_instance_id
        )
        transaction_file_result = TransactionFileResult.objects.get(
            id=transaction_file_result_id
        )

        try:
            _l.debug(
                "complex_transaction_csv_file_import_by_procedure looking for scheme %s "
                % procedure_instance.procedure.scheme_user_code
            )

            scheme = ComplexTransactionImportScheme.objects.get(
                master_user=procedure_instance.master_user,
                user_code=procedure_instance.procedure.scheme_user_code,
            )

            text = "Data File Procedure %s. File is received. Decrypting file" % (
                procedure_instance.procedure.user_code
            )

            send_system_message(
                master_user=procedure_instance.master_user,
                performed_by="System",
                description=text,
            )

            _l.debug("trying to open %s" % transaction_file_result.file_path)

            with storage.open(transaction_file_result.file_path, "rb") as f:
                try:
                    encrypted_text = f.read()

                    rsa_cipher = RSACipher()

                    aes_key = None

                    try:
                        aes_key = rsa_cipher.decrypt(
                            procedure_instance.private_key,
                            procedure_instance.symmetric_key,
                        )

                        _l.debug(
                            "complex_transaction_csv_file_import_by_procedure decrypting symmetric key"
                        )

                    except Exception as e:
                        _l.debug(
                            "complex_transaction_csv_file_import_by_procedure AES Key decryption error %s"
                            % e
                        )

                    aes_cipher = AESCipher(aes_key)

                    decrypt_text = None

                    try:
                        decrypt_text = aes_cipher.decrypt(encrypted_text)

                        _l.debug(
                            "complex_transaction_csv_file_import_by_procedure decrypting text file"
                        )

                    except Exception as e:
                        _l.debug(
                            "complex_transaction_csv_file_import_by_procedure Text decryption error %s"
                            % e
                        )

                    _l.debug(
                        "complex_transaction_csv_file_import_by_procedure file decrypted"
                    )

                    _l.debug("Size of decrypted text: %s" % len(decrypt_text))

                    with NamedTemporaryFile() as tmpf:
                        _l.debug("tmpf.name %s" % tmpf.name)

                        tmpf.write(decrypt_text.encode("utf-8"))
                        tmpf.flush()

                        file_name = "%s-%s" % (
                            timezone.now().strftime("%Y%m%d%H%M%S"),
                            uuid.uuid4().hex,
                        )
                        file_path = "%s/public/%s.csv" % (
                            procedure_instance.master_user.space_code,
                            file_name,
                        )

                        storage.save(file_path, tmpf)

                        _l.debug(
                            "complex_transaction_csv_file_import_by_procedure tmp file filled"
                        )

                        instance = ComplexTransactionCsvFileImport(
                            scheme=scheme,
                            file_path=file_path,
                            missing_data_handler=scheme.missing_data_handler,
                            error_handling=scheme.error_handler,
                            delimiter=scheme.delimiter,
                            member=procedure_instance.member,
                            master_user=procedure_instance.master_user,
                        )

                    _l.debug(
                        "complex_transaction_csv_file_import_by_procedure instance: %s"
                        % instance
                    )

                    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

                    file_name = "%s-%s" % (
                        timezone.now().strftime("%Y%m%d%H%M%S"),
                        uuid.uuid4().hex,
                    )
                    file_name_hash = hashlib.md5(file_name.encode("utf-8")).hexdigest()

                    file_report = FileReport()

                    file_report.upload_file(
                        file_name="Data Procedure %s (%s).csv"
                        % (current_date_time, file_name_hash),
                        text=decrypt_text,
                        master_user=procedure_instance.master_user,
                    )
                    file_report.master_user = procedure_instance.master_user
                    file_report.name = "'Transaction Import File. Procedure ' %s %s" % (
                        procedure_instance.id,
                        current_date_time,
                    )
                    file_report.file_name = "Data Procedure %s (%s).csv" % (
                        current_date_time,
                        file_name_hash,
                    )
                    file_report.type = "transaction_import.import"
                    file_report.notes = (
                        "Transaction Import File. Procedure %s" % procedure_instance.id
                    )
                    file_report.content_type = "text/csv"

                    file_report.save()

                    _l.debug("file_report %s" % file_report)

                    text = "Data File Procedure %s. File is received. Start Import" % (
                        procedure_instance.procedure.user_code
                    )

                    send_system_message(
                        master_user=procedure_instance.master_user,
                        performed_by="System",
                        description=text,
                        attachments=[file_report.id],
                    )

                    options_object = {}
                    options_object["file_path"] = instance.file_path
                    options_object["scheme_id"] = instance.scheme.id
                    options_object["execution_context"] = {"started_by": "procedure"}

                    total_rows = 0

                    with storage.open(options_object["file_path"], "rb") as f1:
                        _l.info("Start reading file to split it into chunks")

                        for lineno, line in enumerate(f1):
                            total_rows = lineno

                    options_object["total_rows"] = total_rows

                    _l.debug(
                        "complex_transaction_csv_file_import_by_procedure total_rows %s"
                        % options_object["total_rows"]
                    )

                    celery_task = CeleryTask.objects.create(
                        master_user=procedure_instance.master_user,
                        member=procedure_instance.member,
                        options_object=options_object,
                        verbose_name="Transaction Import",
                        type="transaction_import",
                    )

                    celery_task.save()

                    # Creating subtask
                    sub_task = CeleryTask.objects.create(
                        master_user=celery_task.master_user,
                        member=celery_task.member,
                        parent=celery_task,
                    )

                    sub_task_options_object = copy.deepcopy(celery_task.options_object)

                    sub_task.options_object = sub_task_options_object

                    sub_task.save()

                    # transaction.on_commit(
                    #     lambda: complex_transaction_csv_file_import.apply_async(kwargs={'instance': instance, 'execution_context': {'started_by': 'procedure'}}))

                    # transaction.on_commit(
                    #     lambda: complex_transaction_csv_file_import.apply_async(kwargs={'task_id': sub_task.pk}))

                    transaction.on_commit(
                        lambda: transaction_import.apply_async(
                            kwargs={"task_id": sub_task.id, 'context': {
                                'space_code': sub_task.master_user.space_code,
                                'realm_code': sub_task.master_user.realm_code
                            }},
                            queue="backend-background-queue",
                        )
                    )

                    # celery_sub_tasks = []
                    #
                    # ct = complex_transaction_csv_file_import.s(task_id=sub_task.id)
                    # celery_sub_tasks.append(ct)
                    #
                    # _l.info("Creating %s subtasks" % len(celery_sub_tasks))
                    #
                    # # chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(task_id=celery_task.pk)).apply_async()
                    #
                    # transaction.on_commit(
                    #     lambda: chord(celery_sub_tasks, complex_transaction_csv_file_import_parallel_finish.si(
                    #         task_id=celery_task.pk)).apply_async())

                except Exception as e:
                    _l.error(
                        "complex_transaction_csv_file_import_by_procedure decryption error %s"
                        % e
                    )

        except ComplexTransactionImportScheme.DoesNotExist:
            text = (
                "Data File Procedure %s. Can't import file, Import scheme %s is not found"
                % (
                    procedure_instance.procedure.user_code,
                    procedure_instance.procedure.scheme_name,
                )
            )

            send_system_message(
                master_user=procedure_instance.master_user,
                performed_by="System",
                description=text,
            )

            _l.error(
                "complex_transaction_csv_file_import_by_procedure scheme %s not found"
                % procedure_instance.procedure.scheme_name
            )

            procedure_instance.status = RequestDataFileProcedureInstance.STATUS_ERROR
            procedure_instance.save()


@finmars_task(
    name="integrations.complex_transaction_csv_file_import_by_procedure_json", bind=True
)
def complex_transaction_csv_file_import_by_procedure_json(
    self, procedure_instance_id, celery_task_id, *args, **kwargs
):
    _l.info(
        f"complex_transaction_csv_file_import_by_procedure_json  procedure_instance_id "
        f"{procedure_instance_id} celery_task_id {celery_task_id}"
    )

    from poms.procedures.models import RequestDataFileProcedureInstance

    procedure_instance = RequestDataFileProcedureInstance.objects.get(
        id=procedure_instance_id
    )
    celery_task = CeleryTask.objects.get(id=celery_task_id)
    celery_task.celery_task_id = self.request.id
    celery_task.save()

    try:
        _l.info(
            f"complex_transaction_csv_file_import_by_procedure_json looking for "
            f"scheme {procedure_instance.procedure.scheme_user_code}"
        )

        scheme = ComplexTransactionImportScheme.objects.get(
            master_user=procedure_instance.master_user,
            user_code=procedure_instance.procedure.scheme_user_code,
        )

        options_object = celery_task.options_object

        options_object["file_path"] = ""
        options_object["file_name"] = ""
        options_object["scheme_id"] = scheme.id
        options_object["execution_context"] = {
            "started_by": "procedure",
            "date_from": str(procedure_instance.date_from),
            "date_to": str(procedure_instance.date_to),
        }

        celery_task.options_object = options_object
        celery_task.save()

        text = (
            f"Data File Procedure {procedure_instance.procedure.user_code}. "
            f"File is received. Importing JSON"
        )

        send_system_message(
            master_user=procedure_instance.master_user,
            performed_by="System",
            description=text,
        )

        transaction.on_commit(
            lambda: transaction_import.apply_async(
                kwargs={
                    "task_id": celery_task.id,
                    "procedure_instance_id": procedure_instance_id, 'context': {
                        'space_code': celery_task.master_user.space_code,
                        'realm_code': celery_task.master_user.realm_code
                    }
                },
                queue="backend-background-queue",
            )
        )

    except Exception as e:
        _l.info(f"complex_transaction_csv_file_import_by_procedure_json err {e}")

        text = (
            f"Data File Procedure {procedure_instance.procedure.user_code}. "
            f"Can't import json, Error {e}"
        )

        send_system_message(
            master_user=procedure_instance.master_user,
            performed_by="System",
            description=text,
        )

        _l.debug(
            f"complex_transaction_csv_file_import_by_procedure scheme "
            f"{procedure_instance.procedure.scheme_name} not found"
        )

        procedure_instance.status = RequestDataFileProcedureInstance.STATUS_ERROR
        procedure_instance.save()


def create_counterparty_from_callback_data(data, master_user, member) -> Counterparty:
    from poms.counterparties.serializers import CounterpartySerializer
    from poms.counterparties.models import CounterpartyGroup

    func = "create_counterparty_from_database"

    proxy_request = ProxyRequest(ProxyUser(member, master_user))
    activate(proxy_request)
    context = {
        "master_user": master_user,
        "request": proxy_request,
        "member": member,
    }

    group = CounterpartyGroup.objects.filter(master_user=master_user).first()
    company_data = {
        "user_code": data.get("user_code"),
        "name": data.get("name"),
        "short_name": data.get("short_name"),
        "public_name": data.get("public_name"),
        "notes": data.get("notes"),
        "group": group.id if group else None,
    }

    _l.info(f"{func} started, company_data={company_data}")

    try:
        instance = Counterparty.objects.get(
            master_user=master_user,
            user_code=company_data["user_code"],
        )
        serializer = CounterpartySerializer(
            data=company_data,
            context=context,
            instance=instance,
        )

    except Counterparty.DoesNotExist:
        serializer = CounterpartySerializer(data=company_data, context=context)

    except Exception as e:
        _l.error(f"{func} unexpected {e}\n {traceback.format_exc()}")
        raise e

    if serializer.is_valid():
        counter_party = serializer.save()
        _l.info(
            f"{func} Counterparty {company_data['user_code']} is imported successfully"
        )
        return counter_party

    else:
        _l.error(f"{func} invalid company data {serializer.errors}")
        raise Exception(serializer.errors)


def create_currency_from_callback_data(data, master_user, member) -> Currency:
    from poms.currencies.serializers import CurrencySerializer

    func = "create_currency_from_finmars_database"

    proxy_request = ProxyRequest(ProxyUser(member, master_user))
    activate(proxy_request)
    context = {
        "master_user": master_user,
        "request": proxy_request,
        "member": member,
    }

    currency_data = {
        "user_code": data.get("user_code"),
        "name": data.get("name"),
        "short_name": data.get("short_name"),
        "pricing_condition": PricingCondition.NO_VALUATION,
    }

    _l.info(f"{func} currency_data={currency_data}")

    try:
        instance = Currency.objects.get(
            master_user=master_user,
            user_code=currency_data["user_code"],
        )
        serializer = CurrencySerializer(
            data=currency_data,
            context=context,
            instance=instance,
        )
    except Currency.DoesNotExist:
        serializer = CurrencySerializer(data=currency_data, context=context)

    except Exception as e:
        _l.error(f"{func} unexpected {e}\n {traceback.format_exc()}")
        raise Exception(e)

    if serializer.is_valid():
        currency = serializer.save()

        for policy in currency.pricing_policies.all():
            policy.default_value = currency_data["user_code"] + ".USD"
            policy.save()

        _l.info(
            f"{func} Currency {currency_data['user_code']} is imported successfully"
        )

        return currency

    else:
        _l.error(f"{func} invalid currency data {serializer.errors}")
        raise RuntimeError(serializer.errors)


def handle_currency_and_instrument_api_data(
    api_data: dict,
    task: CeleryTask,
    caller: str,
) -> Instrument:
    """
    Creates/Updates Currency & Instrument from raw API data
    """
    log = "handle_currency_and_instrument_api_data"

    _l.info(f"{log} called from {caller} with data={api_data}")

    currency = create_currency_from_callback_data(
        api_data["currencies"][0],
        task.master_user,
        task.member,
    )

    instrument_data = api_data["instruments"][0]
    instrument_data["pricing_currency"] = currency.user_code

    instrument = create_instrument_from_finmars_database(
        instrument_data,
        task.master_user,
        task.member,
    )

    _l.info(
        f"{log} successfully created/updated instrument={instrument.user_code} "
        f"currency={currency.user_code}"
    )

    return instrument


def update_task_with_instrument_data(data: dict, task: CeleryTask):
    func = f"update_task_with_instrument_data, task.id={task.id}"

    try:
        instrument = handle_currency_and_instrument_api_data(
            api_data=data,
            task=task,
            caller=func,
        )

    except Exception as e:
        err_msg = f"{func} unexpected {repr(e)}"
        _l.error(err_msg)
        task.status = CeleryTask.STATUS_ERROR
        task.error_message = err_msg
        task.save()

    else:
        task_done_with_instrument_info(instrument, task)


def update_task_with_simple_instrument(remote_task_id: int, task: CeleryTask):
    result = task.result_object or {}
    result["remote_task_id"] = remote_task_id

    instrument = create_simple_instrument(task)

    if instrument:
        result["result_id"] = instrument.pk
        result["name"] = instrument.name
        result["short_name"] = instrument.short_name
        result["user_code"] = instrument.user_code
        # Important: Do not change status, task still in pending

    else:
        task.status = CeleryTask.STATUS_ERROR

    task.result_object = result
    task.save()


def update_task_with_currency_data(currency_data: dict, task: CeleryTask):
    func = f"update_task_with_currency_data, task.id={task.id}"

    _l.info(f"{func} currency_data={currency_data}")

    try:
        currency = create_currency_from_callback_data(
            currency_data, task.master_user, task.member
        )
        result = task.result_object
        result["result_id"] = currency.id
        result["user_code"] = currency.user_code
        result["short_name"] = currency.short_name
        result["name"] = currency.name

        task.result_object = result
        task.status = CeleryTask.STATUS_DONE
        task.save()

    except Exception as e:
        err_msg = f"{func} unexpected {repr(e)}"
        update_task_with_error(task, err_msg)


def update_task_with_company_data(company_data: dict, task: CeleryTask):
    func = f"update_task_with_company_data, task.id={task.id}"

    _l.info(f"{func} company_data={company_data}")

    try:
        company = create_counterparty_from_callback_data(
            company_data, task.master_user, task.member
        )
        result = task.result_object
        result["result_id"] = company.id
        result["user_code"] = company.user_code
        result["short_name"] = company.short_name
        result["name"] = company.name
        task.result_object = result
        task.status = CeleryTask.STATUS_DONE
        task.save()

    except Exception as e:
        err_msg = f"{func} unexpected {repr(e)}"
        update_task_with_error(task, err_msg)


def update_task_with_price_data(price_data: dict, task: CeleryTask):
    func = f"update_task_with_price_data, task.id={task.id}"

    _l.info(f"{func} price_data={price_data}")

    raise NotImplemented

    # try:  # TODO LATER
    #     create_prices_from_callback_data(
    #         price_data, task.master_user, task.member
    #     )
    #     result = task.result_object
    #     result["result_id"] = 0
    #     task.result_object = result
    #     task.status = CeleryTask.STATUS_DONE
    #     task.save()
    #
    # except Exception as e:
    #     err_msg = f"{func} unexpected {repr(e)}"
    #     update_task_with_error(task, err_msg)


def import_from_database_task(task_id: int, operation: str):
    func = f"import_{operation}_finmars_database, task_id={task_id}"
    try:
        task = CeleryTask.objects.get(id=task_id)
    except CeleryTask.DoesNotExist:
        _l.error(f"{func} no task with id={task_id}!")
        return

    if not task.options_object:
        err_msg = f"{func} task id={task_id} no options with {operation} data"
        update_task_with_error(task, err_msg)
        return

    BACKEND_CALLBACK_URLS = get_backend_callback_url()

    if operation not in BACKEND_CALLBACK_URLS:
        _l.error(f"{func} invalid operation {operation}")
        return

    options = {
        "data": task.options_object,
        "request_id": task.pk,
        "base_api_url": task.master_user.space_code,
        "callback_url": BACKEND_CALLBACK_URLS[operation],
    }
    task.options_object = options
    task.save()

    _l.info(f"{func} started, request_options={options}")

    try:
        monad: Monad = DatabaseService().get_monad(operation, options)

        if monad.status == MonadStatus.DATA_READY:
            _l.info(f"{func} received {operation} data={monad.data}")

            if operation == "instrument":
                update_task_with_instrument_data(data=monad.data, task=task)
            elif operation == "currency":
                update_task_with_currency_data(currency_data=monad.data, task=task)
            elif operation == "company":
                update_task_with_company_data(company_data=monad.data, task=task)
            elif operation == "price":
                update_task_with_price_data(price_data=monad.data, task=task)
            else:
                raise RuntimeError(f"invalid DatabaseService operation={operation}")

        elif monad.status == MonadStatus.TASK_CREATED:
            _l.info(f"{func} received remote task_id={monad.task_id}")

            if operation == "instrument":
                update_task_with_simple_instrument(
                    remote_task_id=monad.task_id,
                    task=task,
                )
            else:
                result = task.result_object
                result["remote_task_id"] = monad.task_id
                task.result_object = result
                task.save()

        else:
            err_msg = f"{func} received error={monad.message}"
            update_task_with_error(task, err_msg)

    except Exception as e:
        err_msg = f"{func} unexpected {repr(e)}\n  {traceback.format_exc()}"
        update_task_with_error(task, err_msg)


def import_instrument_finmars_database(task_id: int):
    import_from_database_task(task_id=task_id, operation="instrument")


@finmars_task(name="integrations.download_instrument_finmars_database_async", bind=True)
def download_instrument_finmars_database_async(self, task_id, *args, **kwargs):
    _l.info(f"download_instrument_finmars_database_async {task_id}")
    import_instrument_finmars_database(task_id)


def import_currency_finmars_database(task_id: int):
    import_from_database_task(task_id=task_id, operation="currency")


@finmars_task(name="integrations.download_instrument_finmars_database_async", bind=True)
def download_currency_finmars_database_async(self, task_id, *args, **kwargs):
    _l.info(f"download_currency_finmars_database_async {task_id}")
    import_currency_finmars_database(task_id)


def import_company_finmars_database(task_id: int):
    import_from_database_task(task_id=task_id, operation="company")


@finmars_task(name="integrations.download_instrument_finmars_database_async", bind=True)
def download_company_finmars_database_async(self, task_id, *args, **kwargs):
    _l.info(f"download_company_finmars_database_async {task_id}")
    import_company_finmars_database(task_id)


def import_price_finmars_database(task_id: int):
    import_from_database_task(task_id=task_id, operation="price")


@finmars_task(name="integrations.download_instrument_finmars_database_async", bind=True)
def download_price_finmars_database_async(self, task_id, *args, **kwargs):
    _l.info(f"download_price_finmars_database_async {task_id}")
    import_company_finmars_database(task_id)


FINAL_STATUSES = {
    CeleryTask.STATUS_DONE,
    CeleryTask.STATUS_ERROR,
    CeleryTask.STATUS_TIMEOUT,
    CeleryTask.STATUS_CANCELED,
    CeleryTask.STATUS_TRANSACTIONS_ABORTED,
}


@finmars_task(name="integrations.ttl_finisher")
def ttl_finisher(task_id: int, *args, **kwargs):
    func = f"ttl_finisher for task.id={task_id}"
    _l.info(f"{func} started")

    task = CeleryTask.objects.filter(id=task_id).first()
    if not task:
        _l.error(f"{func} no such task!")
        return

    if task.status not in FINAL_STATUSES:
        _l.warning(f"{func} task.status={task.status} ttl={task.ttl} expired!")
        task.status = CeleryTask.STATUS_TIMEOUT
        task.save()

        return

    _l.info(f"{func} task.status={task.status} no action required")
