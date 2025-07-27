import copy
import csv
import json
import os
import re
import traceback
from datetime import datetime, date
from functools import reduce
from logging import getLogger
from operator import or_
from tempfile import NamedTemporaryFile
from typing import Any, Optional

from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.utils.timezone import now

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from poms.accounts.models import AccountType
from poms.celery_tasks.models import CeleryTask
from poms.common.models import ProxyRequest, ProxyUser
from poms.common.storage import get_storage
from poms.common.utils import get_serializer
from poms.counterparties.models import CounterpartyGroup, ResponsibleGroup
from poms.csv_import.models import (
    CsvImportScheme,
    ProcessType,
    SimpleImportConversionItem,
    SimpleImportImportedItem,
    SimpleImportProcessItem,
    SimpleImportProcessPreprocessItem,
    SimpleImportResult,
)
from poms.csv_import.serializers import SimpleImportResultSerializer
from poms.csv_import.tasks import simple_import_bulk_insert_final_updates_procedure
from poms.currencies.models import Currency
from poms.expressions_engine import formula
from poms.file_reports.models import FileReport
from poms.instruments.models import (
    AccrualCalculationModel,
    Country,
    DailyPricingModel,
    Instrument,
    InstrumentType,
    PaymentSizeDetail,
    Periodicity,
    PricingCondition,
    PricingPolicy,
)
from poms.obj_attrs.models import GenericAttributeType, GenericClassifier
from poms.procedures.models import RequestDataFileProcedureInstance
from poms.strategies.models import (
    Strategy1Subgroup,
    Strategy2Subgroup,
    Strategy3Subgroup,
)
from poms.system_messages.handlers import send_system_message
from poms.users.models import EcosystemDefault
from poms.portfolios.models import PortfolioType
from dateutil.parser import parse


storage = get_storage()

_l = getLogger("poms.csv_import")

ACCRUAL_MAP = {
    "Actual/Actual (AFB)": AccrualCalculationModel.DAY_COUNT_ACT_ACT_AFB,
    "Actual/Actual (ICMA)": AccrualCalculationModel.DAY_COUNT_ACT_ACT_ICMA,
    "Actual/Actual (ISDA)": AccrualCalculationModel.DAY_COUNT_ACT_ACT_ISDA,
    "Actual/360": AccrualCalculationModel.DAY_COUNT_ACT_360,
    "Actual/364": AccrualCalculationModel.DAY_COUNT_ACT_364,
    "Actual/365 (Actual/365F)": AccrualCalculationModel.DAY_COUNT_ACT_365_FIXED,
    "Actual/366": AccrualCalculationModel.DAY_COUNT_ACT_366,
    "Actual/365L": AccrualCalculationModel.DAY_COUNT_ACT_365L,
    "Actual/365A": AccrualCalculationModel.DAY_COUNT_ACT_365A,
    "30/360 US": AccrualCalculationModel.DAY_COUNT_30_360_US,
    "NL/365": AccrualCalculationModel.DAY_COUNT_NL_365,
    "BD/252": AccrualCalculationModel.DAY_COUNT_BD_252,
    "30E+/360": AccrualCalculationModel.DAY_COUNT_30E_PLUS_360,
    "30E/360": AccrualCalculationModel.DAY_COUNT_30E_360,
    "30/360 (30/360 ISDA)": AccrualCalculationModel.DAY_COUNT_30_360_ISDA,
    "30/360 (30/360 ISMA)": AccrualCalculationModel.DAY_COUNT_30_360_ISMA,
    "30/360 German": AccrualCalculationModel.DAY_COUNT_30_360_GERMAN,
    # currently unused by CBOND
    "Actual/365": AccrualCalculationModel.DAY_COUNT_ACT_365,
    "30/365": AccrualCalculationModel.DAY_COUNT_30_365,
    "Simple": AccrualCalculationModel.DAY_COUNT_SIMPLE,
    "none": AccrualCalculationModel.DAY_COUNT_NONE,
}
PERIODICITY_MAP = {
    1: Periodicity.ANNUALLY,
    2: Periodicity.SEMI_ANNUALLY,
    4: Periodicity.QUARTERLY,
    6: Periodicity.BIMONTHLY,
    12: Periodicity.MONTHLY,
}
RELATION_FIELDS_MAP = {
    "instrument": Instrument,
    "currency": Currency,
    "pricing_currency": Currency,
    "accrued_currency": Currency,
    "type": AccountType,
    "pricing_condition": PricingCondition,
    "instrument_type": InstrumentType,
    "country": Country,
    "pricing_policy": PricingPolicy,
    "payment_size_detail": PaymentSizeDetail,
    "co_directional_exposure_currency": Currency,
    "counter_directional_exposure_currency": Currency,
    "daily_pricing_model": DailyPricingModel,
    "portfolio_type": PortfolioType,
    "accrual_calculation_model": AccrualCalculationModel,
    "periodicity": Periodicity,

    "register_currency": Currency,
    "register_pricing_policy": PricingPolicy,
    "register_instrument_type": InstrumentType
}


# TO BE DEPRECATED SOON !
# Use InstrumentTypeProcess.fill_instrument_with_instrument_type_defaults
def set_defaults_from_instrument_type(
    instrument_object, instrument_type, ecosystem_default
):
    try:
        # Set system attributes

        if instrument_type.payment_size_detail_id:
            instrument_object["payment_size_detail"] = (
                instrument_type.payment_size_detail_id
            )
        else:
            instrument_object["payment_size_detail"] = None

        if instrument_type.accrued_currency_id:
            instrument_object["accrued_currency"] = instrument_type.accrued_currency_id
        else:
            instrument_object["accrued_currency"] = None

        instrument_object["price_multiplier"] = instrument_type.price_multiplier
        instrument_object["default_price"] = instrument_type.default_price
        instrument_object["maturity_date"] = instrument_type.maturity_date
        instrument_object["maturity_price"] = instrument_type.maturity_price
        instrument_object["accrued_multiplier"] = instrument_type.accrued_multiplier
        instrument_object["default_accrued"] = instrument_type.default_accrued

        if instrument_type.exposure_calculation_model_id:
            instrument_object["exposure_calculation_model"] = (
                instrument_type.exposure_calculation_model_id
            )
        else:
            instrument_object["exposure_calculation_model"] = None

        if instrument_type.pricing_condition_id:
            instrument_object["pricing_condition"] = (
                instrument_type.pricing_condition_id
            )
        else:
            instrument_object["pricing_condition"] = None

        try:
            instrument_object["long_underlying_instrument"] = Instrument.objects.get(
                master_user=instrument_type.master_user,
                user_code=instrument_type.long_underlying_instrument,
            ).pk
        except Exception:
            _l.info("Could not set long_underlying_instrument, fallback to default")
            instrument_object["long_underlying_instrument"] = (
                ecosystem_default.instrument.pk
            )

        instrument_object["underlying_long_multiplier"] = (
            instrument_type.underlying_long_multiplier
        )

        try:
            instrument_object["short_underlying_instrument"] = Instrument.objects.get(
                master_user=instrument_type.master_user,
                user_code=instrument_type.short_underlying_instrument,
            ).pk
        except Exception:
            _l.info("Could not set short_underlying_instrument, fallback to default")
            instrument_object["short_underlying_instrument"] = (
                ecosystem_default.instrument.pk
            )

        instrument_object["underlying_short_multiplier"] = (
            instrument_type.underlying_short_multiplier
        )
        instrument_object["long_underlying_exposure"] = (
            instrument_type.long_underlying_exposure_id
        )
        instrument_object["short_underlying_exposure"] = (
            instrument_type.short_underlying_exposure_id
        )

        try:
            instrument_object["co_directional_exposure_currency"] = (
                Currency.objects.get(
                    master_user=instrument_type.master_user,
                    user_code=instrument_type.co_directional_exposure_currency,
                ).pk
            )
        except Exception as e:
            _l.info(
                f"Could not set co_directional_exposure_currency, "
                f"fallback to default {repr(e)}"
            )
            instrument_object["co_directional_exposure_currency"] = (
                ecosystem_default.currency.pk
            )

        try:
            instrument_object["counter_directional_exposure_currency"] = (
                Currency.objects.get(
                    master_user=instrument_type.master_user,
                    user_code=instrument_type.counter_directional_exposure_currency,
                ).pk
            )
        except Exception as e:
            _l.info(
                f"Could not set counter_directional_exposure_currency, "
                f"fallback to default {repr(e)}"
            )
            instrument_object["counter_directional_exposure_currency"] = (
                ecosystem_default.currency.pk
            )

        # Set attributes
        instrument_object["attributes"] = []

        content_type = ContentType.objects.get(
            app_label="instruments", model="instrument"
        )

        for attribute in instrument_type.instrument_attributes.all():
            attribute_type = GenericAttributeType.objects.get(
                master_user=instrument_type.master_user,
                content_type=content_type,
                user_code=attribute.attribute_type_user_code,
            )

            attr = {"attribute_type": attribute_type.id}

            if attribute.value_type == 10:
                attr["value_string"] = attribute.value_string

            elif attribute.value_type == 20:
                attr["value_float"] = attribute.value_float

            elif attribute.value_type == 30:
                try:
                    item = GenericClassifier.objects.get(
                        attribute_type__user_code=attribute.attribute_type_user_code,
                        name=attribute.value_classifier,
                    )

                    attr["classifier"] = item.id
                    attr["classifier_object"] = {"id": item.id, "name": item.name}
                except Exception as e:
                    _l.error(f"set_defaults_from_instrument_type {e}")

                    attr["classifier"] = None

            elif attribute.value_type == 40:
                attr["value_date"] = attribute.value_date

            instrument_object["attributes"].append(attr)

        # Set Event Schedules

        instrument_object["event_schedules"] = []

        for instrument_type_event in instrument_type.events.all():
            event_schedule = {
                # 'name': instrument_type_event.name,
                "event_class": instrument_type_event.data["event_class"]
            }

            for item in instrument_type_event.data["items"]:
                # TODO add check for value type
                if "default_value" in item:
                    event_schedule[item["key"]] = item["default_value"]

            if "items2" in instrument_type_event.data:
                for item in instrument_type_event.data["items2"]:
                    if "default_value" in item:
                        event_schedule[item["key"]] = item["default_value"]

            event_schedule["is_auto_generated"] = True
            event_schedule["actions"] = []

            for instrument_type_action in instrument_type_event.data["actions"]:
                action = {
                    "transaction_type": instrument_type_action["transaction_type"],
                    "text": instrument_type_action["text"],
                    "is_sent_to_pending": instrument_type_action["is_sent_to_pending"],
                    "is_book_automatic": instrument_type_action["is_book_automatic"],
                }
                event_schedule["actions"].append(action)

            instrument_object["event_schedules"].append(event_schedule)

        # Set Accruals

        instrument_object["accrual_calculation_schedules"] = []

        for instrument_type_accrual in instrument_type.accruals.all():
            accrual = {
                item["key"]: item["default_value"]
                for item in instrument_type_accrual.data["items"]
                if "default_value" in item
            }
            instrument_object["accrual_calculation_schedules"].append(accrual)

        # Set Pricing Policy

        try:
            instrument_object["pricing_policies"] = []

            for it_pricing_policy in instrument_type.pricing_policies.all():
                pricing_policy = {
                    "pricing_policy_id": it_pricing_policy.pricing_policy.id,
                    "target_pricing_schema_user_code": it_pricing_policy.target_pricing_schema_user_code,
                    "options": it_pricing_policy.options,
                }

                instrument_object["pricing_policies"].append(pricing_policy)

        except Exception as e:
            _l.info(f"Can't set default pricing policy {e}")

        _l.info(f"instrument_object {instrument_object}")

        return instrument_object

    except Exception as e:
        _l.error(f"set_defaults_from_instrument_type {e}\n {traceback.format_exc()}")
        raise RuntimeError(f"Instrument Type is not configured correctly {e}")


def undate_events_for_instrument(instrument_object, arg1, maturity):
    # M
    expiration_event = instrument_object["event_schedules"][arg1]
    expiration_event["effective_date"] = maturity
    expiration_event["final_date"] = maturity


def set_events_for_instrument(instrument_object, data_object, instrument_type_obj):
    instrument_type = instrument_type_obj.user_code.lower()

    maturity = None

    if "maturity" in data_object:
        maturity = data_object["maturity"]

    if "maturity_date" in data_object:
        maturity = data_object["maturity_date"]

    if maturity:
        if (
            instrument_type
            in {
                "bonds",
                "convertible_bonds",
                "index_linked_bonds",
                "short_term_notes",
            }
            and len(instrument_object["event_schedules"]) == 2
        ):
            undate_events_for_instrument(instrument_object, 1, maturity)

        if instrument_type in {
            "bond_futures",
            "fx_forwards",
            "forwards",
            "futures",
            "commodity_futures",
            "call_options",
            "etfs",
            "funds",
            "index_futures",
            "index_options",
            "put_options",
            "tbills",
            "warrants",
        }:
            undate_events_for_instrument(instrument_object, 0, maturity)


def set_default_accrual(instrument_object, instrument_type_obj):
    instrument_type = instrument_type_obj.user_code.lower()
    if "bond" not in instrument_type:
        # if not bond, no accruals
        return

    if instrument_object["accrual_calculation_schedules"]:
        # instrument already has accruals
        return

    instrument_object["accrual_calculation_schedules"] = [
        {
            "accrual_start_date": None,
            "first_payment_date": None,
            "accrual_size": None,
            "accrual_calculation_model": 1,
            "accrual_calculation_model_object": {
                "id": 1,
                "name": "none",
                "short_name": "none",
                "user_code": "none",
                "public_name": "none",
                "notes": None,
            },
            "periodicity": Periodicity.ANNUALLY,
            "periodicity_object": {},
            "periodicity_n": 1,
            "notes": "default settings",
        }
    ]


def set_periodicity_period(source_data, accrual):
    p = int(source_data["accrual_calculation_schedules"][0]["periodicity_n"])

    accrual["periodicity_n"] = p
    try:
        accrual["periodicity"] = PERIODICITY_MAP[p]
    except KeyError:
        _l.error(f"invalid/unknown periodicity_n={p} set default=ANNUALLY")
        accrual["periodicity"] = Periodicity.ANNUALLY

    _l.info(f"periodicity {accrual['periodicity']}")


# Global method for create instrument object from Instrument Json data
def handler_instrument_object(
    source_data, instrument_type, master_user, ecosystem_default, attribute_types
):
    func = "handler_instrument_object"

    object_data = {"instrument_type": instrument_type.id}

    set_defaults_from_instrument_type(object_data, instrument_type, ecosystem_default)

    try:
        # TODO remove, when finmars.database.com will be deployed
        if isinstance(source_data["pricing_currency"], str):
            object_data["pricing_currency"] = Currency.objects.get(
                master_user=master_user,
                user_code=source_data["pricing_currency"],
            ).id
        else:
            object_data["pricing_currency"] = Currency.objects.get(
                master_user=master_user,
                user_code=source_data["pricing_currency"]["code"],
            ).id

    except Exception:
        object_data["pricing_currency"] = ecosystem_default.currency.id

    object_data["accrued_currency"] = object_data["pricing_currency"]
    object_data["co_directional_exposure_currency"] = object_data["pricing_currency"]
    object_data["counter_directional_exposure_currency"] = object_data[
        "pricing_currency"
    ]

    object_data["public_name"] = source_data["name"]
    object_data["user_code"] = source_data["user_code"]
    object_data["name"] = source_data["name"]
    object_data["short_name"] = source_data["short_name"]
    object_data["pricing_condition"] = source_data.get("pricing_condition")
    object_data["payment_size_detail"] = source_data.get("payment_size_detail")
    object_data["daily_pricing_model"] = source_data.get("daily_pricing_model")
    object_data["factor_schedules"] = source_data.get("factor_schedules", [])
    object_data["identifier"] = source_data.get("identifier", {})

    try:
        object_data["payment_size_detail"] = PaymentSizeDetail.objects.get(
            user_code=source_data["payment_size_detail"]
        ).id
    except Exception:
        object_data["payment_size_detail"] = (
            ecosystem_default.payment_size_detail.id
            if ecosystem_default.payment_size_detail
            else PaymentSizeDetail.DEFAULT
        )

    if "maturity_price" in source_data:
        try:
            object_data["maturity_price"] = float(source_data["maturity_price"])
        except Exception as e:
            _l.error(f"{func} Could not set maturity price {repr(e)}")

    if "maturity" in source_data and source_data["maturity"] != "":
        object_data["maturity_date"] = source_data["maturity"]

    elif "maturity_date" in source_data and source_data["maturity_date"] != "":
        if source_data["maturity_date"] in ("null", "9999-00-00"):
            object_data["maturity_date"] = None
        else:
            object_data["maturity_date"] = source_data["maturity_date"]
    else:
        object_data["maturity_date"] = None

    if (
        "country" in source_data
        and source_data["country"]
        and source_data["country"].get("alpha_3")
    ):
        try:
            country = Country.objects.get(alpha_3=source_data["country"]["alpha_3"])
            object_data["country"] = country.id

        except Country.DoesNotExist:
            _l.error(f"{func} no such country {source_data['country']['alpha_3']}")

    if "registration_date" in source_data and source_data["registration_date"] not in (
        "null",
        "",
    ):
        object_data["registration_date"] = source_data["registration_date"]
    else:
        object_data["registration_date"] = None

    try:
        if "sector" in source_data:
            sector_attribute = GenericAttributeType.objects.get(user_code="sector")

            attribute = {}
            exist = False

            for attribute in object_data["attributes"]:
                if attribute["attribute_type"] == sector_attribute.id:
                    exist = True
                    attribute["value_string"] = source_data["sector"]

            if not exist:
                attribute["attribute_type"] = sector_attribute.id
                attribute["value_string"] = source_data["sector"]

                object_data["attributes"].append(attribute)

    except Exception as e:
        _l.error(f"{func} Could not set sector {repr(e)}")

    _tmp_attributes_dict = {
        item["attribute_type"]: item for item in object_data["attributes"]
    }
    try:
        if "attributes" in source_data and isinstance(source_data["attributes"], dict):
            for attribute_type in attribute_types:
                lower_user_code = attribute_type.user_code.lower()

                for key, value in source_data["attributes"].items():
                    _l_key = key.lower()

                    if _l_key == lower_user_code:
                        attribute = {"attribute_type": attribute_type.id}

                        if attribute_type.value_type == 10:
                            attribute["value_string"] = value

                        elif attribute_type.value_type == 20:
                            attribute["value_float"] = value

                        elif attribute_type.value_type == 30:
                            try:
                                classifier = GenericClassifier.objects.get(
                                    attribute_type=attribute_type, name=value
                                )

                                attribute["classifier"] = classifier.id

                            except Exception:
                                attribute["classifier"] = None

                        elif attribute_type.value_type == 40:
                            attribute["value_date"] = value

                        _tmp_attributes_dict[attribute["attribute_type"]] = attribute
    except Exception as e:
        _l.error(
            f"{func} Could not set attributes from finmars database. Error {repr(e)}\n"
            f"{traceback.format_exc()}"
        )

    object_data["attributes"] = []

    for value in _tmp_attributes_dict.values():
        object_data["attributes"].append(value)

    object_data["master_user"] = master_user.id
    object_data["manual_pricing_formulas"] = []

    set_events_for_instrument(object_data, source_data, instrument_type)

    if (
        "accrual_calculation_schedules" in source_data
        and source_data["accrual_calculation_schedules"]
    ):
        _l.info("Setting up accrual schedules. Overwrite Existing")
        object_data["accrual_calculation_schedules"] = []
        for accrual in source_data["accrual_calculation_schedules"]:
            accrual.pop("id")  # remove id of finmars_database accrual schedule object
            set_periodicity_period(source_data, accrual)
            object_data["accrual_calculation_schedules"].append(accrual)
    else:
        set_default_accrual(object_data, instrument_type)

    if "accrual_events" in source_data and source_data["accrual_events"]:
        _l.info("Setting up accrual events. Overwrite Existing")
        object_data["accrual_events"] = []
        for accrual in source_data["accrual_events"]:
            accrual.pop("id", None)  # remove id of accrual object
            accrual.pop("instrument", None)  # remove instrument.id of accrual object
            accrual.pop("source", None)  # remove source.id of accrual object
            object_data["accrual_events"].append(accrual)

    if "name" not in object_data and "user_code" in object_data:
        object_data["name"] = object_data["user_code"]

    if "short_name" not in object_data and "user_code" in object_data:
        object_data["short_name"] = object_data["user_code"]

    _l.info(f"{func} instrument={source_data['user_code']} object_data={object_data}")

    return object_data


class SimpleImportProcess:
    def __init__(self, task_id, procedure_instance_id=None):
        self.task = CeleryTask.objects.get(pk=task_id)
        self.parent_task = self.task.parent

        _l.info(f"SimpleImportProcess.Task {self.task}. init")

        self.task.status = CeleryTask.STATUS_PENDING
        self.task.save()

        self.procedure_instance = None
        if procedure_instance_id:
            self.procedure_instance = RequestDataFileProcedureInstance.objects.get(
                id=procedure_instance_id
            )

            _l.info(
                f"SimpleImportProcess.Task {self.task}. init "
                f"procedure_instance {self.procedure_instance}"
            )

        self.member = self.task.member
        self.master_user = self.task.master_user
        self.proxy_user = ProxyUser(self.member, self.master_user)
        self.proxy_request = ProxyRequest(self.proxy_user)

        if self.task.options_object.get("scheme_id", None):
            self.scheme = CsvImportScheme.objects.get(
                pk=self.task.options_object["scheme_id"]
            )
        elif self.task.options_object.get("scheme_user_code", None):
            self.scheme = CsvImportScheme.objects.get(
                user_code=self.task.options_object["scheme_user_code"]
            )
        else:
            raise RuntimeError(
                f"Import Scheme {self.task.options_object['scheme_user_code']} "
                f"was not found"
            )

        self.execution_context = self.task.options_object["execution_context"]
        self.file_path = self.task.options_object["file_path"]
        self.ecosystem_default = EcosystemDefault.cache.get_cache(
            master_user_pk=self.master_user.pk
        )

        self.result = SimpleImportResult()
        self.result.task = self.task
        self.result.scheme = self.scheme

        self.process_type = ProcessType.CSV

        self.find_process_type()
        # `self.attribute_types` are set inside `self.get_attribute_types()`
        self.get_attribute_types()

        self.file_items = []  # items from provider  (json, csv, excel)
        self.raw_items = []  # items from provider  (json, csv, excel)
        self.conversion_items = []  # items with applied converions
        self.preprocessed_items = []  # items with calculated variables applied
        self.items = []  # result items that will be passed to TransactionTypeProcess

        self.context = {
            "master_user": self.master_user,
            "member": self.member,
            "request": self.proxy_request,
        }

        import_system_message_performed_by = self.member.username
        import_system_message_title = "Simple import (start)"
        if (
            self.execution_context
            and self.execution_context["started_by"] == "procedure"
        ):
            import_system_message_performed_by = "System"
            import_system_message_title = "Simple import from broker (start)"

        send_system_message(
            master_user=self.master_user,
            performed_by=import_system_message_performed_by,
            section="import",
            type="success",
            title=import_system_message_title,
            description=(
                f"{self.member.username} started import with scheme {self.scheme.name}"
            ),
        )

    def get_result_stats(self):
        result_stats = {
            "total": self.result.total_rows,
            "error": 0,
            "success": 0,
            "skip": 0,
        }
        self.preprocess()
        self.process()
        for result_item in self.result.items:
            if result_item.status == "error":
                result_stats["error"] += 1
            elif result_item.status == "success":
                result_stats["success"] += 1
            elif "skip" in result_item.status:
                result_stats["skip"] += 1
        return result_stats

    def generate_file_report(self):
        _l.info(
            f"SimpleImportProcess.generate_file_report "
            f"error_handler {self.scheme.error_handler} "
            f"missing_data_handler {self.scheme.missing_data_handler}"
        )

        result = [
            "Type, Simple Import",
            f"Scheme, {self.scheme.user_code}",
            f"Error handle, {self.scheme.error_handler}",
        ]

        if self.result.file_name:
            result.append(f"Filename, {self.result.file_name}")

        result.append(
            f"Import Rules - if object is not found {self.scheme.missing_data_handler}"
        )

        success_rows_count = 0
        error_rows_count = 0
        skip_rows_count = 0

        for result_item in self.result.items:
            if result_item.status == "error":
                error_rows_count += 1

            elif result_item.status == "success":
                success_rows_count += 1

            if "skip" in result_item.status:
                skip_rows_count += 1

        result.extend(
            (
                f"Rows total, {self.result.total_rows}",
                f"Rows success import, {success_rows_count}",
                f"Rows fail import, {error_rows_count}",
                f"Rows skipped import, {skip_rows_count}",
            )
        )
        columns = ["Row Number", "Status", "Message"]

        column_row_list = [f'"{str(item)}"' for item in columns]
        column_row = ",".join(column_row_list)

        result.append(column_row)

        for result_item in self.result.items:
            content = [
                str(result_item.row_number),
                result_item.status,
            ]

            if result_item.error_message:
                content.append(result_item.error_message)
            elif result_item.message:
                content.append(result_item.message)
            else:
                content.append("")

            content_row_list = [f'"{str(item)}"' for item in content]
            content_row = ",".join(content_row_list)

            result.append(content_row)

        result = "\n".join(result)

        current_date_time = now().strftime("%Y-%m-%d-%H-%M")

        file_name = f"file_report_{current_date_time}_task_{self.task.id}.csv"

        file_report = FileReport()

        _l.info("SimpleImportProcess.generate_file_report uploading file")

        file_report.upload_file(
            file_name=file_name, text=result, master_user=self.master_user
        )
        file_report.master_user = self.master_user
        file_report.name = (
            f"Simple Import {current_date_time} (Task {self.task.id}).csv"
        )
        file_report.file_name = file_name
        file_report.type = "simple_import.import"
        file_report.notes = "System File"
        file_report.content_type = "text/csv"

        file_report.save()

        _l.info(f"SimpleImportProcess.file_report {file_report} {file_report.file_url}")

        return file_report

    def generate_json_report(self):
        # _l.debug('self.result %s' % self.result.__dict__)

        # _l.debug('generate_json_report.result %s' % result)

        current_date_time = now().strftime("%Y-%m-%d-%H-%M")
        file_name = f"file_report_{current_date_time}_task_{self.task.id}.json"

        file_report = FileReport()

        _l.info("SimplemportProcess.generate_json_report uploading file")

        # file_report.upload_json_as_local_file(
        file_report.upload_file(
            file_name=file_name,
            # dict_to_json=self.task.result_object,
            text=json.dumps(self.task.result_object, indent=4, default=str),
            master_user=self.master_user,
        )
        file_report.master_user = self.master_user
        file_report.name = (
            f"Simple Import {current_date_time} (Task {self.task.id}).json"
        )
        file_report.file_name = file_name
        file_report.type = "simple_import.import"
        file_report.notes = "System File"
        file_report.content_type = "application/json"

        file_report.save()

        _l.info(f"SimpleImportProcess.json_report {file_report} {file_report.file_url}")

        return file_report

    def get_attribute_types(self):
        attribute_types = []

        try:
            attribute_types = GenericAttributeType.objects.filter(
                master_user=self.master_user, content_type=self.scheme.content_type
            )

        except Exception as e:
            _l.error(f"Get attribute types exception {repr(e)}")

        self.attribute_types = attribute_types

    def find_process_type(self):
        if self.task.options_object and "items" in self.task.options_object:
            self.process_type = ProcessType.JSON
        elif ".json" in self.file_path:
            self.process_type = ProcessType.JSON
        elif ".xlsx" in self.file_path:
            self.process_type = ProcessType.EXCEL
        elif ".csv" in self.file_path:
            self.process_type = ProcessType.CSV

        _l.info(
            f"SimpleImportProcess.Task {self.task}. process_type {self.process_type}"
        )

    def get_verbose_result(self):
        imported_count = 0
        error_count = 0

        for item in self.result.items:
            if item.status == "error":
                error_count += 1
            else:
                imported_count += 1

        result = (
            f"Processed {len(self.items)} rows and successfully imported "
            f"{imported_count} items. Error rows {error_count}"
        )

        return result

    def fill_with_file_items(self):
        _l.info(
            f"SimpleImportProcess.Task {self.task}. fill_with_raw_items "
            f"INIT {self.process_type}"
        )

        try:
            if self.process_type == ProcessType.JSON:
                try:
                    _l.info("Trying to get json items from task object options")

                    self.file_items = self.task.options_object["items"]

                except Exception:
                    _l.info("Trying to get json items from file")

                    with storage.open(self.file_path, "rb") as f:
                        self.file_items = json.loads(f.read())

                if not isinstance(self.file_items, list):
                    raise ValueError(
                        f"File {self.file_path} of type json is not a List. "
                        f"Did you forget to wrap it into []?"
                    )

            elif self.process_type == ProcessType.CSV:
                _l.info(f"ProcessType.CSV self.file_path {self.file_path}")

                with storage.open(self.file_path, "rb") as f:
                    with NamedTemporaryFile() as tmpf:
                        for chunk in f.chunks():
                            tmpf.write(chunk)
                        tmpf.flush()

                        # TODO check encoding (maybe should be taken from scheme)
                        with open(
                            tmpf.name,
                            mode="rt",
                            encoding="utf_8_sig",
                            errors="ignore",
                        ) as cf:
                            # TODO check quotechar (maybe should be taken from scheme)
                            reader = csv.reader(
                                cf,
                                delimiter=self.scheme.delimiter,
                                quotechar='"',
                                strict=False,
                                skipinitialspace=True,
                            )

                            self.append_and_count_file_items(reader)

            elif self.process_type == ProcessType.EXCEL:
                with storage.open(self.file_path, "rb") as f:
                    with NamedTemporaryFile() as tmpf:
                        self.read_from_excel_file(f, tmpf)

            else:
                raise ValueError(
                    f"File {self.file_path} is of invalid type {self.process_type}. "
                    f"Import impossible"
                )

            self.result.total_rows = len(self.file_items)

            if self.result.total_rows == 0:
                raise ValueError(
                    f"File {self.file_path} has no items. Nothing to import"
                )
            if self.result.total_rows > settings.MAX_ITEMS_IMPORT:
                raise ValueError(
                    f"File {self.file_path} has more than {settings.MAX_ITEMS_IMPORT} "
                    f"items. Import impossible"
                )

            _l.info(
                f"SimpleImportProcess.Task {self.task}. fill_with_raw_items "
                f"{self.process_type} DONE items {len(self.raw_items)}"
            )

        except Exception as e:
            err_msg = (
                f"SimpleImportProcess.Task {self.task}. fill_with_raw_items "
                f"{self.process_type} Exception {repr(e)} "
                f"Traceback {traceback.format_exc()}"
            )
            _l.error(err_msg)
            raise e

    def read_from_excel_file(self, f, tmpf):
        for chunk in f.chunks():
            tmpf.write(chunk)

        tmpf.flush()

        os.link(tmpf.name, f"{tmpf.name}.xlsx")

        _l.info(f"self.file_path {self.file_path}")
        _l.info(f"tmpf.name {tmpf.name}")

        wb = load_workbook(filename=f"{tmpf.name}.xlsx")

        ws = (
            wb[self.scheme.spreadsheet_active_tab_name]
            if (
                self.scheme.spreadsheet_active_tab_name
                and self.scheme.spreadsheet_active_tab_name in wb.sheetnames
            )
            else wb.active
        )
        reader = []

        if self.scheme.spreadsheet_start_cell == "A1":
            reader.extend([cell.value for cell in r] for r in ws.rows)
        else:
            start_cell_row_number = int(
                re.search(r"\d+", self.scheme.spreadsheet_start_cell)[0]
            )
            start_cell_letter = self.scheme.spreadsheet_start_cell.split(
                str(start_cell_row_number)
            )[0]

            start_cell_column_number = column_index_from_string(start_cell_letter)

            for row_number, r in enumerate(ws.rows, start=1):
                if row_number >= start_cell_row_number:
                    row_values = [
                        cell.value
                        for cell in r
                        if cell.column >= start_cell_column_number
                    ]
                    reader.append(row_values)

        self.append_and_count_file_items(reader)

    def append_and_count_file_items(self, reader):
        column_row = None

        for row_index, row in enumerate(reader):
            if row_index == 0:
                column_row = row

            else:
                file_item = {
                    column_row[column_index]: value
                    for column_index, value in enumerate(row)
                }
                self.file_items.append(file_item)

    def whole_file_preprocess(self):
        if self.scheme.data_preprocess_expression:
            names = {"data": self.file_items}

            try:
                # _l.info("whole_file_preprocess  names %s" % names)

                self.file_items = formula.safe_eval(
                    self.scheme.data_preprocess_expression,
                    names=names,
                    context=self.context,
                )

                # _l.info("whole_file_preprocess  self.raw_items %s" % self.raw_items)

            except Exception as e:
                _l.error(f"Could not execute preprocess expression. Error {e}")
                raise e

        _l.info(f"whole_file_preprocess.file_items {len(self.file_items)}")

        return self.file_items

    def fill_with_raw_items(self):
        _l.info(
            f"SimpleImportProcess.Task {self.task}. fill_with_raw_items "
            f"INIT {self.process_type}"
        )

        try:
            for file_item in self.file_items:
                item = {}
                for scheme_input in self.scheme.csv_fields.all():
                    try:
                        item[scheme_input.name] = file_item[scheme_input.column_name]
                    except Exception:
                        item[scheme_input.name] = None

                self.raw_items.append(item)

            _l.info(
                f"SimpleImportProcess.Task {self.task}. fill_with_raw_items "
                f"{self.process_type} DONE items {len(self.raw_items)}"
            )

        except Exception as e:
            _l.error(
                f"SimpleImportProcess.Task {self.task}. fill_with_raw_items "
                f"{self.process_type}  error {e} trace {traceback.format_exc()}"
            )
            raise e

    def apply_conversion_to_raw_items(self):
        for row_number, raw_item in enumerate(self.raw_items, start=1):
            conversion_item = SimpleImportConversionItem()
            conversion_item.file_inputs = self.file_items[row_number - 1]
            conversion_item.raw_inputs = raw_item
            conversion_item.conversion_inputs = {}
            conversion_item.row_number = row_number

            for scheme_input in self.scheme.csv_fields.all():
                try:
                    names = raw_item
                    conversion_item.conversion_inputs[scheme_input.name] = (
                        formula.safe_eval(
                            scheme_input.name_expr, names=names, context=self.context
                        )
                    )
                except Exception:
                    conversion_item.conversion_inputs[scheme_input.name] = None

            self.conversion_items.append(conversion_item)

    # We have formulas that lookup for rows
    # e.g. transaction_import.find_row
    # so it means, in first iterations we will got errors in that inputs
    def recursive_preprocess(self, deep=1, current_level=0):
        if len(self.preprocessed_items) == 0:
            for row_number, conversion_item in enumerate(
                self.conversion_items, start=1
            ):
                preprocess_item = SimpleImportProcessPreprocessItem()
                preprocess_item.file_inputs = conversion_item.file_inputs
                preprocess_item.raw_inputs = conversion_item.raw_inputs
                preprocess_item.conversion_inputs = conversion_item.conversion_inputs
                preprocess_item.row_number = row_number
                preprocess_item.inputs = {}

                self.preprocessed_items.append(preprocess_item)

        for preprocess_item in self.preprocessed_items:
            # CREATE SCHEME INPUTS
            for scheme_input in self.scheme.csv_fields.all():
                key_column_name = scheme_input.column_name

                try:
                    preprocess_item.inputs[scheme_input.name] = (
                        preprocess_item.conversion_inputs[scheme_input.name]
                    )

                except Exception as e:
                    preprocess_item.inputs[scheme_input.name] = None

                    if current_level == deep:
                        _l.error(
                            f"key_column_name {key_column_name} scheme_input.name "
                            f"{scheme_input.name} preprocess_item.raw_inputs "
                            f"{preprocess_item.conversion_inputs} Task {self.task}. "
                            f"recursive_preprocess init input {scheme_input} err {e}"
                        )

            # CREATE CALCULATED INPUTS
            for scheme_calculated_input in self.scheme.calculated_inputs.all():
                try:
                    names = preprocess_item.inputs

                    value = formula.safe_eval(
                        scheme_calculated_input.name_expr,
                        names=names,
                        context={
                            "master_user": self.master_user,
                            "member": self.member,
                            "request": self.proxy_request,
                            "transaction_import": {"items": self.preprocessed_items},
                        },
                    )

                    preprocess_item.inputs[scheme_calculated_input.name] = value

                except Exception as e:
                    preprocess_item.inputs[scheme_calculated_input.name] = None

                    if current_level == deep:
                        _l.error(
                            f"SimpleImportProcess.Task {self.task} recursive_preprocess"
                            f" calculated_input {scheme_calculated_input} err {e}"
                        )

        if current_level < deep:
            self.recursive_preprocess(deep, current_level + 1)

    def preprocess(self):
        _l.info(f"SimpleImportProcess.Task {self.task}. preprocess INIT")

        self.recursive_preprocess(deep=2)

        for preprocessed_item in self.preprocessed_items:
            item = SimpleImportProcessItem()
            item.row_number = preprocessed_item.row_number
            item.file_inputs = preprocessed_item.file_inputs
            item.raw_inputs = preprocessed_item.raw_inputs
            item.conversion_inputs = preprocessed_item.conversion_inputs
            item.inputs = preprocessed_item.inputs

            self.items.append(item)

        _l.info(
            f"SimpleImportProcess.Task {self.task}. preprocess "
            f"DONE items {len(self.preprocessed_items)}"
        )

    def fill_result_item_with_attributes(self, item, all_entity_fields_models=None):
        if not all_entity_fields_models:
            all_entity_fields_models = self.scheme.entity_fields.all()

        result = []
        for attribute_type in self.attribute_types:
            for entity_field in all_entity_fields_models:
                if (
                    entity_field.attribute_user_code
                    and entity_field.attribute_user_code == attribute_type.user_code
                ):
                    attribute = {"attribute_type": attribute_type.id}

                    if (
                        attribute_type.value_type == GenericAttributeType.STRING
                        and item.final_inputs[entity_field.attribute_user_code]
                    ):
                        attribute["value_string"] = item.final_inputs[
                            entity_field.attribute_user_code
                        ]

                    if attribute_type.value_type == GenericAttributeType.NUMBER and (
                        item.final_inputs[entity_field.attribute_user_code]
                        or item.final_inputs[entity_field.attribute_user_code] == 0
                    ):
                        attribute["value_float"] = item.final_inputs[
                            entity_field.attribute_user_code
                        ]

                    if (
                        attribute_type.value_type == GenericAttributeType.CLASSIFIER
                        and item.final_inputs[entity_field.attribute_user_code]
                    ):
                        try:
                            attribute["classifier"] = GenericClassifier.objects.get(
                                attribute_type=attribute_type,
                                name=item.final_inputs[
                                    entity_field.attribute_user_code
                                ],
                            ).id
                        except Exception as e:
                            _l.error(
                                f"fill_result_item_with_attributes classifier error - "
                                f"item {item} e {e}"
                            )

                            item.status = "error"
                            
                            if not item.error_message:
                                item.error_message = ""

                            item.error_message = f"{item.error_message}%s: %s, " % (
                                entity_field.attribute_user_code,
                                str(e),
                            )

                            attribute["classifier"] = None

                    if (
                        item.final_inputs[entity_field.attribute_user_code]
                        and attribute_type.value_type == GenericAttributeType.DATE
                    ):
                        attribute["value_date"] = item.final_inputs[
                            entity_field.attribute_user_code
                        ]

                    result.append(attribute)

        return result

    def overwrite_item_attributes(
        self, result_item, item, all_entity_fields_models=None
    ):
        if not all_entity_fields_models:
            all_entity_fields_models = self.scheme.entity_fields.all()

        for attribute in result_item["attributes"]:
            for entity_field in all_entity_fields_models:
                if entity_field.attribute_user_code and (
                    entity_field.attribute_user_code
                    == attribute["attribute_type_object"]["user_code"]
                ):
                    if (
                        attribute["attribute_type_object"]["value_type"]
                        == GenericAttributeType.STRING
                    ):
                        if item.final_inputs[entity_field.attribute_user_code]:
                            attribute["value_string"] = item.final_inputs[
                                entity_field.attribute_user_code
                            ]

                    elif (
                        attribute["attribute_type_object"]["value_type"]
                        == GenericAttributeType.NUMBER
                    ):
                        if item.final_inputs[entity_field.attribute_user_code]:
                            attribute["value_float"] = item.final_inputs[
                                entity_field.attribute_user_code
                            ]

                    elif (
                        attribute["attribute_type_object"]["value_type"]
                        == GenericAttributeType.CLASSIFIER
                    ):
                        if item.final_inputs[entity_field.attribute_user_code]:
                            try:
                                attribute["classifier"] = GenericClassifier.objects.get(
                                    attribute_type_id=attribute[
                                        "attribute_type_object"
                                    ]["id"],
                                    name=item.final_inputs[
                                        entity_field.attribute_user_code
                                    ],
                                ).id
                            except Exception as e:
                                _l.error(
                                    f"fill_result_item_with_attributes classifier error"
                                    f" - item {item} e {e}"
                                )

                                item.status = "error"

                                if not item.error_message:
                                    item.error_message = ""

                                item.error_message = f"{item.error_message}%s: %s, " % (
                                    entity_field.attribute_user_code,
                                    str(e),
                                )

                                attribute["classifier"] = None

                    elif (
                        attribute["attribute_type_object"]["value_type"]
                        == GenericAttributeType.DATE
                    ):
                        if item.final_inputs[entity_field.attribute_user_code]:
                            attribute["value_date"] = item.final_inputs[
                                entity_field.attribute_user_code
                            ]

    def __get_key_for_matching_model(
        self, key_model_user_code="-", pricing_policy__user_code="-", date="-"
    ):
        return f"{key_model_user_code}-{pricing_policy__user_code}-{date}"

    def __relation_fields_map_for_content_type(self):
        relation_fields_map = RELATION_FIELDS_MAP

        if self.scheme.content_type.model == "counterparty":
            relation_fields_map["group"] = CounterpartyGroup

        if self.scheme.content_type.model == "responsible":
            relation_fields_map["group"] = ResponsibleGroup

        if self.scheme.content_type.model == "strategy1":
            relation_fields_map["subgroup"] = Strategy1Subgroup

        if self.scheme.content_type.model == "strategy2":
            relation_fields_map["subgroup"] = Strategy2Subgroup

        if self.scheme.content_type.model == "strategy3":
            relation_fields_map["subgroup"] = Strategy3Subgroup

        return relation_fields_map

    def __get_relation_to_convert(
        self, item, relation_models_user_codes, all_entity_fields_models
    ):
        relation_fields_map = self.__relation_fields_map_for_content_type()

        for entity_field in all_entity_fields_models:
            key = entity_field.system_property_key

            if key in relation_fields_map and isinstance(item[key], str):
                if key not in relation_models_user_codes:
                    relation_models_user_codes[key] = []

                if item[key] not in relation_models_user_codes[key]:
                    relation_models_user_codes[key].append(item[key])

        return relation_models_user_codes

    def __get_relation_to_ids(
        self, relation_models_user_codes, all_entity_fields_models
    ):
        relation_fields_map = self.__relation_fields_map_for_content_type()

        relation_models_to_ids = {}

        for entity_field in all_entity_fields_models:
            key = entity_field.system_property_key

            relation_models_to_ids[key] = {}

            if key in relation_models_user_codes:
                relation_filter_result = relation_fields_map[key].objects.filter(
                    user_code__in=relation_models_user_codes[key]
                )

                if relation_filter_result:
                    for result_row in relation_filter_result:
                        relation_id = result_row.id
                        relation_user_code = result_row.user_code
                        relation_models_to_ids[key][relation_user_code] = result_row

        return relation_models_to_ids

    def replace_item_relations_by_ids(
        self, item, result_item, relation_models_to_ids, all_entity_fields_models
    ):
        relation_fields_map = self.__relation_fields_map_for_content_type()

        for entity_field in all_entity_fields_models:
            key = entity_field.system_property_key

            if (
                key in relation_fields_map
                and key in relation_models_to_ids
                and key in result_item
                and isinstance(result_item[key], str)
            ):
                if result_item[key] in relation_models_to_ids[key]:
                    result_item[key] = relation_models_to_ids[key][result_item[key]]
                else:
                    result_item[key] = None

        return result_item

    def convert_relation_to_ids(self, item, result_item, all_entity_fields_models=None):
        if not all_entity_fields_models:
            all_entity_fields_models = self.scheme.entity_fields.all()

        relation_fields_map = self.__relation_fields_map_for_content_type()

        for entity_field in all_entity_fields_models:
            key = entity_field.system_property_key

            if key in result_item:

                if key in relation_fields_map and isinstance(result_item[key], str):
                    try:
                        result_item[key] = (
                            relation_fields_map[key]
                            .objects.get(user_code=result_item[key])
                            .id
                        )
                    except Exception as e:
                        _l.warning(f"{key} {result_item[key]}: {e}")
                        result_item[key] = None

        return result_item

    def remove_nullable_attributes(self, result_item):
        for key, value in list(result_item.items()):
            if value is None:
                result_item.pop(key)

        return result_item

    def get_final_inputs(self, item, all_entity_fields_models=None):
        result = {}

        if not all_entity_fields_models:
            all_entity_fields_models = self.scheme.entity_fields.all()

        for entity_field in all_entity_fields_models:
            if entity_field.expression:
                try:
                    value = formula.safe_eval(
                        entity_field.expression, names=item.inputs, context=self.context
                    )

                    if entity_field.system_property_key:
                        result[entity_field.system_property_key] = value

                    elif entity_field.attribute_user_code:
                        result[entity_field.attribute_user_code] = value

                except Exception as e:
                    _l.warning(f"get_final_inputs.error {repr(e)}")

            elif entity_field.system_property_key:
                result[entity_field.system_property_key] = None

            elif entity_field.attribute_user_code:
                result[entity_field.attribute_user_code] = None

        return result

    def import_item(self, item: dict[str, Any]):
        from poms.instruments.handlers import InstrumentTypeProcess

        content_type_key = (
            f"{self.scheme.content_type.app_label}.{self.scheme.content_type.model}"
        )

        serializer_class = get_serializer(content_type_key)

        if not item.imported_items:
            item.imported_items = []

        try:
            item.final_inputs = self.get_final_inputs(item)

            result_item = {}
            if self.scheme.content_type.model == "instrument":
                instrument_type = InstrumentType.objects.get(
                    user_code=item.final_inputs["instrument_type"]
                )

                process = InstrumentTypeProcess(instrument_type=instrument_type)

                default_instrument_object = process.instrument

                default_instrument_object.update(result_item)

                result_item = default_instrument_object

            for key, value in item.final_inputs.items():
                if item.final_inputs[key] is not None:
                    result_item[key] = item.final_inputs[key]

            # TODO do not overwrite existing values from Instrument Type for Instrument
            result_item["attributes"] = self.fill_result_item_with_attributes(item)
            result_item = self.convert_relation_to_ids(item, result_item)
            result_item = self.remove_nullable_attributes(result_item)

            serializer = serializer_class(data=result_item, context=self.context)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            if self.scheme.item_post_process_script:
                # POST SUCCESS SCRIPT

                try:
                    formula.safe_eval(
                        self.scheme.item_post_process_script,
                        names=item.inputs,
                        context=self.context,
                    )

                except Exception as e:
                    item.status = "error"

                    if not item.error_message:
                        item.error_message = ""

                    item.error_message = (
                        f"{item.error_message} Post script error: {repr(e)}, "
                    )

            self.handle_successful_item_import(item, serializer)

        except Exception as e:
            if self.scheme.mode == "overwrite":
                try:
                    model = self.scheme.content_type.model_class()

                    if self.scheme.content_type.model == "pricehistory":
                        instance = model.objects.get(
                            key_model_user_code=item.final_inputs["instrument"],
                            pricing_policy__user_code=item.final_inputs[
                                "pricing_policy"
                            ],
                            date=item.final_inputs["date"],
                        )
                    elif self.scheme.content_type.model == "currencyhistory":
                        instance = model.objects.get(
                            currency__user_code=item.final_inputs["currency"],
                            pricing_policy__user_code=item.final_inputs[
                                "pricing_policy"
                            ],
                            date=item.final_inputs["date"],
                        )
                    elif self.scheme.content_type.model == "accrualcalculationschedule":
                        accrual_start_date = item.final_inputs["accrual_start_date"]
                        if not isinstance(accrual_start_date, date):
                            accrual_start_date = parse(str(accrual_start_date))

                        instance = model.objects.get(
                            instrument__user_code=item.final_inputs["instrument"],
                            accrual_start_date=accrual_start_date.strftime(
                                settings.API_DATE_FORMAT
                            ),
                        )
                    else:
                        instance = model.objects.get(
                            master_user=self.master_user,
                            user_code=item.final_inputs["user_code"],
                        )

                    item.final_inputs = self.get_final_inputs(item)

                    result_item = copy.copy(
                        serializer_class(instance=instance, context=self.context).data
                    )

                    for key, value in item.final_inputs.items():
                        if item.final_inputs[key] is not None:
                            result_item[key] = item.final_inputs[key]

                    if self.scheme.content_type.model not in [
                        "pricehistory",
                        "currencyhistory",
                        "accrualcalculationschedule",
                    ]:
                        self.overwrite_item_attributes(result_item, item)

                    result_item = self.convert_relation_to_ids(item, result_item)

                    serializer = serializer_class(
                        data=result_item,
                        instance=instance,
                        partial=True,
                        context=self.context,
                    )
                    serializer.is_valid(raise_exception=True)
                    serializer.save()

                    if self.scheme.item_post_process_script:
                        # POST SUCCESS SCRIPT
                        try:
                            formula.safe_eval(
                                self.scheme.item_post_process_script,
                                names=item.inputs,
                                context=self.context,
                            )
                        except Exception as e:
                            item.status = "error"

                            if not item.error_message:
                                item.error_message = ""

                            item.error_message = (
                                f"{item.error_message} Post script error: {repr(e)}, "
                            )

                    self.handle_successful_item_import(item, serializer)

                except Exception as e:
                    item.status = "error"

                    if not item.error_message:
                        item.error_message = ""

                    item.error_message = (
                        f"{item.error_message} ==== Overwrite Exception {e}"
                    )
                    _l.error(
                        f"import_item.overwrite model={self.scheme.content_type.model}"
                        f" final_inputs={item.final_inputs} error {e} traceback "
                        f"{traceback.format_exc()}"
                    )
            else:
                if "make a unique set" in str(e.__dict__):
                    item.status = "skip"
                    item.error_message = None

                else:
                    _l.info("traceback %s" % traceback.format_exc())
                
                    item.status = "error"
                    
                    if not item.error_message:
                        item.error_message = ""

                    item.error_message = (
                        f"{item.error_message} ==== Create Exception {e}"
                    )

    @staticmethod
    def calculate_pricehistory_null_fields(
        model: str, final_inputs: dict
    ) -> Optional[str]:
        """
        Calculates accrued_price & factor for PriceHistory if in the file
        their values are null, and update final_inputs dict
        """
        if model.lower() != "pricehistory" or not final_inputs:
            return None

        _l.info(f"calculate_null_fields: {model} final_inputs={final_inputs}")

        date_str = final_inputs.get("date")
        try:
            effective_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            err_msg = f"calculate_null_fields: invalid date_str={date_str}"
            _l.error(err_msg)
            return err_msg

        user_code = final_inputs.get("instrument")
        try:
            instrument = Instrument.objects.get(user_code=user_code)
        except Exception:
            err_msg = f"calculate_null_fields: no such instrument user_code={user_code}"
            _l.error(err_msg)
            return err_msg

        error_messages = []
        for key, value in final_inputs.items():
            if value is None:
                try:
                    if key == "accrued_price":
                        final_inputs[key] = instrument.get_accrued_price(
                            price_date=effective_date
                        )

                    elif key == "factor":
                        final_inputs[key] = instrument.get_factor(fdate=effective_date)

                except Exception as e:
                    final_inputs[key] = 1 if key == "factor" else 0
                    err_msg = f"calculate_null_fields: {key} {repr(e)}"
                    _l.error(err_msg)
                    error_messages.append(err_msg)

        if error_messages:
            return "; ".join(error_messages)

    def import_items_by_batch_indexes(
        self, batch_indexes, filter_for_async_functions_eval
    ):
        all_entity_fields_models = self.scheme.entity_fields.all()
        relation_models_user_codes = {}
        for item_index in batch_indexes:
            self.items[item_index].final_inputs = self.get_final_inputs(
                self.items[item_index], all_entity_fields_models
            )
            # dict for getting relation models at the next step
            relation_models_user_codes = self.__get_relation_to_convert(
                self.items[item_index].final_inputs,  # self.items[item_index],
                relation_models_user_codes,
                all_entity_fields_models,
            )

        # getting relation models
        relation_models_to_ids = self.__get_relation_to_ids(
            relation_models_user_codes, all_entity_fields_models
        )

        for item_index in batch_indexes:
            errors = None
            result_item = {
                key: self.items[item_index].final_inputs[key]
                for key, value in self.items[item_index].final_inputs.items()
                if self.items[item_index].final_inputs[key] is not None
            }
            result_item = self.replace_item_relations_by_ids(
                self.items[item_index],
                result_item,
                relation_models_to_ids,
                all_entity_fields_models,
            )
            result_item = self.remove_nullable_attributes(result_item)

            self.items[item_index].final_inputs = result_item
            if errors:
                self.items[item_index].status = "error"
                self.items[item_index].error_message += f"; {errors}"

        models_q_filter_list = []
        for item_index in batch_indexes:
            try:
                if self.scheme.content_type.model == "pricehistory":
                    models_q_filter_list.append(
                        Q(
                            instrument_id=self.items[item_index]
                            .final_inputs["instrument"]
                            .id,
                            pricing_policy_id=self.items[item_index]
                            .final_inputs["pricing_policy"]
                            .id,
                            date=parse_date(
                                self.items[item_index].final_inputs["date"]
                            ),
                        )
                    )
                else:  # "currencyhistory":
                    models_q_filter_list.append(
                        Q(
                            currency_id=self.items[item_index]
                            .final_inputs["currency"]
                            .id,
                            pricing_policy_id=self.items[item_index]
                            .final_inputs["pricing_policy"]
                            .id,
                            date=parse_date(
                                self.items[item_index].final_inputs["date"]
                            ),
                        )
                    )
            except Exception as e:
                self.items[item_index].status = "error"
                self.items[item_index].error_message = (
                    f"{self.items[item_index].error_message} "
                    f"Relation model error: {repr(e)}"
                )

        model = self.scheme.content_type.model_class()
        if models_q_filter_list:
            conditions = reduce(or_, models_q_filter_list)
            model_objects_for_update = model.objects.filter(conditions)
        else:
            model_objects_for_update = []

        # collecting keys of models for select models for update
        model_for_update_ids = {}
        if model_objects_for_update:
            for model_object in model_objects_for_update:
                if self.scheme.content_type.model == "pricehistory":
                    model_key_for_matching_model = self.__get_key_for_matching_model(
                        key_model_user_code=model_object.instrument_id,
                        pricing_policy__user_code=model_object.pricing_policy_id,
                        date=model_object.date,
                    )
                else:  # "currencyhistory"
                    model_key_for_matching_model = self.__get_key_for_matching_model(
                        key_model_user_code=model_object.currency_id,
                        pricing_policy__user_code=model_object.pricing_policy_id,
                        date=model_object.date,
                    )
                model_for_update_ids[model_key_for_matching_model] = model_object

        # dict for filtering models by key
        models_for_bulk_insert = {}
        models_for_bulk_update = {}
        for item_index in batch_indexes:
            # skip error status items
            if self.items[item_index].status == "error":
                continue

            if self.scheme.content_type.model == "pricehistory":
                item_key_for_matching_model = self.__get_key_for_matching_model(
                    key_model_user_code=self.items[item_index]
                    .final_inputs["instrument"]
                    .id,
                    pricing_policy__user_code=self.items[item_index]
                    .final_inputs["pricing_policy"]
                    .id,
                    date=self.items[item_index].final_inputs["date"],
                )
            else:  # "currencyhistory"
                item_key_for_matching_model = self.__get_key_for_matching_model(
                    key_model_user_code=self.items[item_index]
                    .final_inputs["currency"]
                    .id,
                    pricing_policy__user_code=self.items[item_index]
                    .final_inputs["pricing_policy"]
                    .id,
                    date=self.items[item_index].final_inputs["date"],
                )

            if item_key_for_matching_model in model_for_update_ids:
                # if overwrite allowed
                if self.scheme.mode == "overwrite":
                    # updating dict from model by data from import file
                    for key, value in self.items[item_index].final_inputs.items():
                        if self.items[item_index].final_inputs[key] is not None:
                            setattr(
                                model_for_update_ids[item_key_for_matching_model],
                                key,
                                value,
                            )

                    try:
                        # models_for_bulk_update.append(model_for_update_ids[item_key_for_matching_model])
                        models_for_bulk_update[item_key_for_matching_model] = (
                            model_for_update_ids[item_key_for_matching_model]
                        )

                    except Exception as e:
                        self.items[item_index].status = "error"
                        self.items[item_index].error_message = (
                            f"{self.items[item_index].error_message}==== "
                            f"Overwrite Exception {e}"
                        )

                else:
                    self.items[item_index].status = "skip"
                    self.items[item_index].error_message = None

                    # self.items[item_index].status = "error"
                    # self.items[item_index].error_message = (
                    #     f"{self.items[item_index].error_message}"
                    #     f"====  Overwrite disabled"
                    # )
            else:
                result_item = {
                    key: self.items[item_index].final_inputs[key]
                    for key, value in self.items[item_index].final_inputs.items()
                    if self.items[item_index].final_inputs[key] is not None
                }
                result_item = self.replace_item_relations_by_ids(
                    self.items[item_index],
                    result_item,
                    relation_models_to_ids,
                    all_entity_fields_models,
                )
                result_item = self.remove_nullable_attributes(result_item)
                try:
                    # models_for_bulk_insert.append(model(**result_item))
                    models_for_bulk_insert[item_key_for_matching_model] = model(
                        **result_item
                    )

                    if self.scheme.content_type.model == "pricehistory":
                        filter_for_async_functions_eval.append(
                            {
                                "instrument_id": self.items[item_index]
                                .final_inputs["instrument"]
                                .id,
                                "pricing_policy_id": self.items[item_index]
                                .final_inputs["pricing_policy"]
                                .id,
                                "date": self.items[item_index].final_inputs["date"],
                            }
                        )
                    else:  # "currencyhistory"
                        filter_for_async_functions_eval.append(
                            {
                                "currency_id": self.items[item_index]
                                .final_inputs["currency"]
                                .id,
                                "pricing_policy_id": self.items[item_index]
                                .final_inputs["pricing_policy"]
                                .id,
                                "date": self.items[item_index].final_inputs["date"],
                            }
                        )

                except Exception as e:
                    _l.info(
                        f"{self.items[item_index].error_message}"
                        f"==== Create Exception dcsd {e}"
                    )
                    self.items[item_index].status = "error"
                    self.items[item_index].error_message = (
                        f"{self.items[item_index].error_message}"
                        f"==== Create Exception {e}"
                    )

        # mass inserting models
        batch_rows_count = 0
        if models_for_bulk_insert:
            # Attention! bulk_create doesn't use save() method of the model
            model.objects.bulk_create(models_for_bulk_insert.values())
            batch_rows_count = batch_rows_count + len(models_for_bulk_insert.values())
            _l.info(
                f"SimpleImportProcess.Task bulk_insert count. "
                f"{len(models_for_bulk_insert.values())} "
            )

        if models_for_bulk_update:
            for model_obj in models_for_bulk_update.values():
                model_obj.save()
            # todo: by bulk update
            # model.objects.bulk_update(models_for_bulk_update)
            batch_rows_count = batch_rows_count + len(models_for_bulk_update.values())
            _l.info(
                f"SimpleImportProcess.Task bulk_update count. "
                f"{len(models_for_bulk_update.values())} "
            )

        _l.info(
            f"SimpleImportProcess.Task filter_for_async_functions_eval count."
            f" {len(filter_for_async_functions_eval)} "
        )

        self.handle_successful_items_by_batch_import(batch_indexes)
        return batch_rows_count

    def handle_successful_items_by_batch_import(self, batch_indexes: list[int]):
        for item_index in batch_indexes:

            self.items[item_index].status = "success"
            self.items[item_index].message = f"Item Imported {self.scheme.content_type.model}"

    def handle_successful_item_import(self, item, serializer):
        item.status = "success"
        item.message = f"Item Imported {serializer.instance}"

        trn = SimpleImportImportedItem(
            id=serializer.instance.id, user_code=str(serializer.instance)
        )

        item.imported_items.append(trn)

    def process_items(self):
        _l.info(f"SimpleImportProcess.Task {self.task}. process_items INIT")

        for item in self.items:
            try:
                _l.info(
                    f"SimpleImportProcess.Task {self.task}. ========= process row "
                    f"{item.row_number}/{self.result.total_rows} ========"
                )

                if self.scheme.filter_expr:
                    # expr = Expression.parseString("a == 1 and b == 2")
                    # expr = Expression.parseString(self.scheme.filter_expr)

                    success = bool(
                        formula.safe_eval(
                            self.scheme.filter_expr,
                            names=item.inputs,
                            context=self.context,
                        )
                    )

                    if not success:
                        item.status = "skip"
                        item.message = "Skipped due filter"
                        _l.info(
                            f"SimpleImportProcess.Task {self.task}. Row skipped "
                            f"due filter {item.row_number}"
                        )
                        continue

                self.import_item(item)

                self.result.processed_rows += 1

                self.task.update_progress(
                    {
                        "current": self.result.processed_rows,
                        "total": len(self.items),
                        "percent": round(
                            self.result.processed_rows / (len(self.items) / 100)
                        ),
                        "description": f"Row {self.result.processed_rows} processed",
                    }
                )

            except Exception as e:
                item.status = "error"
                item.message = f"item.row_number {item.row_number} error {repr(e)}"

                _l.error(
                    f"SimpleImportProcess.Task {self.task}.  ========= process row "
                    f"{str(item.row_number)} ======== Exception {e} ====== "
                    f"Traceback {traceback.format_exc()}"
                )

        self.result.items = self.items

        _l.info(f"SimpleImportProcess.Task {self.task}. process_items DONE")

    def process_items_batches(self):
        _l.info(f"SimpleImportProcess.Task {self.task}. process_items_batches INIT")
        # mb aren't needed
        self.result.processed_rows = 0
        items_per_batch = 100
        batch_indexes = []
        item_index = 0

        filter_for_async_functions_eval = []

        while item_index < len(self.items):
            success = True
            if self.scheme.filter_expr:
                try:
                    success = bool(
                        formula.safe_eval(
                            self.scheme.filter_expr,
                            names=self.items[item_index].inputs,
                            context=self.context,
                        )
                    )

                    if not success:
                        self.items[item_index].status = "skip"
                        self.items[item_index].message = "Skipped due filter"
                        _l.info(
                            f"SimpleImportProcess.Task {self.task}. Row skipped "
                            f"due filter {self.items[item_index].row_number}"
                        )

                except Exception as e:
                    success = False
                    self.items[item_index].status = "error"
                    self.items[item_index].message = (
                        f"item.row_number {self.items[item_index].row_number} "
                        f"error {repr(e)}"
                    )
                    _l.error(
                        f"SimpleImportProcess.Task {self.task}.  ========= process row "
                        f"{str(self.items[item_index].row_number)} ======== Exception "
                        f"{repr(e)} ====== Trace {traceback.format_exc()}"
                    )

            if success:
                batch_indexes.append(item_index)

            # increment while index
            item_index = item_index + 1

            if len(batch_indexes) >= items_per_batch or item_index >= len(self.items):
                batche_rows_count = self.import_items_by_batch_indexes(
                    batch_indexes, filter_for_async_functions_eval
                )
                self.result.processed_rows = (
                    self.result.processed_rows + batche_rows_count
                )
                batch_indexes = []

                self.task.update_progress(
                    {
                        "current": self.result.processed_rows,
                        "total": len(self.items),
                        "percent": round(
                            self.result.processed_rows / (len(self.items) / 100)
                        ),
                        "description": f"Row {self.result.processed_rows} processed",
                    }
                )

        self.result.items = self.items

        _l.info(f"SimpleImportProcess.Task {self.task}. process_items_batches DONE")

    def process(self):
        error_flag = False

        try:
            if self.scheme.content_type.model in ("pricehistory", "currencyhistory"):
                self.process_items_batches()
            else:
                self.process_items()

        except Exception as e:
            _l.error(
                f"SimpleImportProcess.Task {self.task}.process "
                f"Exception {repr(e)} Traceback {traceback.format_exc()}"
            )

            error_flag = True

            self.result.error_message = f"General Import Error. Exception {repr(e)}"

            if (
                self.execution_context
                and self.execution_context["started_by"] == "procedure"
            ):
                send_system_message(
                    master_user=self.master_user,
                    performed_by="System",
                    description=f"Can't process file. Exception {e}",
                )

        finally:
            self.task.result_object = SimpleImportResultSerializer(
                instance=self.result, context=self.context
            ).data

            # _l.info(f"self.task.result_object {self.task.result_object}")

            self.result.reports = []
            self.result.reports.append(self.generate_file_report())
            self.result.reports.append(self.generate_json_report())
            self.task.save()

            error_rows_count = sum(
                result_item.status == "error" for result_item in self.result.items
            )
            if error_rows_count:
                # Ignore item errors https://finmars2018.atlassian.net/browse/FN-2318
                # error_flag = True

                send_system_message(
                    master_user=self.master_user,
                    action_status="required",
                    type="warning",
                    title=f"Simple Import Partially Failed. Task id: {self.task.id}",
                    description=f"Error rows {error_rows_count}/{len(self.result.items)}",
                )

            system_message_description = (
                f"New items created (Import scheme - {str(self.scheme.name)}) -"
                f" {len(self.items)}"
            )

            import_system_message_title = "Simple import (finished)"

            system_message_performed_by = self.member.username

            system_message_title = "New Items (import from file)"
            if self.process_type == ProcessType.JSON and (
                self.execution_context
                and self.execution_context["started_by"] == "procedure"
            ):
                system_message_title = "New itmes (import from broker)"
                system_message_performed_by = "System"
                import_system_message_title = "Simple import from broker (finished)"

            send_system_message(
                master_user=self.master_user,
                performed_by=system_message_performed_by,
                section="import",
                type="error" if error_flag else "success",
                title="Import Finished. Prices Recalculation Required",
                description=(
                    "Please, run schedule or execute procedures to calculate portfolio "
                    "prices and nav history"
                ),
            )

        send_system_message(
            master_user=self.master_user,
            performed_by=system_message_performed_by,
            section="import",
            type="error" if error_flag else "success",
            title=import_system_message_title,
            attachments=[self.result.reports[0].id, self.result.reports[1].id],
        )

        send_system_message(
            master_user=self.master_user,
            performed_by=system_message_performed_by,
            section="import",
            type="error" if error_flag else "success",
            title=system_message_title,
            description=system_message_description,
        )

        if self.procedure_instance and self.procedure_instance.schedule_instance:
            self.procedure_instance.schedule_instance.run_next_procedure()

        self.task.add_attachment(self.result.reports[0].id)
        self.task.add_attachment(self.result.reports[1].id)
        self.task.verbose_result = self.get_verbose_result()
        self.task.status = (
            CeleryTask.STATUS_ERROR if error_flag else CeleryTask.STATUS_DONE
        )
        self.task.mark_task_as_finished()
        self.task.save()


class SimpleImportFinalUpdatesProcess(object):
    def __init__(self, task_id, procedure_instance_id=None):
        self.task = CeleryTask.objects.get(pk=task_id)
        self.parent_task = self.task.parent

        _l.info(f"SimpleImportFinalUpdatesProcess.Task {self.task}. init")

        self.task.status = CeleryTask.STATUS_PENDING
        self.task.save()

        self.member = self.task.member
        self.master_user = self.task.master_user
        self.proxy_user = ProxyUser(self.member, self.master_user)
        self.proxy_request = ProxyRequest(self.proxy_user)

        self.scheme = CsvImportScheme.objects.get(
            pk=self.task.options_object["scheme_id"]
        )

        self.model = self.scheme.content_type.model_class()

        content_type_key = (
            f"{self.scheme.content_type.app_label}.{self.scheme.content_type.model}"
        )
        # получаем сериалайзер под выбранный тип импорта
        self.serializer_class = get_serializer(content_type_key)

        self.filter_for_async_functions_eval = self.task.options_object.get(
            "filter_for_async_functions_eval", []
        )

        self.context = {
            "master_user": self.master_user,
            "member": self.member,
            "request": self.proxy_request,
        }

    def process(self):
        error_flag = False

        total_models_for_update = 0
        success_models_updates_count = 0
        error_models_updates_count = 0
        filter_for_async_functions_eval = []

        if self.task.options_object.get("filter_for_async_functions_eval", []):
            filter_for_async_functions_eval = self.task.options_object.get(
                "filter_for_async_functions_eval"
            )
            total_models_for_update = len(filter_for_async_functions_eval)

            # обновляем пачками
            items_per_batche = 100
            models_q_filter_list = []
            item_index = 0

            while item_index < total_models_for_update:
                models_q_filter_list.append(
                    Q(**filter_for_async_functions_eval[item_index])
                )
                item_index = item_index + 1

                if (
                    len(models_q_filter_list) >= items_per_batche
                    or item_index >= total_models_for_update
                ):
                    conditions = reduce(or_, models_q_filter_list)
                    model_objects_for_update = self.model.objects.filter(conditions)

                    for model_object in model_objects_for_update:
                        try:
                            model_object.save()
                            success_models_updates_count = (
                                success_models_updates_count + 1
                            )
                        except Exception:
                            error_models_updates_count = error_models_updates_count + 1

                    self.task.update_progress(
                        {
                            "current": success_models_updates_count,
                            "total": len(filter_for_async_functions_eval),
                            "percent": round(
                                (
                                    success_models_updates_count
                                    / len(filter_for_async_functions_eval)
                                )
                                * 100
                            ),
                            "description": f"Row finalization {success_models_updates_count} processed",
                        }
                    )
                    models_q_filter_list = []

        self.task.result_object = {
            "total_models_for_update": total_models_for_update,
            "success_models_updates_count": success_models_updates_count,
            "error_models_updates_count": error_models_updates_count,
            "not_found_models": total_models_for_update
            - (success_models_updates_count + error_models_updates_count),
        }

        self.task.status = (
            CeleryTask.STATUS_ERROR if error_flag else CeleryTask.STATUS_DONE
        )
        self.task.mark_task_as_finished()
        self.task.save()

        send_system_message(
            master_user=self.master_user,
            performed_by=self.member.username,
            section="import",
            type="error" if error_flag else "success",
            title="Simple Import Final Updates Process (finished)",
            description=(
                f"Final items count {success_models_updates_count}"
                f"/{len(filter_for_async_functions_eval)}"
            ),
        )
