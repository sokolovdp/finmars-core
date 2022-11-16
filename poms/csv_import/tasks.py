import re
import uuid
import hashlib

from celery import shared_task, chord, current_task
from django.contrib.auth.models import Permission
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.utils.timezone import now
from openpyxl.utils import column_index_from_string

from poms.common.crypto.AESCipher import AESCipher
from poms.common.crypto.RSACipher import RSACipher
from poms.common.formula import safe_eval, ExpressionEvalError
from poms.common.utils import date_now
from poms.file_reports.models import FileReport
from poms.obj_perms.models import GenericObjectPermission
from poms.pricing.models import InstrumentPricingPolicy
from poms.system_messages.handlers import send_system_message

from poms.users.models import EcosystemDefault, Group
from django.apps import apps
from django.contrib.contenttypes.models import ContentType

from poms.integrations.models import CounterpartyMapping, AccountMapping, ResponsibleMapping, PortfolioMapping, \
    PortfolioClassifierMapping, AccountClassifierMapping, ResponsibleClassifierMapping, CounterpartyClassifierMapping, \
    PricingPolicyMapping, InstrumentMapping, CurrencyMapping, InstrumentTypeMapping, PaymentSizeDetailMapping, \
    DailyPricingModelMapping, InstrumentClassifierMapping, AccountTypeMapping, \
    Task, PricingConditionMapping, TransactionFileResult

from poms.portfolios.models import Portfolio
from poms.currencies.models import Currency
from poms.instruments.models import PricingPolicy, Instrument, InstrumentType, DailyPricingModel, PaymentSizeDetail, \
    PricingCondition, AccrualCalculationModel, Periodicity
from poms.counterparties.models import Counterparty, Responsible
from poms.accounts.models import Account, AccountType

from poms.obj_attrs.models import GenericAttributeType, GenericAttribute, GenericClassifier
from poms.common import formula
from django.core.exceptions import ValidationError as CoreValidationError
from rest_framework.exceptions import ValidationError

from poms_app import settings
from .filters import SchemeContentTypeFilter
from .handlers import handler_instrument_object
from .models import CsvDataImport, CsvImportScheme
from .serializers import CsvDataImportSerializer, CsvImportSchemeSerializer, CsvDataFileImport

from django.utils.translation import gettext_lazy
from logging import getLogger
from openpyxl import load_workbook

from poms.celery_tasks.models import CeleryTask
from poms.common.websockets import send_websocket_message

import traceback

from poms.instruments.handlers import InstrumentTypeProcess
from poms.procedures.models import RequestDataFileProcedureInstance
from poms.common.models import ProxyUser, ProxyRequest

_l = getLogger('poms.csv_import')

import csv

from poms.common.storage import get_storage
storage = get_storage()

from tempfile import NamedTemporaryFile


def generate_file_report(instance, master_user, scheme, type, name, procedure_instance):
    columns = ['Row number']

    columns = columns + instance.stats[0]['error_data']['columns']['imported_columns']
    columns = columns + instance.stats[0]['error_data']['columns']['data_matching']

    columns.append('Error Message')
    columns.append('Reaction')

    rows_content = []

    for errorRow in instance.stats:

        localResult = []

        localResult.append(errorRow['original_row_index'])

        localResult = localResult + errorRow['error_data']['data']['imported_columns']
        localResult = localResult + errorRow['error_data']['data']['data_matching']

        localResult.append(str(errorRow['error_message']))
        localResult.append(str(errorRow['error_reaction']))

        localResultWrapper = []

        for item in localResult:
            localResultWrapper.append('"' + str(item) + '"')

        rows_content.append(localResultWrapper)

    columns = list(map(str, columns))

    columnRow = ','.join(columns)

    result = []

    result.append('Type, ' + 'Simple Import')
    result.append('Error handle, ' + scheme.error_handler)
    # result.append('Filename, ' + instance.file.name)
    result.append('Mode, ' + scheme.mode)
    result.append('Import Rules - if object is not found, ' + scheme.missing_data_handler)
    # result.push('Entity, ' + vm.scheme.content_type)

    result.append('Rows total, ' + str(instance.total_rows - 1))

    rowsSuccessTotal = 0
    rowsSkippedCount = 0
    rowsFailedCount = 0

    # rowsSkippedCount = instance.stats.filter(function (item) {
    # return item.error_reaction === 'Skipped';
    # }).length
    #
    #  rowsFailedCount = instance.stats.filter(function (item) {
    # return item.error_reaction !== 'Skipped';
    # }).length

    for item in instance.stats:
        if item['error_reaction'] == 'Skipped':
            rowsSkippedCount = rowsSkippedCount + 1

    for item in instance.stats:
        if item['error_reaction'] != 'Skipped' and item['error_reaction'] != 'Success':
            rowsFailedCount = rowsFailedCount + 1

    if instance.error_handler == 'break':

        index = instance.stats[0]['original_row_index']

        rowsSuccessTotal = index - 1  # get row before error
        rowsSuccessTotal = rowsSuccessTotal - 1  # exclude headers

    else:
        rowsSuccessTotal = instance.total_rows - 1 - rowsFailedCount - rowsSkippedCount

    if rowsSuccessTotal < 0:
        rowsSuccessTotal = 0

    result.append('Rows success import, ' + str(rowsSuccessTotal))
    result.append('Rows omitted, ' + str(rowsSkippedCount))
    result.append('Rows fail import, ' + str(rowsFailedCount))

    result.append('\n')
    result.append(columnRow)

    for contentRow in rows_content:
        contentRowStr = list(map(str, contentRow))

        result.append(','.join(contentRowStr))

    result = '\n'.join(result)

    current_date_time = now().strftime("%Y-%m-%d-%H-%M")

    file_name = 'file_report_%s.csv' % current_date_time
    name = "%s %s" % (name, current_date_time)

    file_report = FileReport()

    if procedure_instance:
        file_name = 'file_report_%s_procedure_instance_%s.csv' % (current_date_time, str(procedure_instance.id))
        name = "%s %s Procedure Instance %s" % (name, current_date_time, str(procedure_instance.id))
    else:
        file_name = 'file_report_%s.csv' % current_date_time
        name = "%s %s" % (name, current_date_time)

    file_report.upload_file(file_name=file_name, text=result, master_user=master_user)
    file_report.master_user = master_user
    file_report.name = name
    file_report.file_name = file_name
    file_report.type = type
    file_report.notes = 'System File'
    file_report.content_type = 'text/csv'

    file_report.save()

    return file_report.pk


def generate_file_report_simple(instance, type, name):
    try:
        columns = ['Row number']

        columns.append('Message')

        rows_content = []

        for errorRow in instance.items:

            if errorRow['original_row_index'] != 0:

                localResult = []

                localResult.append(errorRow['original_row_index'])

                if errorRow['error_message']:
                    localResult.append(str(errorRow['error_message']))
                else:
                    localResult.append('OK')

                localResultWrapper = []

                for item in localResult:
                    localResultWrapper.append('"' + str(item) + '"')

                rows_content.append(localResultWrapper)

        columnRow = ','.join(columns)

        result = []

        result.append('Type, ' + type)
        # result.append('Filename, ' + instance.file.name)
        result.append('Mode, ' + instance.mode)
        # result.append('Import Rules - if object is not found, ' + instance.missing_data_handler)
        # result.push('Entity, ' + vm.scheme.content_type)

        result.append('Rows total, ' + str(instance.total_rows))

        rowsSuccessTotal = 0
        rowsSkippedCount = 0
        rowsFailedCount = 0

        # result.append('Rows success import, ' + str(rowsSuccessTotal))
        # result.append('Rows omitted, ' + str(rowsSkippedCount))
        # result.append('Rows fail import, ' + str(rowsFailedCount))

        result.append('\n')
        result.append(columnRow)

        for contentRow in rows_content:
            contentRowStr = list(map(str, contentRow))

            result.append(','.join(contentRowStr))

        result = '\n'.join(result)

        current_date_time = now().strftime("%Y-%m-%d-%H-%M")

        file_name = 'Unified Instrument Import %s.csv' % current_date_time

        file_report = FileReport()

        file_report.upload_file(file_name=file_name, text=result, master_user=instance.master_user)
        file_report.master_user = instance.master_user
        file_report.name = "%s %s" % (name, current_date_time) + '.csv'
        file_report.file_name = file_name
        file_report.type = type
        file_report.notes = 'System File'
        file_report.content_type = 'text/csv'

        file_report.save()

        return file_report.pk
    except Exception as e:
        _l.info("Generate file error occured %s" % e)
        _l.info(traceback.format_exc())
        return None


def get_row_data(row, csv_fields, scheme, first_row):
    csv_row_dict = {}

    if scheme.column_matcher == 'index':

        for csv_field in csv_fields:

            if csv_field.column - 1 < len(row):
                row_value = row[csv_field.column - 1]

                csv_row_dict[csv_field.name] = row_value

            else:

                csv_row_dict[csv_field.name] = ''

    if scheme.column_matcher == 'name':

        if type(row) is dict:

            for csv_field in csv_fields:

                if csv_field.column_name in row:
                    csv_row_dict[csv_field.name] = row[csv_field.column_name]
                else:
                    csv_row_dict[csv_field.name] = None

        else:

            index_map = {}
            index = 0
            for item in first_row:
                index_map[item] = index
                index = index + 1

            for key, value in index_map.items():

                for csv_field in csv_fields:
                    if key == csv_field.column_name:
                        csv_row_dict[csv_field.name] = row[value]

    return csv_row_dict


def get_row_data_converted(row, csv_fields, csv_row_dict_raw, context, conversion_errors):
    csv_row_dict = {}

    for csv_field in csv_fields:

        if csv_field.column - 1 < len(row):

            try:

                executed_expression = safe_eval(csv_field.name_expr, names=csv_row_dict_raw, context=context)

                csv_row_dict[csv_field.name] = executed_expression

            except (ExpressionEvalError, TypeError, Exception, KeyError):

                csv_row_dict[csv_field.name] = gettext_lazy('Invalid expression')

                error = {
                    'name': csv_field.name,
                    'value': gettext_lazy('Invalid expression')
                }

                conversion_errors.append(error)

        else:

            csv_row_dict[csv_field.name] = ''

    return csv_row_dict


def get_field_type(field):
    if field.system_property_key is not None:
        return 'system_attribute'
    else:
        return 'dynamic_attribute'


def get_item(scheme, result):
    Model = apps.get_model(app_label=scheme.content_type.app_label, model_name=scheme.content_type.model)

    item_result = None

    if scheme.content_type.model == 'pricehistory':

        try:

            item_result = Model.objects.get(instrument=result['instrument'],
                                            pricing_policy=result['pricing_policy'],
                                            date=result['date'])
        except Exception as e:

            # _l.debug('get_item pricehistory exception %s' % e)

            item_result = None


    elif scheme.content_type.model == 'currencyhistory':

        try:

            item_result = Model.objects.get(currency=result['currency'], pricing_policy=result['pricing_policy'],
                                            date=result['date'])

        except Exception as e:

            traceback.format_exc()

            # _l.debug('get_item currencyhistory exception %s' % e)

            item_result = None

    else:

        try:

            if 'user_code' in result:
                item_result = Model.objects.get(master_user_id=result['master_user'], user_code=result['user_code'])

        except Exception as e:

            # _l.debug('get_item entity exception %s' % e)

            item_result = None

    return item_result


def update_row_with_calculated_data(csv_row_dict, inputs, calculated_inputs):
    for i in calculated_inputs:

        try:
            value = formula.safe_eval(i.name_expr, names=inputs)
            csv_row_dict[i.name] = value

        except Exception:
            _l.debug('can\'t process calculated input: %s|%s', i.name, i.column, exc_info=True)
            csv_row_dict[i.name] = None

    return csv_row_dict


def process_csv_file(master_user,
                     scheme,
                     original_file,
                     file,
                     error_handler,
                     missing_data_handler,
                     classifier_handler,
                     context,
                     task_instance,
                     update_state,
                     mode,
                     process_result_handler,
                     member,
                     execution_context=None,
                     celery_task=None):
    csv_fields = scheme.csv_fields.all()
    entity_fields = scheme.entity_fields.all()
    calculated_inputs = list(scheme.calculated_inputs.all())

    errors = []
    results = []

    reader = []

    processed_row_index = 0

    if celery_task and celery_task.options_object and 'items' in celery_task.options_object:

        reader = celery_task.options_object['items']

    elif '.csv' in task_instance.filename or (execution_context and execution_context["started_by"] == 'procedure'):

        delimiter = task_instance.scheme.delimiter.encode('utf-8').decode('unicode_escape')

        reader = csv.reader(file, delimiter=delimiter, quotechar=task_instance.quotechar,
                            strict=False, skipinitialspace=True)

    elif '.xlsx' in task_instance.filename:
        _l.info('trying to parse excel %s ' % task_instance.filename)

        wb = load_workbook(filename=original_file)

        # ws = wb.active

        if task_instance.scheme.spreadsheet_active_tab_name and task_instance.scheme.spreadsheet_active_tab_name in wb.sheetnames:
            ws = wb[task_instance.scheme.spreadsheet_active_tab_name]
        else:
            ws = wb.active

        _l.info('ws %s' % ws)
        _l.info('task_instance.scheme.spreadsheet_start_cell %s' % task_instance.scheme.spreadsheet_start_cell)

        reader = []

        if task_instance.scheme.spreadsheet_start_cell == 'A1':

            for r in ws.rows:
                reader.append([cell.value for cell in r])

        else:

            start_cell_row_number = int(re.search(r'\d+', task_instance.scheme.spreadsheet_start_cell)[0])
            start_cell_letter = task_instance.scheme.spreadsheet_start_cell.split(str(start_cell_row_number))[0]

            start_cell_column_number = column_index_from_string(start_cell_letter)

            _l.info('start_cell_column_number %s ' % start_cell_column_number)
            _l.info('start_cell_row_number %s ' % start_cell_row_number)

            row_number = 1

            for r in ws.rows:

                row_values = []

                if row_number >= start_cell_row_number:

                    for cell in r:

                        _l.info('cell.column  %s' % cell.column)

                        if cell.column >= start_cell_column_number:
                            row_values.append(cell.value)

                    reader.append(row_values)

                row_number = row_number + 1

        _l.info('reader %s' % reader)

    first_row = None

    for row_index, row in enumerate(reader):

        if row_index == 0:
            first_row = row

        if row_index != 0 or (celery_task and celery_task.options_object and 'items' in celery_task.options_object):

            try:

                instance = {}
                instance['_row_index'] = row_index
                instance['_row'] = row

                if scheme.content_type.model != 'pricehistory' and scheme.content_type.model != 'currencyhistory':
                    instance['master_user'] = master_user
                    instance['attributes'] = []

                inputs_error = []
                executed_expressions = []

                csv_row_dict_raw = {}

                error_row = {
                    'level': 'info',
                    'error_message': '',
                    'original_row_index': row_index,
                    'inputs': csv_row_dict_raw,
                    'original_row': row,
                    'error_data': {
                        'columns': {
                            'imported_columns': [],
                            'converted_imported_columns': [],
                            'data_matching': []
                        },
                        'data': {
                            'imported_columns': [],
                            'converted_imported_columns': [],
                            'data_matching': []
                        }
                    },
                    'error_reaction': 'Success'
                }

                try:
                    csv_row_dict_raw = get_row_data(row, csv_fields, scheme, first_row)
                except Exception as e:
                    raise Exception("Can't get row data")

                # _l.info('csv_row_dict_raw %s ' % csv_row_dict_raw)

                executed_filter_expression = True

                if scheme.filter_expr:

                    try:
                        executed_filter_expression = safe_eval(scheme.filter_expr, names=csv_row_dict_raw, context={})
                    except (ExpressionEvalError, TypeError, Exception, KeyError):
                        raise Exception("Can evaluate filter expression")

                if executed_filter_expression:

                    error_row['inputs'] = csv_row_dict_raw

                    for key, value in csv_row_dict_raw.items():
                        error_row['error_data']['columns']['imported_columns'].append(key)
                        error_row['error_data']['data']['imported_columns'].append(value)

                    conversion_errors = []

                    csv_row_dict = get_row_data_converted(row, csv_fields, csv_row_dict_raw, {}, conversion_errors)

                    csv_row_dict = update_row_with_calculated_data(csv_row_dict, csv_fields, calculated_inputs)

                    for key, value in csv_row_dict.items():
                        error_row['error_data']['columns']['converted_imported_columns'].append(
                            key + ': Conversion Expression')
                        error_row['error_data']['data']['converted_imported_columns'].append(value)

                    if len(conversion_errors) > 0:

                        inputs_messages = []

                        for input_error in inputs_error:
                            message = '[{0}] (Imported column conversion expression, value: "{1}")'.format(
                                input_error['field'].name, input_error['reason'])

                            inputs_messages.append(message)

                        error_row['error_message'] = error_row['error_message'] + '\n' + '\n' + gettext_lazy(
                            'Can\'t process fields: %(messages)s') % {
                                                         'messages': ', '.join(str(m) for m in inputs_messages)
                                                     }

                        error_row['level'] = 'error'
                        error_row['error_reaction'] = 'Continue import'

                    mapping_map = {
                        'counterparties': CounterpartyMapping,
                        'responsibles': ResponsibleMapping,
                        'accounts': AccountMapping,
                        'portfolios': PortfolioMapping,
                        'pricing_policy': PricingPolicyMapping,
                        'instrument': InstrumentMapping,
                        'instrument_type': InstrumentTypeMapping,
                        'type': AccountTypeMapping,
                        'payment_size_detail': PaymentSizeDetailMapping,
                        'pricing_condition': PricingConditionMapping,
                        'currency': CurrencyMapping,
                        'pricing_currency': CurrencyMapping,
                        'accrued_currency': CurrencyMapping
                    }

                    relation_map = {
                        'counterparties': Counterparty,
                        'responsibles': Responsible,
                        'accounts': Account,
                        'portfolios': Portfolio,
                        'pricing_policy': PricingPolicy,
                        'instrument': Instrument,
                        'instrument_type': InstrumentType,
                        'type': AccountType,
                        'currency': Currency,
                        'pricing_currency': Currency,
                        'accrued_currency': Currency
                    }

                    classifier_mapping_map = {
                        'portfolio': PortfolioClassifierMapping,
                        'instrument': InstrumentClassifierMapping,
                        'account': AccountClassifierMapping,
                        'responsible': ResponsibleClassifierMapping,
                        'counterparty': CounterpartyClassifierMapping
                    }

                    instance_property_to_default_ecosystem_property = {
                        'pricing_currency': 'currency',
                        'accrued_currency': 'currency',
                        'type': 'account_type'
                    }

                    excluded_fields = ['price_download_scheme', 'daily_pricing_model']

                    for entity_field in entity_fields:

                        key = entity_field.system_property_key

                        if entity_field.expression != '' and key not in excluded_fields:

                            error_row['error_data']['columns']['data_matching'].append(entity_field.name)

                            if get_field_type(entity_field) == 'system_attribute':

                                executed_expression = None

                                try:
                                    # context=self.report.context

                                    if entity_field.use_default and scheme.content_type.model == 'instrument':
                                        instance[key] = None  # will be set from instrument type
                                    else:

                                        executed_expression = safe_eval(entity_field.expression, names=csv_row_dict,
                                                                        context={})

                                        executed_expressions.append(executed_expression)

                                        if key in mapping_map:

                                            try:

                                                instance[key] = mapping_map[key].objects.get(master_user=master_user,
                                                                                             value=executed_expression).content_object

                                            except (mapping_map[key].DoesNotExist, KeyError):

                                                try:

                                                    instance[key] = relation_map[key].objects.get(
                                                        master_user=master_user,
                                                        user_code=executed_expression)

                                                except (relation_map[key].DoesNotExist, KeyError):

                                                    if missing_data_handler == 'set_defaults':

                                                        ecosystem_default = EcosystemDefault.objects.get(
                                                            master_user=master_user)

                                                        eco_key = key

                                                        if key in instance_property_to_default_ecosystem_property:
                                                            eco_key = instance_property_to_default_ecosystem_property[
                                                                key]

                                                        if hasattr(ecosystem_default, eco_key):
                                                            instance[key] = getattr(ecosystem_default, eco_key)
                                                        else:
                                                            _l.debug("Can't set default value for %s" % key)
                                                    else:

                                                        inputs_error.append(
                                                            {"field": entity_field,
                                                             "reason": "Relation does not exists"}
                                                        )

                                        else:

                                            if key == 'maturity_date':
                                                instance[key] = str(executed_expression)
                                            else:
                                                instance[key] = executed_expression

                                except (ExpressionEvalError, TypeError, Exception, KeyError):

                                    if missing_data_handler == 'set_defaults':

                                        ecosystem_default = EcosystemDefault.objects.get(
                                            master_user=master_user)

                                        eco_key = key

                                        if key in instance_property_to_default_ecosystem_property:
                                            eco_key = instance_property_to_default_ecosystem_property[key]

                                        if hasattr(ecosystem_default, eco_key):
                                            instance[key] = getattr(ecosystem_default, eco_key)
                                        else:
                                            _l.debug("Can't set default value for %s" % key)

                                    else:

                                        _l.debug('ExpressionEvalError Appending Error %s key %s' % (
                                            ExpressionEvalError, key))

                                        instance[key] = None

                                        inputs_error.append(
                                            {"field": entity_field,
                                             "reason": "Invalid Expression"}
                                        )

                                        executed_expressions.append(gettext_lazy('Invalid expression'))

                                # _l.debug('executed_expression %s' % executed_expression)

                            if get_field_type(entity_field) == 'dynamic_attribute':

                                executed_attr = {}
                                executed_attr['dynamic_attribute_id'] = entity_field.dynamic_attribute_id

                                executed_expression = None

                                try:
                                    # context=self.report.context
                                    executed_expression = safe_eval(entity_field.expression, names=csv_row_dict,
                                                                    context={})

                                    executed_expressions.append(executed_expression)

                                except (ExpressionEvalError, TypeError, Exception, KeyError):

                                    # inputs_error.append(entity_field)
                                    inputs_error.append(
                                        {"field": entity_field,
                                         "reason": "Invalid Expression"}
                                    )

                                    executed_expressions.append(gettext_lazy('Invalid expression'))

                                attr_type = GenericAttributeType.objects.get(pk=executed_attr['dynamic_attribute_id'])

                                if attr_type.value_type == 30:

                                    if scheme.content_type.model in classifier_mapping_map:

                                        try:
                                            executed_attr['executed_expression'] = classifier_mapping_map[
                                                scheme.content_type.model].objects.get(
                                                master_user=master_user,
                                                value=executed_expression, attribute_type=attr_type).content_object

                                        except (
                                                classifier_mapping_map[scheme.content_type.model].DoesNotExist,
                                                KeyError):

                                            try:

                                                # _l.debug('Lookup by name in classifier')

                                                executed_attr['executed_expression'] = GenericClassifier.objects.get(
                                                    attribute_type=attr_type, name=executed_expression)

                                            except (GenericClassifier.DoesNotExist, KeyError):

                                                if classifier_handler == 'append':

                                                    classifier_obj = GenericClassifier.objects.create(
                                                        attribute_type=attr_type,
                                                        name=executed_expression)

                                                    executed_attr['executed_expression'] = classifier_obj

                                                else:

                                                    executed_attr['executed_expression'] = None

                                                    # inputs_error.append(entity_field)
                                                    #
                                                    # _l.debug('%s classifier mapping  does not exist' % scheme.content_type.model)
                                                    # _l.debug('expresion: %s ' % executed_expression)

                                else:

                                    executed_attr['executed_expression'] = executed_expression

                                instance['attributes'].append(executed_attr)

                    error_row['error_data']['data']['data_matching'] = executed_expressions

                    if inputs_error:

                        inputs_messages = []

                        for input_error in inputs_error:
                            message = '[{0}] ({1})'.format(input_error['field'].name, input_error['reason'])

                            inputs_messages.append(message)

                        error_row['level'] = 'error'
                        error_row['error_message'] = error_row['error_message'] + '\n' + '\n' + gettext_lazy(
                            'Can\'t process fields: %(inputs)s') % {
                                                         'inputs': ', '.join(str(m) for m in inputs_messages)}

                        error_row['error_reaction'] = 'Continue import'

                    else:

                        try:

                            instance, error_row = process_result_handler(instance, error_row, scheme, error_handler,
                                                                         mode,
                                                                         member, master_user)

                            results.append(instance)

                        except Exception as e:

                            _l.info('Result processing error %s ' % e)

                            traceback.format_exc()

                            raise Exception("Result processing error")

                else:

                    error_row['level'] = 'info'
                    error_row['error_message'] = 'Row was skipped'
                    error_row['error_reaction'] = 'Skipped'

            except Exception as e:

                error_row['error_message'] = error_row['error_message'] + gettext_lazy(
                    'Unhandled Error. %s' % e)

            finally:

                if row_index != 0:  # skip header from counting
                    processed_row_index = processed_row_index + 1

                errors.append(error_row)

                task_instance.processed_rows = processed_row_index

                _l.debug('task_instance.processed_rows: %s', task_instance.processed_rows)

                send_websocket_message(data={
                    'type': 'simple_import_status',
                    'payload': {'task_id': task_instance.task_id,
                                'state': Task.STATUS_PENDING,
                                'processed_rows': task_instance.processed_rows,
                                'total_rows': task_instance.total_rows,
                                'user_code': scheme.user_code,
                                'file_name': task_instance.filename}
                }, level="member",
                    context={"master_user": master_user, "member": member})

                # Deprecated
                update_state(task_id=task_instance.task_id, state=Task.STATUS_PENDING,
                             meta={'processed_rows': task_instance.processed_rows,
                                   'total_rows': task_instance.total_rows, 'user_code': scheme.user_code,
                                   'file_name': task_instance.filename})

                if error_handler == 'break' and error_row['level'] == 'error':
                    error_row['error_reaction'] = 'Break import'

                    return results, errors

    return results, errors


class ValidateHandler:

    def create_simple_instance(self, scheme, result):

        Model = apps.get_model(app_label=scheme.content_type.app_label, model_name=scheme.content_type.model)

        result_without_many_to_many = {}

        many_to_many_fields = ['counterparties', 'responsibles', 'accounts', 'portfolios']
        system_fields = ['_row_index', '_row']

        for key, value in result.items():

            if key != 'attributes':

                if key not in many_to_many_fields and key not in system_fields:
                    result_without_many_to_many[key] = value

        try:
            instance = Model(**result_without_many_to_many)
        except (ValidationError, IntegrityError):
            instance = None

        return instance

    def attributes_full_clean(self, instance, attributes, error_handler, error_row):

        attr_type_user_code = 'Unknown'

        for result_attr in attributes:

            attr_type_user_code = 'Unknown'

            try:

                attr_type = GenericAttributeType.objects.get(pk=result_attr['dynamic_attribute_id'])

                attr_type_user_code = attr_type.user_code

                if attr_type:

                    attribute = GenericAttribute(content_object=instance, attribute_type=attr_type)

                    # _l.debug('result_attr', result_attr)
                    # _l.debug('attribute', attribute)

                    if attr_type.value_type == 10:
                        attribute.value_string = str(result_attr['executed_expression'])
                    elif attr_type.value_type == 20:
                        attribute.value_float = float(result_attr['executed_expression'])
                    elif attr_type.value_type == 30:

                        attribute.classifier = result_attr['executed_expression']

                    elif attr_type.value_type == 40:
                        attribute.value_date = formula._parse_date(result_attr['executed_expression'])
                    else:
                        pass

                    attribute.object_id = 1  # To pass object id check

                    attribute.full_clean()

            except Exception as e:

                # _l.info("e %s" % e)

                error_row['error_reaction'] = 'Continue import'
                error_row['level'] = 'error'
                error_row['error_message'] = error_row['error_message'] + gettext_lazy(
                    'Validation error %(error)s ') % {
                                                 'error': 'Cannot create attribute. Attribute type %s, value %s. Exception: %s' % (
                                                     attr_type_user_code, result_attr['executed_expression'], e)
                                             }

    def instance_full_clean(self, scheme, result, error_handler, error_row):

        try:

            instance = self.create_simple_instance(scheme, result)

            if instance:

                # self.fill_with_relation_attributes(instance, result)
                if scheme.content_type.model != 'pricehistory' and scheme.content_type.model != 'currencyhistory':
                    self.attributes_full_clean(instance, result['attributes'], error_handler, error_row)

                instance.full_clean()

        except CoreValidationError as e:

            _l.info('instance_full_clean  e %s' % e)

            error_row['error_reaction'] = 'Continue import'
            error_row['level'] = 'error'
            error_row['error_message'] = error_row['error_message'] + gettext_lazy(
                'Validation error %(error)s ') % {
                                             'error': 'Cannot create instance'
                                         }

    def instance_overwrite_full_clean(self, scheme, result, item, error_handler, error_row):

        # _l.debug('Overwrite item %s' % item)

        try:

            many_to_many_fields = ['counterparties', 'responsibles', 'accounts', 'portfolios']
            system_fields = ['_row_index', '_row']

            for key, value in result.items():

                if key != 'attributes':

                    if key not in many_to_many_fields and key not in system_fields:
                        setattr(item, key, value)

            # TODO import attribute validation later
            # self.fill_with_relation_attributes(item, result)
            # if scheme.content_type.model != 'pricehistory' and scheme.content_type.model != 'currencyhistory':
            #     self.attributes_full_clean(item, result['attributes'], error_handler, error_row)

            item.full_clean()

        except CoreValidationError as e:

            error_row['error_reaction'] = 'Continue import'
            error_row['level'] = 'error'
            error_row['error_message'] = error_row['error_message'] + gettext_lazy(
                'Validation error %(error)s ') % {
                                             'error': e
                                         }

            if error_handler == 'break':
                error_row['error_reaction'] = 'Break import'

    def full_clean_result(self, result_item, error_row, scheme, error_handler, mode, member, master_user):

        item = get_item(scheme, result_item)

        if mode == 'overwrite' and item:

            self.instance_overwrite_full_clean(scheme, result_item, item, error_handler, error_row)

        elif mode == 'overwrite' and not item:

            self.instance_full_clean(scheme, result_item, error_handler, error_row)

        elif mode == 'skip' and not item:

            self.instance_full_clean(scheme, result_item, error_handler, error_row)

        elif mode == 'skip' and item:

            error_row['level'] = 'error'
            error_row['error_reaction'] = 'Skipped'
            error_row['error_message'] = error_row['error_message'] + str(gettext_lazy(
                'Entry already exists '))

        return result_item, error_row

    def _row_count(self, file, instance):

        delimiter = instance.scheme.delimiter.encode('utf-8').decode('unicode_escape')

        reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                            strict=False, skipinitialspace=True)

        row_index = 0

        for row_index, row in enumerate(reader):
            pass

        return row_index

    def process(self, celery_task, update_state):

        _l.debug('ValidateHandler.process: initialized')

        scheme = CsvImportScheme.objects.get(pk=celery_task.options_object['scheme_id'])

        instance = CsvDataFileImport(task_id=celery_task.id,
                                     master_user=celery_task.master_user,
                                     member=celery_task.member,
                                     scheme=scheme,
                                     file_path=celery_task.options_object['file_path'],
                                     filename=celery_task.options_object['filename']
                                     )

        error_handler = scheme.error_handler
        missing_data_handler = scheme.missing_data_handler
        classifier_handler = scheme.classifier_handler
        delimiter = scheme.delimiter
        mode = scheme.mode
        master_user = instance.master_user
        member = instance.member

        execution_context = celery_task.options_object['execution_context']

        try:
            with storage.open(instance.file_path, 'rb') as f:

                _l.info('ValidateHandler.proces nstance.file_path %s' % instance.file_path)
                _l.info('ValidateHandler.proces f %s' % f)

                with NamedTemporaryFile() as tmpf:

                    for chunk in f.chunks():
                        tmpf.write(chunk)
                    tmpf.flush()

                    tmpf.flush()

                    if '.csv' in instance.filename:

                        with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cfr:
                            instance.total_rows = self._row_count(cfr, instance)
                            update_state(task_id=instance.task_id, state=Task.STATUS_PENDING,
                                         meta={'total_rows': instance.total_rows,
                                               'user_code': instance.scheme.user_code, 'file_name': instance.filename})

                    elif '.xlsx' in instance.filename:

                        wb = load_workbook(filename=tmpf.name)

                        if instance.scheme.spreadsheet_active_tab_name and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames:
                            ws = wb[instance.scheme.spreadsheet_active_tab_name]
                        else:
                            ws = wb.active

                        _l.info('ws %s' % ws)

                        reader = []

                        row_index = 0

                        for r in ws.rows:
                            row_index = row_index + 1

                        instance.total_rows = row_index

                        update_state(task_id=instance.task_id, state=Task.STATUS_PENDING,
                                     meta={'total_rows': instance.total_rows,
                                           'user_code': instance.scheme.user_code, 'file_name': instance.filename})

                    with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cf:
                        context = {}

                        results, process_errors = process_csv_file(master_user, scheme, f, cf, error_handler,
                                                                   missing_data_handler,
                                                                   classifier_handler,
                                                                   context, instance, update_state, mode,
                                                                   self.full_clean_result, member)

                        _l.debug('ValidateHandler.process_csv_file: finished')
                        _l.debug('ValidateHandler.process_csv_file process_errors %s: ' % len(process_errors))

                        instance.imported = len(results)
                        instance.stats = process_errors
        except Exception as e:

            _l.debug(e)
            _l.debug('Can\'t process file', exc_info=True)
            instance.error_message = gettext_lazy("Invalid file format or file already deleted.")
        finally:
            # import_file_storage.delete(instance.file_path)
            storage.delete(instance.file_path)

        if instance.stats and len(instance.stats):
            instance.stats_file_report = generate_file_report(instance, master_user, scheme, 'csv_import.validate',
                                                              'Simple Data Import Validation')

            send_websocket_message(data={
                'type': 'simple_import_status',
                'payload': {'task_id': instance.task_id,
                            'state': Task.STATUS_DONE,
                            'processed_rows': instance.processed_rows,
                            'total_rows': instance.total_rows,
                            'file_name': instance.filename,
                            'stats': instance.stats,
                            'stats_file_report': instance.stats_file_report,
                            'scheme': scheme.id,
                            'scheme_object': {
                                'id': scheme.id,
                                'user_code': scheme.user_code,
                                'classifier_handler': scheme.classifier_handler,
                                'delimiter': scheme.delimiter,
                                'error_handler': scheme.error_handler,
                                'missing_data_handler': scheme.missing_data_handler,
                                'mode': scheme.mode,
                            }}
            }, level="member",
                context={"master_user": master_user, "member": member})

        return instance


@shared_task(name='csv_import.data_csv_file_import_validate', bind=True)
def data_csv_file_import_validate(self, task_id):
    handler = ValidateHandler()

    celery_task = CeleryTask.objects.get(pk=task_id)

    handler.process(celery_task, self.update_state)


class ImportHandler:

    def delete_dynamic_attributes(self, instance, attributes):

        for result_attr in attributes:

            attr_type = GenericAttributeType.objects.get(pk=result_attr['dynamic_attribute_id'])

            if attr_type:
                attribute = GenericAttribute.objects.filter(object_id=instance.pk, attribute_type=attr_type)

                attribute.delete()

    def fill_with_dynamic_attributes(self, instance, attributes):

        for result_attr in attributes:

            attr_type = GenericAttributeType.objects.get(pk=result_attr['dynamic_attribute_id'])

            if attr_type:

                attribute = GenericAttribute(content_object=instance, attribute_type=attr_type)

                if attr_type.value_type == 10:
                    attribute.value_string = str(result_attr['executed_expression'])
                elif attr_type.value_type == 20:
                    attribute.value_float = float(result_attr['executed_expression'])
                elif attr_type.value_type == 30:

                    attribute.classifier = result_attr['executed_expression']

                elif attr_type.value_type == 40:
                    attribute.value_date = formula._parse_date(result_attr['executed_expression'])
                else:
                    pass

                attribute.save()

    def fill_with_relation_attributes(self, instance, result):

        for key, value in result.items():

            if key == 'counterparties':
                getattr(instance, key, False).add(result[key])
            elif key == 'responsibles':
                getattr(instance, key, False).add(result[key])
            elif key == 'accounts':
                getattr(instance, key, False).add(result[key])
            elif key == 'portfolios':
                getattr(instance, key, False).add(result[key])

    def create_simple_instance(self, scheme, result, error_row):

        Model = apps.get_model(app_label=scheme.content_type.app_label, model_name=scheme.content_type.model)

        result_without_many_to_many = {}

        if scheme.content_type.model == 'instrument':
            process = InstrumentTypeProcess(instrument_type=result['instrument_type'])

            result_without_many_to_many = process.instrument

            _l.info('create_simple_instance result_without_many_to_many %s ' % result_without_many_to_many)

        many_to_many_fields = ['counterparties', 'responsibles', 'accounts', 'portfolios']
        system_fields = ['_row_index', '_row']

        _l.info('create_simple_instance result %s ' % result)

        for key, value in result.items():

            if key != 'attributes':

                if key not in many_to_many_fields and key not in system_fields:
                    result_without_many_to_many[key] = value

        instance = None

        try:

            if scheme.content_type.model == 'instrument':
                from poms.instruments.serializers import InstrumentSerializer

                if 'accrued_currency' in result_without_many_to_many:
                    result_without_many_to_many['accrued_currency'] = result_without_many_to_many['accrued_currency'].id

                if 'instrument_type' in result_without_many_to_many:
                    result_without_many_to_many['instrument_type'] = result_without_many_to_many['instrument_type'].id

                if 'pricing_currency' in result_without_many_to_many:
                    result_without_many_to_many['pricing_currency'] = result_without_many_to_many['pricing_currency'].id

                serializer = InstrumentSerializer(data=result_without_many_to_many, context=self.context)

                is_valid = serializer.is_valid()

                if is_valid:
                    serializer.save()

                    instance = serializer.instance

                else:

                    _l.info('Serialize is not valid %s' % serializer.errors)

                    error_row['level'] = 'error'
                    error_row['error_message'] = error_row['error_message'] + str(serializer.errors)

                    raise Exception("Validation failed")

            else:
                instance = Model.objects.create(**result_without_many_to_many)
        except (ValidationError, IntegrityError, ValueError) as e:

            _l.info('create_simple_instance %s' % e)

            instance = None

            error_row['error_reaction'] = 'Continue import'
            error_row['level'] = 'error'
            error_row['error_message'] = error_row['error_message'] + gettext_lazy(
                'Cannot create simple instance %(error)s ') % {
                                             'error': e}

        return instance

    def add_permissions(self, instance, scheme, member, master_user):

        groups = Group.objects.filter(master_user=master_user)

        _l.debug('len groups for %s' % len(list(groups)))

        for group in groups:

            permission_table = group.permission_table

            if permission_table and 'data' in permission_table:

                table = None

                for item in permission_table['data']:
                    if item['content_type'] == scheme.content_type.app_label + '.' + scheme.content_type.model:
                        table = item['data']

                _l.debug('content_type %s' % scheme.content_type.app_label + '.' + scheme.content_type.model)
                # _l.debug('table %s' % table)

                if table:

                    manage_codename = 'manage_' + scheme.content_type.model
                    change_codename = 'change_' + scheme.content_type.model
                    view_codename = 'view_' + scheme.content_type.model

                    manage_permission = Permission.objects.get(content_type=scheme.content_type,
                                                               codename=manage_codename)
                    change_permission = Permission.objects.get(content_type=scheme.content_type,
                                                               codename=change_codename)
                    view_permission = Permission.objects.get(content_type=scheme.content_type, codename=view_codename)

                    for member_group in member.groups.all():

                        if member_group.id == group.id:

                            if 'creator_manage' in table and table['creator_manage']:
                                perm = GenericObjectPermission.objects.update_or_create(object_id=instance.id,
                                                                                        content_type=scheme.content_type,
                                                                                        group=group,
                                                                                        permission=manage_permission)

                            if 'creator_change' in table and table['creator_change']:
                                perm = GenericObjectPermission.objects.update_or_create(object_id=instance.id,
                                                                                        content_type=scheme.content_type,
                                                                                        group=group,
                                                                                        permission=change_permission)

                            if 'creator_view' in table and table['creator_view']:
                                perm = GenericObjectPermission.objects.update_or_create(object_id=instance.id,
                                                                                        content_type=scheme.content_type,
                                                                                        group=group,
                                                                                        permission=view_permission)
                        else:

                            if 'other_manage' in table and table['other_manage']:
                                perm = GenericObjectPermission.objects.update_or_create(object_id=instance.id,
                                                                                        content_type=scheme.content_type,
                                                                                        group=group,
                                                                                        permission=manage_permission)
                            if 'other_change' in table and table['other_change']:
                                perm = GenericObjectPermission.objects.update_or_create(object_id=instance.id,
                                                                                        content_type=scheme.content_type,
                                                                                        group=group,
                                                                                        permission=change_permission)
                            if 'other_view' in table and table['other_view']:
                                perm = GenericObjectPermission.objects.update_or_create(object_id=instance.id,
                                                                                        content_type=scheme.content_type,
                                                                                        group=group,
                                                                                        permission=view_permission)

    def is_inherit_rights(self, scheme, member):

        result = False

        if scheme.content_type.model == 'account' or scheme.content_type.model == 'instrument':

            for group in member.groups.all():

                permission_table = group.permission_table

                if permission_table and 'data' in permission_table:

                    table = None

                    for item in permission_table['data']:
                        if item['content_type'] == scheme.content_type.app_label + '.' + scheme.content_type.model:
                            table = item['data']

                    if table:

                        if table['inherit_rights']:
                            result = True

        return result

    def add_inherited_permissions(self, instance, scheme, member, master_user):

        _l.debug('add_inherited_permissions')

        if scheme.content_type.model == 'account':

            if instance.type:
                for perm in instance.type.object_permissions.all():
                    type_codename = perm.permission.codename.split('_')[0]
                    account_perm = Permission.objects.get(codename=type_codename + '_' + 'account')

                    GenericObjectPermission.objects.create(group=perm.group, object_id=instance.id,
                                                           content_type=scheme.content_type,
                                                           permission=account_perm)

        if scheme.content_type.model == 'instrument':
            if instance.instrument_type:
                for perm in instance.instrument_type.object_permissions.all():
                    type_codename = perm.permission.codename.split('_')[0]
                    account_perm = Permission.objects.get(codename=type_codename + '_' + 'instrument')

                    GenericObjectPermission.objects.create(group=perm.group, object_id=instance.id,
                                                           content_type=scheme.content_type,
                                                           permission=account_perm)

    def add_instrument_type_pricing_policies(self, instance, scheme, member, master_user):

        _l.debug("Add Pricing Policies instrument_type %s" % instance.instrument_type)

        if instance.instrument_type:

            InstrumentPricingPolicy.objects.filter(instrument=instance).delete()

            for pp_item in instance.instrument_type.pricing_policies.all():
                item = InstrumentPricingPolicy()

                item.instrument = instance
                item.pricing_policy = pp_item.pricing_policy
                item.pricing_scheme = pp_item.pricing_scheme
                item.notes = pp_item.notes
                item.default_value = pp_item.default_value
                item.attribute_key = pp_item.attribute_key
                item.json_data = pp_item.json_data

                item.save()

    def save_instance(self, scheme, result, error_handler, error_row, member, master_user):

        try:

            instance = self.create_simple_instance(scheme, result, error_row)

            if instance:

                self.fill_with_relation_attributes(instance, result)
                if scheme.content_type.model != 'pricehistory' and scheme.content_type.model != 'currencyhistory':
                    self.fill_with_dynamic_attributes(instance, result['attributes'])
                    self.add_permissions(instance, scheme, member, master_user)

                    if self.is_inherit_rights(scheme, member):
                        self.add_inherited_permissions(instance, scheme, member, master_user)

                if scheme.content_type.model == 'instrument':
                    self.add_instrument_type_pricing_policies(instance, scheme, member, master_user)

                instance.save()

        except Exception as e:

            _l.info("save_instance Exception %s" % e)

            error_row['error_reaction'] = 'Continue import'
            error_row['level'] = 'error'
            error_row['error_message'] = error_row['error_message'] + gettext_lazy(
                'Validation error %(error)s ') % {
                                             'error': e
                                         }

    def overwrite_instance(self, scheme, result, item, error_handler, error_row, member, master_user):

        # _l.debug('Overwrite item %s' % item)

        # _l.debug('ImportHandler overwrite_instance %s ' % item)

        try:

            many_to_many_fields = ['counterparties', 'responsibles', 'accounts', 'portfolios']
            system_fields = ['_row_index', '_row']

            for key, value in result.items():

                if key != 'attributes':

                    if key not in many_to_many_fields and key not in system_fields:
                        setattr(item, key, value)

            self.fill_with_relation_attributes(item, result)
            if scheme.content_type.model != 'pricehistory' and scheme.content_type.model != 'currencyhistory':
                self.delete_dynamic_attributes(item, result['attributes'])
                self.fill_with_dynamic_attributes(item, result['attributes'])
                self.add_permissions(item, scheme, member, master_user)

            if scheme.content_type.model == 'instrument':
                self.add_instrument_type_pricing_policies(item, scheme, member, master_user)

            item.save()

        except ValidationError as e:

            error_row['error_reaction'] = 'Continue import'
            error_row['level'] = 'error'
            error_row['error_message'] = error_row['error_message'] + str(gettext_lazy(
                'Validation error %(error)s ') % {
                                                                              'error': e
                                                                          })

    def add_instrument_to_download_queue(self, scheme, result_item, error_handler, error_row, member, master_user):

        from poms.integrations.tasks import download_instrument_cbond_task

        if scheme.instrument_reference_column:
            user_code = result_item[scheme.instrument_reference_column]

            # DEPRECATED, REFACTOR SOON
            task = Task(
                master_user=master_user,
                member=member,
                status=Task.STATUS_PENDING,
                action='csv_import_instrument_download_from_provider',
            )

            task.options_object = {
                'user_code': user_code
            }
            task.save()

            download_instrument_cbond_task.apply_async(kwargs={'task_id': task.id})
        else:
            _l.info("ISIN For instrument download is not found")

    def import_result(self, result_item, error_row, scheme, error_handler, mode, member, master_user):

        # _l.debug('ImportHandler.result_item %s' % result_item)

        item = get_item(scheme, result_item)

        if mode == 'overwrite' and item:

            # _l.debug('Overwrite instance')

            self.overwrite_instance(scheme, result_item, item, error_handler, error_row, member, master_user)

        elif mode == 'overwrite' and not item:

            _l.debug('Create instance')

            self.save_instance(scheme, result_item, error_handler, error_row, member, master_user)

        elif mode == 'skip' and not item:

            _l.debug('Create instance')

            self.save_instance(scheme, result_item, error_handler, error_row, member, master_user)

        elif mode == 'skip' and item:

            # _l.debug('Skip instance %s')

            error_row['level'] = 'error'
            error_row['error_message'] = error_row['error_message'] + error_row[
                'error_message']

        if scheme.content_type.model == 'instrument' and scheme.instrument_reference_column:
            self.add_instrument_to_download_queue(scheme, result_item, error_handler, error_row, member, master_user)

        return result_item, error_row

    def _row_count(self, file, instance):

        delimiter = instance.scheme.delimiter.encode('utf-8').decode('unicode_escape')

        reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                            strict=False, skipinitialspace=True)

        row_index = 0

        for row_index, row in enumerate(reader):
            pass

        # return plain index, -1 row because of ignoring csv header row

        return row_index

    def process(self, celery_task, update_state, procedure_instance=None):

        _l.debug('ImportHandler.process: initialized')

        scheme = CsvImportScheme.objects.get(pk=celery_task.options_object['scheme_id'])

        instance = CsvDataFileImport(task_id=celery_task.id,
                                     master_user=celery_task.master_user,
                                     member=celery_task.member,
                                     scheme=scheme,
                                     file_path=celery_task.options_object['file_path'],
                                     filename=celery_task.options_object['filename'])

        error_handler = scheme.error_handler
        missing_data_handler = scheme.missing_data_handler
        classifier_handler = scheme.classifier_handler
        delimiter = scheme.delimiter
        mode = scheme.mode
        master_user = instance.master_user
        member = instance.member

        execution_context = celery_task.options_object['execution_context']

        proxy_user = ProxyUser(member, master_user)
        proxy_request = ProxyRequest(proxy_user)

        self.context = {'master_user': master_user,
                        'request': proxy_request}

        _l.debug('ImportHandler.mode %s' % mode)

        try:

            if celery_task and celery_task.options_object and 'items' in celery_task.options_object:

                _l.info("Parse json data")

                items = celery_task.options_object['items']

                context = {}

                instance.total_rows = len(items)

                results, process_errors = process_csv_file(master_user, scheme, None, None, error_handler,
                                                           missing_data_handler,
                                                           classifier_handler,
                                                           context, instance, update_state, mode,
                                                           self.import_result, member, execution_context, celery_task)

                instance.imported = len(results)
                instance.stats = process_errors

            else:

                with storage.open(instance.file_path, 'rb') as f:

                    with NamedTemporaryFile() as tmpf:
                        for chunk in f.chunks():
                            tmpf.write(chunk)
                        tmpf.flush()

                        if '.csv' in instance.filename or (
                                execution_context and execution_context["started_by"] == 'procedure'):

                            with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cfr:
                                instance.total_rows = self._row_count(cfr, instance)
                                update_state(task_id=instance.task_id, state=Task.STATUS_PENDING,
                                             meta={'total_rows': instance.total_rows,
                                                   'user_code': instance.scheme.user_code,
                                                   'file_name': instance.filename})

                        elif '.xlsx' in instance.filename:

                            wb = load_workbook(filename=f)

                            if instance.scheme.spreadsheet_active_tab_name and instance.scheme.spreadsheet_active_tab_name in wb.sheetnames:
                                ws = wb[instance.scheme.spreadsheet_active_tab_name]
                            else:
                                ws = wb.active

                            _l.info('ws %s' % ws)

                            reader = []

                            row_index = 0

                            for r in ws.rows:
                                row_index = row_index + 1

                            instance.total_rows = row_index

                            update_state(task_id=instance.task_id, state=Task.STATUS_PENDING,
                                         meta={'total_rows': instance.total_rows,
                                               'user_code': instance.scheme.user_code, 'file_name': instance.filename})

                        with open(tmpf.name, mode='rt', encoding=instance.encoding, errors='ignore') as cf:
                            context = {}

                            results, process_errors = process_csv_file(master_user, scheme, f, cf, error_handler,
                                                                       missing_data_handler,
                                                                       classifier_handler,
                                                                       context, instance, update_state, mode,
                                                                       self.import_result, member, execution_context,
                                                                       celery_task)

                            _l.debug('ImportHandler.process_csv_file: finished')
                            _l.debug('ImportHandler.process_csv_file process_errors %s: ' % len(process_errors))

                            instance.imported = len(results)
                            instance.stats = process_errors
        except Exception as e:

            _l.debug(e)
            _l.debug('Can\'t process file', exc_info=True)
            instance.error_message = gettext_lazy("Invalid file format or file already deleted.")
        finally:
            # import_file_storage.delete(instance.file_path)
            if instance.file_path:
                storage.delete(instance.file_path)

        if instance.stats and len(instance.stats):
            instance.stats_file_report = generate_file_report(instance, master_user, scheme, 'csv_import.import',
                                                              'Simple Data Import', procedure_instance)

            send_websocket_message(data={
                'type': 'simple_import_status',
                'payload': {'task_id': instance.task_id,
                            'state': Task.STATUS_DONE,
                            'processed_rows': instance.processed_rows,
                            'total_rows': instance.total_rows,
                            'file_name': instance.filename,
                            'stats': instance.stats,
                            'stats_file_report': instance.stats_file_report,
                            'scheme': scheme.id,
                            'scheme_object': {
                                'id': scheme.id,
                                'user_code': scheme.user_code,
                                'classifier_handler': scheme.classifier_handler,
                                'delimiter': scheme.delimiter,
                                'error_handler': scheme.error_handler,
                                'missing_data_handler': scheme.missing_data_handler,
                                'mode': scheme.mode,
                            }}
            }, level="member",
                context={"master_user": master_user, "member": member})

            send_websocket_message(data={'type': "simple_message",
                                         'payload': {
                                             'message': "Member %s imported data with Simple Import Service" % member.username
                                         }
                                         }, level="master_user", context={"master_user": master_user, "member": member})

            if execution_context and execution_context["started_by"] == 'procedure':
                send_system_message(master_user=instance.master_user,
                                    performed_by="System",
                                    description="Import Finished",
                                    attachments=[instance.stats_file_report])
            else:
                send_system_message(master_user=instance.master_user,
                                    performed_by="System",
                                    description="User %s Import Finished" % member.username,
                                    attachments=[instance.stats_file_report])

        if procedure_instance and procedure_instance.schedule_instance:
            procedure_instance.schedule_instance.run_next_procedure()

        celery_task.status = CeleryTask.STATUS_DONE
        celery_task.save()

        return instance


@shared_task(name='csv_import.data_csv_file_import', bind=True)
def data_csv_file_import(self, task_id, procedure_instance_id=None):
    try:

        handler = ImportHandler()

        celery_task = CeleryTask.objects.get(pk=task_id)
        procedure_instance = None

        if procedure_instance_id:
            procedure_instance = RequestDataFileProcedureInstance.objects.get(id=procedure_instance_id)

        handler.process(celery_task, self.update_state, procedure_instance)

    except Exception as e:

        _l.debug('data_csv_file_import exception error %s' % e)
        _l.info(traceback.format_exc())


@shared_task(name='csv_import.data_csv_file_import_by_procedure', bind=True)
def data_csv_file_import_by_procedure(self, procedure_instance_id, transaction_file_result_id):
    with transaction.atomic():

        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.procedures.models import RequestDataFileProcedureInstance

        procedure_instance = RequestDataFileProcedureInstance.objects.get(id=procedure_instance_id)
        transaction_file_result = TransactionFileResult.objects.get(id=transaction_file_result_id)

        try:

            _l.debug(
                'data_csv_file_import_by_procedure looking for scheme_user_code %s ' % procedure_instance.procedure.scheme_user_code)

            scheme = CsvImportScheme.objects.get(master_user=procedure_instance.master_user,
                                                 user_code=procedure_instance.procedure.scheme_user_code)

            text = "Data File Procedure %s. File is received, start data import" % (
                procedure_instance.procedure.user_code)

            send_system_message(master_user=procedure_instance.master_user,
                                performed_by='System',
                                description=text)

            with storage.open(transaction_file_result.file_path, 'rb') as f:

                try:

                    encrypted_text = f.read()

                    rsa_cipher = RSACipher()

                    aes_key = None

                    try:
                        aes_key = rsa_cipher.decrypt(procedure_instance.private_key, procedure_instance.symmetric_key)
                    except Exception as e:
                        _l.debug('data_csv_file_import_by_procedure AES Key decryption error %s' % e)

                    aes_cipher = AESCipher(aes_key)

                    decrypt_text = None

                    try:
                        decrypt_text = aes_cipher.decrypt(encrypted_text)
                    except Exception as e:
                        _l.debug('data_csv_file_import_by_procedure Text decryption error %s' % e)

                    _l.debug('data_csv_file_import_by_procedure file decrypted')

                    _l.debug('Size of decrypted text: %s' % len(decrypt_text))

                    with NamedTemporaryFile() as tmpf:

                        tmpf.write(decrypt_text.encode('utf-8'))
                        tmpf.flush()

                        file_name = '%s-%s' % (timezone.now().strftime('%Y%m%d%H%M%S'), uuid.uuid4().hex)
                        file_path = '%s/public/%s.csv' % (settings.BASE_API_URL, file_name)

                        storage.save(file_path, tmpf)

                        _l.debug('data_csv_file_import_by_procedure tmp file filled')

                        instance = CsvDataFileImport(scheme=scheme,
                                                     file_path=file_path,
                                                     filename=transaction_file_result.file_path,
                                                     member=procedure_instance.member,
                                                     master_user=procedure_instance.master_user,
                                                     delimiter=scheme.delimiter,
                                                     error_handler=scheme.error_handler,
                                                     mode=scheme.mode,
                                                     classifier_handler=scheme.classifier_handler,
                                                     missing_data_handler=scheme.missing_data_handler)

                        _l.debug('data_csv_file_import_by_procedure instance: %s' % instance)

                        current_date_time = now().strftime("%Y-%m-%d-%H-%M")

                        file_name = '%s-%s' % (timezone.now().strftime('%Y%m%d%H%M%S'), uuid.uuid4().hex)
                        file_name_hash = hashlib.md5(file_name.encode('utf-8')).hexdigest()

                        file_report = FileReport()

                        file_report.upload_file(
                            file_name='Data Procedure %s (%s).csv' % (current_date_time, file_name_hash),
                            text=decrypt_text, master_user=procedure_instance.master_user)
                        file_report.master_user = procedure_instance.master_user
                        file_report.name = "'Data Import File. Procedure ' %s %s" % (
                            procedure_instance.id, current_date_time)
                        file_report.file_name = 'Data Procedure %s (%s).csv' % (current_date_time, file_name_hash)
                        file_report.type = 'csv_import.import'
                        file_report.notes = 'Data Import File. Procedure %s' % procedure_instance.id
                        file_report.content_type = 'text/csv'

                        file_report.save()

                        _l.debug('file_report %s' % file_report)

                        text = "Data File Procedure %s. File is received. Start Import" % (
                            procedure_instance.procedure.user_code)

                        send_system_message(master_user=procedure_instance.master_user,
                                            performed_by='System',
                                            description=text,
                                            attachments=[file_report.id])

                        transaction.on_commit(
                            lambda: data_csv_file_import.apply_async(
                                kwargs={'instance': instance, 'execution_context': {'started_by': 'procedure'}}))


                except Exception as e:

                    traceback.format_exc()

                    _l.debug('data_csv_file_import_by_procedure decryption error %s' % e)

        except CsvImportScheme.DoesNotExist:

            text = "Data File Procedure %s. Can't import file, Import scheme %s is not found" % (
                procedure_instance.procedure.user_code, procedure_instance.procedure.user_code)

            send_system_message(master_user=procedure_instance.master_user,
                                performed_by='System',
                                description=text)

            _l.debug(
                'data_csv_file_import_by_procedure scheme %s not found' % procedure_instance.procedure.user_code)

            procedure_instance.status = RequestDataFileProcedureInstance.STATUS_ERROR
            procedure_instance.save()


@shared_task(name='csv_import.data_csv_file_import_by_procedure_json', bind=True)
def data_csv_file_import_by_procedure_json(self, procedure_instance_id, celery_task_id):
    with transaction.atomic():

        _l.info('data_csv_file_import_by_procedure_json  procedure_instance_id %s celery_task_id %s' % (
            procedure_instance_id, celery_task_id))

        from poms.integrations.serializers import ComplexTransactionCsvFileImport
        from poms.procedures.models import RequestDataFileProcedureInstance

        procedure_instance = RequestDataFileProcedureInstance.objects.get(id=procedure_instance_id)
        celery_task = CeleryTask.objects.get(id=celery_task_id)
        celery_task.status = CeleryTask.STATUS_PENDING
        celery_task.celery_task_id = self.request.id
        celery_task.save()

        try:

            _l.info(
                'data_csv_file_import_by_procedure_json looking for scheme %s ' % procedure_instance.procedure.scheme_user_code)

            scheme = CsvImportScheme.objects.get(master_user=procedure_instance.master_user,
                                                 user_code=procedure_instance.procedure.scheme_user_code)

            options_object = celery_task.options_object

            options_object['file_path'] = ''
            options_object['filename'] = ''
            options_object['scheme_id'] = scheme.id
            options_object['execution_context'] = {'started_by': 'procedure'}

            celery_task.options_object = options_object
            celery_task.save()

            text = "Data File Procedure %s. Procedure Instance %s. File is received. Importing JSON" % (
                procedure_instance.id,
                procedure_instance.procedure.user_code)

            send_system_message(master_user=procedure_instance.master_user,
                                performed_by='System',
                                description=text)

            transaction.on_commit(lambda: data_csv_file_import.apply_async(
                kwargs={"task_id": celery_task.id, "procedure_instance_id": procedure_instance_id}))


        except Exception as e:

            _l.info('data_csv_file_import_by_procedure_json e %s' % e)

            text = "Data File Procedure %s. Can't import json, Error %s" % (
                procedure_instance.procedure.user_code, e)

            send_system_message(master_user=procedure_instance.master_user,
                                performed_by='System',
                                description=text)

            _l.debug(
                'data_csv_file_import_by_procedure_json scheme %s not found' % procedure_instance.procedure.scheme_name)

            procedure_instance.status = RequestDataFileProcedureInstance.STATUS_ERROR
            procedure_instance.save()


class UnifiedImportHandler():

    def __init__(self, instance, update_state, execution_context):
        self.instance = instance
        self.update_state = update_state
        self.execution_context = execution_context

    def get_row_data(self, row, first_row):
        csv_row_dict = {}

        index = 0
        for col in first_row:
            col_name = col.lower().replace(" ", "_")

            csv_row_dict[col_name] = row[index]

            index = index + 1

        return csv_row_dict

    def _row_count(self, file, instance):

        delimiter = instance.scheme.delimiter.encode('utf-8').decode('unicode_escape')

        reader = csv.reader(file, delimiter=delimiter, quotechar=instance.quotechar,
                            strict=False, skipinitialspace=True)

        row_index = 0

        for row_index, row in enumerate(reader):
            pass

        # return plain index, -1 row because of ignoring csv header row

        return row_index

    def process_row(self, first_row, row, item, context):

        from poms.instruments.serializers import InstrumentSerializer

        try:

            row_as_dict = self.get_row_data(row, first_row)

            item['row_as_dict'] = row_as_dict

            row_data = {}
            row_data = row_as_dict  # tmp

            instrument_type = None

            # _l.info('row_data %s' % row_data)

            skip = False

            try:

                instrument_type = InstrumentType.objects.get(master_user=self.instance.master_user,
                                                             user_code=row_as_dict['instrument_type'])


            except Exception as e:

                item['error_message'] = 'Instrument Type Does not Find'

                skip = True

            if skip == False:

                row_data = handler_instrument_object(row_as_dict, instrument_type, self.instance.master_user,
                                                     self.ecosystem_default, self.attribute_types)

                if self.instance.mode == 'skip':

                    serializer = InstrumentSerializer(data=row_data, context=context)

                    is_valid = serializer.is_valid()

                    item['row_data'] = row_data

                    if is_valid:
                        serializer.save()
                    else:
                        item['error_message'] = serializer.errors

                if self.instance.mode == 'overwrite':

                    instrument = None

                    try:
                        instrument = Instrument.objects.get(user_code=row_data['user_code'],
                                                            master_user=self.instance.master_user)
                    except Instrument.DoesNotExist:
                        instrument = None

                    serializer = InstrumentSerializer(data=row_data, context=context, instance=instrument)

                    is_valid = serializer.is_valid()

                    item['row_data'] = row_data

                    if is_valid:
                        serializer.save()
                    else:
                        item['error_message'] = serializer.errors

        except Exception as e:

            _l.info("Error %s" % e)
            _l.info(traceback.format_exc())

            item['error_message'] = 'Unhandled error in row processing. Exception %s' % str(e)

        finally:

            self.instance.processed_rows = item['original_row_index']

            send_websocket_message(data={
                'type': 'simple_import_status',
                'payload': {'task_id': self.instance.task_id,
                            'state': Task.STATUS_PENDING,
                            'processed_rows': self.instance.processed_rows,
                            'total_rows': self.instance.total_rows,
                            'file_name': self.instance.filename}
            }, level="member",
                context={"master_user": self.instance.master_user, "member": self.instance.member})

    def unified_process_csv_file(self, file):

        _l.info('unified_process_csv_file mode %s' % self.instance.mode)

        errors = []
        results = []

        delimiter = self.instance.scheme.delimiter.encode('utf-8').decode('unicode_escape')

        reader = csv.reader(file, delimiter=delimiter, quotechar=self.instance.quotechar,
                            strict=False, skipinitialspace=True)

        first_row = None

        proxy_user = ProxyUser(self.instance.member, self.instance.master_user)
        proxy_request = ProxyRequest(proxy_user)

        context = {'master_user': self.instance.master_user,
                   'request': proxy_request}

        instrument_content_type = ContentType.objects.get(app_label="instruments", model='instrument')

        self.ecosystem_default = EcosystemDefault.objects.get(master_user=self.instance.master_user)
        self.attribute_types = GenericAttributeType.objects.filter(master_user=self.instance.master_user,
                                                                   content_type=instrument_content_type)

        items = []

        for row_index, row in enumerate(reader):

            item = {
                'original_row_index': row_index,
                'error_message': ''
            }

            if row_index == 0:
                first_row = row
            else:

                self.process_row(first_row, row, item, context)

            items.append(item)

        return items

    def process(self):

        _l.debug('UnifiedImportHandler.process: initialized')

        mode = self.instance.mode
        master_user = self.instance.master_user
        member = self.instance.member

        _l.debug('UnifiedImportHandler.mode %s' % mode)

        try:
            with storage.open(self.instance.file_path, 'rb') as f:
                with NamedTemporaryFile() as tmpf:
                    for chunk in f.chunks():
                        tmpf.write(chunk)
                    tmpf.flush()

                    with open(tmpf.name, mode='rt', encoding=self.instance.encoding, errors='ignore') as cfr:
                        self.instance.total_rows = self._row_count(cfr, self.instance)
                        self.update_state(task_id=self.instance.task_id, state=Task.STATUS_PENDING,
                                          meta={'total_rows': self.instance.total_rows,
                                                'file_name': self.instance.filename})

                    with open(tmpf.name, mode='rt', encoding=self.instance.encoding, errors='ignore') as cf:
                        context = {}

                        items = self.unified_process_csv_file(cf)

                        _l.debug('UnifiedImportHandler.process_csv_file: finished')

                        self.instance.items = items
        except Exception as e:

            _l.debug(e)
            _l.debug('Can\'t process file', exc_info=True)
            self.instance.error_message = gettext_lazy("Invalid file format or file already deleted.")
        finally:
            # import_file_storage.delete(instance.file_path)
            storage.delete(self.instance.file_path)

        _l.info("Import here? %s" % len(self.instance.items))

        if self.instance.items and len(self.instance.items):

            self.instance.stats_file_report = generate_file_report_simple(self.instance, 'csv_import.unified_import',
                                                                          'Unified Data Import')

            _l.info('self.instance.stats_file_report %s' % self.instance.stats_file_report)

            send_websocket_message(data={
                'type': 'simple_import_status',
                'payload': {'task_id': self.instance.task_id,
                            'state': Task.STATUS_DONE,
                            'processed_rows': self.instance.processed_rows,
                            'total_rows': self.instance.total_rows,
                            'file_name': self.instance.filename,
                            'stats': self.instance.stats,
                            'stats_file_report': self.instance.stats_file_report
                            }
            }, level="member",
                context={"master_user": master_user, "member": member})

            send_websocket_message(data={'type': "simple_message",
                                         'payload': {
                                             'message': "Member %s imported data with Simple Import Service" % member.username
                                         }
                                         }, level="master_user", context={"master_user": master_user, "member": member})

            if self.execution_context and self.execution_context["started_by"] == 'procedure':
                send_system_message(master_user=self.instance.master_user,
                                    performed_by='System',
                                    description="Import Finished",
                                    attachments=[self.instance.stats_file_report])
            else:
                send_system_message(master_user=self.instance.master_user,
                                    performed_by='System',
                                    description="User %s Import Finished" % member.username,
                                    attachments=[self.instance.stats_file_report])

        return self.instance


@shared_task(name='csv_import.unified_data_csv_file_import', bind=True)
def unified_data_csv_file_import(self, instance, execution_context=None):
    handler = UnifiedImportHandler(instance, self.update_state, execution_context)

    setattr(instance, 'task_id', current_task.request.id)

    handler.process()

    return instance
