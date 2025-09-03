import csv
import json
import logging
import os
import re
import time
import traceback
from copy import deepcopy
from datetime import date
from tempfile import NamedTemporaryFile

from django.utils.timezone import now
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from poms.accounts.models import Account
from poms.celery_tasks.models import CeleryTask
from poms.common.models import ProxyRequest, ProxyUser
from poms.common.storage import get_storage
from poms.counterparties.models import Counterparty, Responsible
from poms.currencies.models import Currency
from poms.expressions_engine import formula
from poms.file_reports.models import FileReport
from poms.instruments.models import (
    AccrualCalculationModel,
    DailyPricingModel,
    Instrument,
    InstrumentType,
    PaymentSizeDetail,
    Periodicity,
)
from poms.integrations.models import ComplexTransactionImportScheme
from poms.portfolios.models import Portfolio
from poms.procedures.models import RequestDataFileProcedureInstance
from poms.strategies.models import Strategy1, Strategy2, Strategy3
from poms.system_messages.handlers import send_system_message
from poms.transaction_import.exceptions import (
    BookException,
    BookSkipException,
    BookUnhandledException,
)
from poms.transaction_import.models import (
    ProcessType,
    TransactionImportBookedTransaction,
    TransactionImportConversionItem,
    TransactionImportProcessItem,
    TransactionImportProcessPreprocessItem,
    TransactionImportResult,
)
from poms.transaction_import.serializers import TransactionImportResultSerializer
from poms.transactions.handlers import TransactionTypeProcess
from poms.transactions.models import TransactionType, TransactionTypeInput
from poms.users.models import EcosystemDefault

storage = get_storage()


_l = logging.getLogger("poms.transaction_import")

props_map = {
    Account: "account",
    Currency: "currency",
    Instrument: "instrument",
    InstrumentType: "instrument_type",
    Counterparty: "counterparty",
    Responsible: "responsible",
    Strategy1: "strategy1",
    Strategy2: "strategy2",
    Strategy3: "strategy3",
    DailyPricingModel: "daily_pricing_model",
    PaymentSizeDetail: "payment_size_detail",
    Portfolio: "portfolio",
    Periodicity: "periodicity",
    AccrualCalculationModel: "accrual_calculation_model",
}


class TransactionImportProcess:
    def __init__(self, task_id, procedure_instance_id=None):
        self.task = CeleryTask.objects.get(pk=task_id)
        self.parent_task = self.task.parent

        _l.info(f"TransactionImportProcess.Task {self.task}. init")

        self.task.status = CeleryTask.STATUS_PENDING
        self.task.save()

        self.procedure_instance = None
        if procedure_instance_id:
            self.procedure_instance = RequestDataFileProcedureInstance.objects.get(id=procedure_instance_id)

            _l.info(f"TransactionImportProcess.Task {self.task}. init procedure_instance {self.procedure_instance}")

        self.member = self.task.member
        self.master_user = self.task.master_user
        self.proxy_user = ProxyUser(self.member, self.master_user)
        self.proxy_request = ProxyRequest(self.proxy_user)

        if self.task.options_object.get("scheme_id", None):
            self.scheme = ComplexTransactionImportScheme.objects.get(pk=self.task.options_object["scheme_id"])
        elif self.task.options_object.get("scheme_user_code", None):
            self.scheme = ComplexTransactionImportScheme.objects.get(
                user_code=self.task.options_object["scheme_user_code"]
            )
        else:
            raise RuntimeError("Import Scheme not found")

        self.execution_context = self.task.options_object["execution_context"]
        self.file_path = self.task.options_object["file_path"]

        self.ecosystem_default = EcosystemDefault.cache.get_cache(master_user_pk=self.master_user.pk)

        self.find_default_rule_scenario()
        self.find_error_rule_scenario()

        self.result = TransactionImportResult()
        self.result.task = self.task
        self.result.scheme = self.scheme

        self.process_type = ProcessType.CSV  # default type
        self.find_process_type()

        self.file_items = []  # items from provider  (json, csv, excel)
        self.raw_items = []  # items from provider  (json, csv, excel)
        self.conversion_items = []  # items with applied converions
        self.preprocessed_items = []  # items with calculated variables applied
        self.items = []  # result items that will be passed to TransactionTypeProcess

        self.context = {
            "master_user": self.master_user,
            "member": self.member,
            "request": self.proxy_request,
        }

        import_system_message_performed_by = self.member.username
        import_system_message_title = "Transaction import (start)"

        if self.execution_context and self.execution_context["started_by"] == "procedure":
            import_system_message_performed_by = "System"
            import_system_message_title = "Transaction import from broker (start)"

        self.prefetched_relations = {}
        self.prefetch_relations()

        send_system_message(
            master_user=self.master_user,
            performed_by=import_system_message_performed_by,
            section="import",
            type="success",
            title=import_system_message_title,
            description=f"{self.member.username} started import with scheme {self.scheme.name}",
        )

        _l.info(f"self.scheme.book_uniqueness_settings {self.scheme.book_uniqueness_settings}")

    def prefetch_relations(self):
        def as_dict(items):
            result_dict = {}
            for item in items:
                result_dict[item.user_code] = item
            return result_dict

        st = time.perf_counter()

        result = {
            "accounts.account": as_dict(Account.objects.all()),
            "portfolios.portfolio": as_dict(Portfolio.objects.all()),
            "currencies.currency": as_dict(Currency.objects.all()),
            "instruments.instrument": as_dict(Instrument.objects.all()),
            "strategies.strategy1": as_dict(Strategy1.objects.all()),
            "strategies.strategy2": as_dict(Strategy2.objects.all()),
            "strategies.strategy3": as_dict(Strategy3.objects.all()),
            "counterparties.counterparty": as_dict(Counterparty.objects.all()),
            "counterparties.responsible": as_dict(Responsible.objects.all()),
            "instruments.instrumenttype": as_dict(InstrumentType.objects.all()),
        }

        self.prefetched_relations = result

        _l.info(
            "TransactionImportProcess: prefetch_relations done: %s",
            f"{time.perf_counter() - st:3.3f}",
        )

    def items_has_error(self):
        return any(result_item.status == "error" for result_item in self.result.items)

    def generate_file_report(self):
        _l.info("TransactionImportProcess.generate_file_report error_handler %s", self.scheme.error_handler)
        _l.info(
            "TransactionImportProcess.generate_file_report missing_data_handler %s", self.scheme.missing_data_handler
        )

        result = [
            "Type, Transaction Import",
            "Scheme, " + self.scheme.user_code,
            "Error handle, " + self.scheme.error_handler,
            "Uniqueness settings, " + str(self.scheme.book_uniqueness_settings),
        ]

        if self.result.file_name:
            result.append("Filename, " + self.result.file_name)

        result.append("Import Rules - if object is not found, " + self.scheme.missing_data_handler)

        success_rows_count = 0
        error_rows_count = 0
        skip_rows_count = 0

        for result_item in self.result.items:
            if result_item.status == "error":
                error_rows_count = error_rows_count + 1

            if result_item.status == "success":
                success_rows_count = success_rows_count + 1

            if "skip" in result_item.status:
                skip_rows_count = skip_rows_count + 1

        result.append(f"Rows total, {self.result.total_rows}")
        result.append(f"Rows success import, {success_rows_count}")
        result.append(f"Rows fail import, {error_rows_count}")
        result.append(f"Rows skipped import, {skip_rows_count}")

        columns = ["Row Number", "Status", "Message"]

        column_row_list = []

        for item in columns:
            column_row_list.append('"' + str(item) + '"')

        column_row = ",".join(column_row_list)

        result.append(column_row)

        for result_item in self.result.items:
            content = []

            content.append(str(result_item.row_number))
            content.append(result_item.status)

            if result_item.error_message:
                content.append(result_item.error_message)
            elif result_item.message:
                content.append(result_item.message)
            else:
                content.append("")

            content_row_list = []

            for item in content:
                content_row_list.append('"' + str(item) + '"')

            content_row = ",".join(content_row_list)

            result.append(content_row)

        result = "\n".join(result)

        current_date_time = now().strftime("%Y-%m-%d-%H-%M")

        file_name = f"file_report_{current_date_time}_task_{self.task.id}.csv"

        file_report = FileReport()

        # _l.info('TransactionImportProcess.generate_file_report uploading file')

        file_report.upload_file(file_name=file_name, text=result, master_user=self.master_user)
        file_report.master_user = self.master_user
        file_report.name = f"Transaction Import {current_date_time} (Task {self.task.id}).csv"
        file_report.file_name = file_name
        file_report.type = "transaction_import.import"
        file_report.notes = "System File"
        file_report.content_type = "text/csv"

        file_report.save()

        # _l.info('TransactionImportProcess.file_report %s' % file_report)
        # _l.info('TransactionImportProcess.file_report %s' % file_report.file_url)

        return file_report

    def generate_json_report(self):
        serializer = TransactionImportResultSerializer(instance=self.result, context=self.context)

        result = serializer.data

        # _l.debug('self.result %s' % self.result.__dict__)

        # _l.debug('generate_json_report.result %s' % result)

        current_date_time = now().strftime("%Y-%m-%d-%H-%M")
        file_name = f"file_report_{current_date_time}_task_{self.task.id}.json"

        file_report = FileReport()

        _l.info("TransactionImportProcess.generate_json_report uploading file")

        # _l.info('Uploading result len %s' % len(result))

        file_report.upload_file(
            file_name=file_name,
            text=json.dumps(result, indent=4, default=str),
            master_user=self.master_user,
        )
        file_report.master_user = self.master_user
        file_report.name = f"Transaction Import {current_date_time} (Task {self.task.id}).json"
        file_report.file_name = file_name
        file_report.type = "transaction_import.import"
        file_report.notes = "System File"
        file_report.content_type = "application/json"

        file_report.save()

        # _l.info('TransactionImportProcess.json_report %s' % file_report)
        # _l.info('TransactionImportProcess.json_report %s' % file_report.file_url)

        return file_report

    def find_process_type(self):
        if self.task.options_object and "items" in self.task.options_object or ".json" in self.file_path:
            self.process_type = ProcessType.JSON
        elif ".xlsx" in self.file_path:
            self.process_type = ProcessType.EXCEL
        elif ".csv" in self.file_path:
            self.process_type = ProcessType.CSV

        _l.info("TransactionImportProcess.Task %s. process_type %s", self.task, self.process_type)

    def get_default_relation(self, rule_scenario, field):
        # i = field.transaction_type_input
        # TODO PERFORMANCE_ISSUE Maybe performance issue
        i = TransactionTypeInput.objects.get(
            transaction_type__user_code=rule_scenario.transaction_type,
            name=field.transaction_type_input,
        )

        model_class = i.content_type.model_class()

        key = props_map[model_class]

        v = None

        if hasattr(self.ecosystem_default, key):
            v = getattr(self.ecosystem_default, key)

        return v

    def find_default_rule_scenario(self):
        rule_scenarios = self.scheme.rule_scenarios.prefetch_related("fields").all()

        self.default_rule_scenario = None

        for scenario in rule_scenarios:
            if scenario.is_default_rule_scenario:
                self.default_rule_scenario = scenario

    def find_error_rule_scenario(self):
        rule_scenarios = self.scheme.rule_scenarios.prefetch_related("fields").all()

        self.error_rule_scenario = None

        for scenario in rule_scenarios:
            if scenario.is_error_rule_scenario:
                self.error_rule_scenario = scenario

    def get_rule_value_for_item(self, item):
        try:
            return formula.safe_eval(self.scheme.rule_expr, names=item.inputs, context=self.context)

        except Exception as e:
            _l.info(
                f"TransactionImportProcess.Task {self.task}. get_rule_value_for_item "
                f"Exception {e} Traceback {traceback.format_exc()}"
            )
            return None

    def convert_value(self, item, rule_scenario, field, value):
        # TODO PERFORMANCE_ISSUE Maybe performance issue
        i = TransactionTypeInput.objects.get(
            transaction_type__user_code=rule_scenario.transaction_type,
            name=field.transaction_type_input,
        )

        if i.value_type == TransactionTypeInput.STRING:
            return str(value)

        if i.value_type == TransactionTypeInput.SELECTOR:
            return str(value)

        elif i.value_type == TransactionTypeInput.NUMBER:
            return float(value)

        elif i.value_type == TransactionTypeInput.DATE:
            if not isinstance(value, date):
                return formula._parse_date(value)
            else:
                return value

        elif i.value_type == TransactionTypeInput.RELATION:
            model_class = i.content_type.model_class()

            v = None

            try:
                try:
                    # optimized way of getting from prefetched dictionary

                    content_type_key = f"{i.content_type.app_label}.{i.content_type.model}"

                    v = self.prefetched_relations[content_type_key][value]

                except Exception:
                    # old way of getting relations

                    v = model_class.objects.get(master_user=self.master_user, user_code=value)

            except Exception:
                _l.error("User code %s not found for %s", value, field.transaction_type_input)

            if not v:
                if self.scheme.missing_data_handler == "set_defaults":
                    v = self.get_default_relation(rule_scenario, field)

                else:
                    item.status = "error"
                    item.error_message = (
                        f"{item.error_message} Can't find relation of [{field.transaction_type_input}](value:{value})",
                    )

                    raise BookException(code=400, error_message=item.error_message)

            return v

    def get_fields_for_item(self, item, rule_scenario):
        fields = {}

        for field in rule_scenario.fields.all():
            try:
                field_value = formula.safe_eval(field.value_expr, names=item.inputs, context=self.context)
                field_value = self.convert_value(item, rule_scenario, field, field_value)
                fields[field.transaction_type_input] = field_value

            except Exception as e:
                item.status = "error"
                item.error_message = f"{item.error_message}Exception {str(e)}"

                _l.error(
                    f"TransactionImportProcess.Task {self.task}. "
                    f"get_fields_for_item {item} field {field} Exception {e} "
                    f"Traceback {traceback.format_exc()}"
                )

                raise Exception(e) from e

        return fields

    def book(self, item, rule_scenario, error=None):  # noqa: PLR0912, PLR0915
        # _l.info(
        #     'TransactionImportProcess.Task %s. book INIT item %s rule_scenario %s' % (self.task, item, rule_scenario)) # noqa: E501

        try:
            if item.status == "error":
                raise BookException(code=400, error_message=item.error_message)

            fields = self.get_fields_for_item(item, rule_scenario)

            if error:
                fields["error_message"] = str(error)

            if self.scheme.book_uniqueness_settings == ComplexTransactionImportScheme.USE_TRANSACTION_TYPE_SETTING:
                uniqueness_reaction = None
            else:
                uniqueness_reaction = self.scheme.book_uniqueness_settings

            transaction_type_process_instance = TransactionTypeProcess(
                linked_import_task=self.task,
                transaction_type=TransactionType.objects.get(user_code=rule_scenario.transaction_type),
                default_values=fields,
                context=self.context,
                uniqueness_reaction=uniqueness_reaction,
                member=self.member,
                source=item.file_inputs,
                execution_context="import",
            )

            if not item.transaction_inputs:
                item.transaction_inputs = {}

            fields_dict = {}

            for key, value in fields.items():
                fields_dict[key] = str(value)

            if error:
                fields_dict["error_message"] = str(error)

            item.transaction_inputs[rule_scenario.transaction_type] = fields_dict

            transaction_type_process_instance.process()

            item.processed_rule_scenarios.append(rule_scenario)

            if transaction_type_process_instance.complex_transaction:
                trn = TransactionImportBookedTransaction(
                    code=transaction_type_process_instance.complex_transaction.code,
                    text=transaction_type_process_instance.complex_transaction.text,
                    transaction_unique_code=transaction_type_process_instance.complex_transaction.transaction_unique_code,
                )

                item.booked_transactions.append(trn)

            if transaction_type_process_instance.has_errors:
                if transaction_type_process_instance.uniqueness_status == "skip":
                    errors = []

                    if transaction_type_process_instance.general_errors:
                        errors = errors + transaction_type_process_instance.general_errors

                    item.status = "skip"
                    item.message = "Transaction Skipped %s", json.dumps(errors, default=str)

                    self.task.update_progress(
                        {
                            "current": self.result.processed_rows,
                            "total": len(self.items),
                            "percent": round(self.result.processed_rows / (len(self.items) / 100)),
                            "description": f"Going to skip {rule_scenario.transaction_type}",
                        }
                    )

                    # raise BookSkipException(code=409, error_message=item.error_message)

                else:
                    item.status = "error"

                    errors = []

                    if transaction_type_process_instance.general_errors:
                        errors = errors + transaction_type_process_instance.general_errors

                    if transaction_type_process_instance.instruments_errors:
                        errors = errors + transaction_type_process_instance.instruments_errors

                    if transaction_type_process_instance.value_errors:
                        errors = errors + transaction_type_process_instance.value_errors

                    if transaction_type_process_instance.complex_transaction_errors:
                        errors = errors + transaction_type_process_instance.complex_transaction_errors

                    if transaction_type_process_instance.transactions_errors:
                        errors = errors + transaction_type_process_instance.transactions_errors

                    item.error_message = item.error_message + " Book Exception: " + json.dumps(errors, default=str)

                    raise BookException(code=400, error_message=item.error_message)

            else:
                item.status = "success"
                item.message = "Transaction Booked %s", transaction_type_process_instance.complex_transaction

                # _l.info('TransactionImportProcess.Task %s. book SUCCESS item %s rule_scenario %s' % (
                #     self.task, item, rule_scenario))

                self.task.update_progress(
                    {
                        "current": self.result.processed_rows,
                        "total": len(self.items),
                        "percent": round(self.result.processed_rows / (len(self.items) / 100)),
                        "description": f"Going to book {rule_scenario.transaction_type}",
                    }
                )

        except Exception as e:
            _l.error("TransactionImportProcess.Task %s. book Exception %s ", self.task, e)

            if e.__class__.__name__ == "BookException":
                raise BookException(code=400, error_message=str(e)) from e
            elif e.__class__.__name__ == "BookSkipException":
                raise BookSkipException(code=400, error_message=str(e)) from e

            else:
                item.status = "error"
                item.error_message = item.error_message + "Unhandled Exception: " + str(e)

                _l.error("TransactionImportProcess.Task %s. book Traceback %s ", self.task, traceback.format_exc())

                raise BookUnhandledException(code=500, error_message=str(e)) from e

    def fill_with_file_items(self):  # noqa: PLR0912, PLR0915
        _l.info("TransactionImportProcess.Task %s. fill_with_raw_items INIT %s", self.task, self.process_type)

        st = time.perf_counter()

        try:
            if self.process_type == ProcessType.JSON:
                try:
                    _l.info("Trying to get json items from task object options")
                    items = self.task.options_object["items"]

                    self.result.total_rows = len(items)

                    self.file_items = items

                except Exception:
                    _l.info("Trying to get json items from file")

                    with storage.open(self.file_path, "rb") as f:
                        self.file_items = json.loads(f.read())

            if self.process_type == ProcessType.CSV:
                _l.info("ProcessType.CSV self.file_path %s", self.file_path)

                with storage.open(self.file_path, "rb") as f, NamedTemporaryFile() as tmpf:
                    for chunk in f.chunks():
                        tmpf.write(chunk)
                    tmpf.flush()

                    # TODO check encoding (maybe should be taken from scheme)
                    with open(tmpf.name, encoding="utf_8_sig", errors="ignore") as cf:
                        # TODO check quotechar (maybe should be taken from scheme)
                        reader = csv.reader(
                            cf,
                            delimiter=self.scheme.delimiter,
                            quotechar='"',
                            strict=False,
                            skipinitialspace=True,
                        )

                        column_row = None

                        for row_index, row in enumerate(reader):
                            if row_index == 0:
                                column_row = row

                            else:
                                file_item = {}

                                for column_index, value in enumerate(row):
                                    key = column_row[column_index]
                                    file_item[key] = value

                                self.file_items.append(file_item)

                        self.result.total_rows = len(self.file_items)

            if self.process_type == ProcessType.EXCEL:
                with storage.open(self.file_path, "rb") as f, NamedTemporaryFile() as tmpf:
                    for chunk in f.chunks():
                        tmpf.write(chunk)
                    tmpf.flush()

                    os.link(tmpf.name, tmpf.name + ".xlsx")

                    # _l.info('self.file_path %s' % self.file_path)
                    # _l.info('tmpf.name %s' % tmpf.name)

                    wb = load_workbook(filename=tmpf.name + ".xlsx")

                    if (
                        self.scheme.spreadsheet_active_tab_name
                        and self.scheme.spreadsheet_active_tab_name in wb.sheetnames
                    ):
                        ws = wb[self.scheme.spreadsheet_active_tab_name]
                    else:
                        ws = wb.active

                    reader = []

                    if self.scheme.spreadsheet_start_cell == "A1":
                        for r in ws.rows:
                            reader.append([cell.value for cell in r])

                    else:
                        start_cell_row_number = int(re.search(r"\d+", self.scheme.spreadsheet_start_cell)[0])
                        start_cell_letter = self.scheme.spreadsheet_start_cell.split(str(start_cell_row_number))[0]

                        start_cell_column_number = column_index_from_string(start_cell_letter)

                        row_number = 1

                        for r in ws.rows:
                            row_values = []

                            if row_number >= start_cell_row_number:
                                for cell in r:
                                    if cell.column >= start_cell_column_number:
                                        row_values.append(cell.value)

                                reader.append(row_values)

                            row_number = row_number + 1

                    column_row = None

                    for row_index, row in enumerate(reader):
                        if row_index == 0:
                            column_row = row

                        else:
                            file_item = {}

                            for column_index, value in enumerate(row):
                                key = column_row[column_index]
                                file_item[key] = value

                            self.file_items.append(file_item)

                    self.result.total_rows = len(self.file_items)

            _l.info(
                "TransactionImportProcess.Task %s. fill_with_raw_items %s DONE items %s",
                self.task,
                self.process_type,
                len(self.raw_items),
            )

            _l.info(
                "TransactionImportProcess: fill_with_raw_items done: %s",
                f"{time.perf_counter() - st:3.3f}",
            )

        except Exception as e:
            _l.error(
                "TransactionImportProcess.Task %s. fill_with_raw_items %s Exception %s",
                self.task,
                self.process_type,
                e,
            )
            _l.error(
                "TransactionImportProcess.Task %s. fill_with_raw_items %s Traceback %s",
                self.task,
                self.process_type,
                traceback.format_exc(),
            )

    def fill_with_raw_items(self):
        _l.info(f"TransactionImportProcess.Task {self.task}. fill_with_raw_items INIT {self.process_type}")
        st = time.perf_counter()

        for file_item in self.file_items:
            item = {}
            for scheme_input in self.scheme.inputs.all():
                try:
                    item[scheme_input.name] = file_item[scheme_input.column_name]

                except Exception as e:
                    _l.error(
                        f"TransactionImportProcess.Task {self.task}. "
                        f"fill_with_raw_items {self.process_type} Exception "
                        f"{repr(e)}"
                    )
                    continue

            self.raw_items.append(item)

        _l.info(
            f"TransactionImportProcess.Task {self.task}. fill_with_raw_items "
            f"{self.process_type} DONE items {len(self.raw_items)}"
        )
        _l.info(
            "TransactionImportProcess: fill_with_raw_items done: %s",
            f"{time.perf_counter() - st:3.3f}",
        )

    def apply_conversion_to_raw_items(self):
        # EXECUTE CONVERSIONS ON SCHEME INPUTS

        st = time.perf_counter()

        row_number = 1
        for raw_item in self.raw_items:
            conversion_item = TransactionImportConversionItem()
            conversion_item.file_inputs = self.file_items[row_number - 1]
            conversion_item.raw_inputs = raw_item
            conversion_item.conversion_inputs = {}
            conversion_item.row_number = row_number

            for scheme_input in self.scheme.inputs.all():
                ## passing first column from shema page
                try:
                    ## functional by FN-2436: Pass None value to Imported Columns expression when
                    ## the field is absent but expected in the data received
                    # deepcopy raw_item.source item data remains untouched
                    names = deepcopy(raw_item)

                    if scheme_input.name not in names:
                        names[scheme_input.name] = None

                    conversion_item.conversion_inputs[scheme_input.name] = formula.safe_eval(
                        scheme_input.name_expr, names=names, context=self.context
                    )
                except Exception:
                    conversion_item.conversion_inputs[scheme_input.name] = None

            self.conversion_items.append(conversion_item)

            row_number = row_number + 1

        _l.info(
            "TransactionImportProcess: apply_conversion_to_raw_items done: %s",
            f"{time.perf_counter() - st:3.3f}",
        )

    # We have formulas that lookup for rows
    # e.g. transaction_import.find_row
    # so it means, in first iterations we will got errors in that inputs
    def recursive_preprocess(self, deep=1, current_level=0):
        if len(self.preprocessed_items) == 0:
            row_number = 1

            for conversion_item in self.conversion_items:
                preprocess_item = TransactionImportProcessPreprocessItem()
                preprocess_item.file_inputs = conversion_item.file_inputs
                preprocess_item.raw_inputs = conversion_item.raw_inputs
                preprocess_item.conversion_inputs = conversion_item.conversion_inputs
                preprocess_item.row_number = row_number
                preprocess_item.inputs = {}

                self.preprocessed_items.append(preprocess_item)

                row_number = row_number + 1

        for preprocess_item in self.preprocessed_items:
            # CREATE SCHEME INPUTS

            for scheme_input in self.scheme.inputs.all():
                key_column_name = scheme_input.column_name

                try:
                    preprocess_item.inputs[scheme_input.name] = preprocess_item.conversion_inputs[scheme_input.name]

                except Exception as e:
                    preprocess_item.inputs[scheme_input.name] = None

                    if current_level == deep:
                        _l.error("key_column_name %s", key_column_name)
                        _l.error("scheme_input.name %s", scheme_input.name)
                        _l.error("preprocess_item.raw_inputs %s", preprocess_item.conversion_inputs)
                        _l.error(
                            "TransactionImportProcess.Task %s. recursive_preprocess init input %s Exception %s",
                            self.task,
                            scheme_input,
                            e,
                        )

            # CREATE CALCULATED INPUTS

            for scheme_calculated_input in self.scheme.calculated_inputs.all():
                try:
                    names = preprocess_item.inputs

                    # passing second column formulas on schema page
                    value = formula.safe_eval(
                        scheme_calculated_input.name_expr,
                        names=names,
                        context={
                            "master_user": self.master_user,
                            "member": self.member,
                            "request": self.proxy_request,
                            "transaction_import": {"items": self.preprocessed_items},
                        },
                    )

                    preprocess_item.inputs[scheme_calculated_input.name] = value

                except Exception as e:
                    preprocess_item.inputs[scheme_calculated_input.name] = None

                    if current_level == deep:
                        _l.error(
                            f"TransactionImportProcess.Task {self.task}. recursive_preprocess "
                            f"calculated_input {scheme_calculated_input} Exception {e}"
                        )

        if current_level < deep:
            self.recursive_preprocess(deep, current_level + 1)

    def preprocess(self):
        _l.info(f"TransactionImportProcess.Task {self.task}. preprocess INIT")
        st = time.perf_counter()
        # self.recursive_preprocess(deep=2)

        self.scheme.expression_iterations_count = max(self.scheme.expression_iterations_count, 1)

        self.recursive_preprocess(deep=self.scheme.expression_iterations_count)

        for preprocessed_item in self.preprocessed_items:
            item = TransactionImportProcessItem()
            item.row_number = preprocessed_item.row_number
            item.file_inputs = preprocessed_item.file_inputs
            item.raw_inputs = preprocessed_item.raw_inputs
            item.conversion_inputs = preprocessed_item.conversion_inputs
            item.inputs = preprocessed_item.inputs

            self.items.append(item)

        _l.info(f"TransactionImportProcess.Task {self.task}. preprocess DONE items {len(self.preprocessed_items)}")
        _l.info(
            "TransactionImportProcess: preprocess done: %s",
            f"{time.perf_counter() - st:3.3f}",
        )

    def process_items(self):  # noqa: PLR0912, PLR0915
        _l.info(f"TransactionImportProcess.Task {self.task}. process_items INIT")
        st = time.perf_counter()
        index = 0

        # with transaction.atomic():

        for item in self.items:
            item.processed_rule_scenarios = []
            item.booked_transactions = []

            _l.info(
                f"TransactionImportProcess.Task {self.task}. ========= process "
                f"row {str(item.row_number)}/{str(self.result.total_rows)} ========"
            )

            try:
                if self.scheme.filter_expression:
                    # expr = Expression.parseString("a == 1 and b == 2")

                    result = bool(
                        formula.safe_eval(
                            self.scheme.filter_expression,
                            names=item.inputs,
                            context=self.context,
                        )
                    )

                    if not result:
                        item.status = "skip"
                        item.message = "Skipped due filter"

                        _l.info(
                            f"TransactionImportProcess.Task {self.task}. Row skipped due filter {str(item.row_number)}"
                        )
                        continue

                rule_value = self.get_rule_value_for_item(item)

                _l.info(
                    f"TransactionImportProcess.Task {self.task}. ========= process row "
                    f"{str(item.row_number)}/{str(self.result.total_rows)} "
                    f"======== {rule_value} "
                )

                if rule_value:
                    found = False

                    for rule_scenario in self.scheme.rule_scenarios.all():
                        if rule_scenario.status != "skip":
                            selector_values = rule_scenario.selector_values.all()

                            for selector_value in selector_values:
                                if selector_value.value == rule_value:
                                    # sid = transaction.savepoint()
                                    # _l.info("Create checkpoint for %s" % index)

                                    found = True
                                    try:
                                        self.book(item, rule_scenario)

                                        # _l.info("Savepoint commit for %s" % index)
                                        # _l.error("Could not book error scenario %s" % e)
                                        # transaction.savepoint_commit(sid)

                                    except BookSkipException:
                                        # _l.info("BookSkipException")
                                        # transaction.savepoint_rollback(sid)
                                        continue

                                    except (
                                        Exception,
                                        BookUnhandledException,
                                        BookException,
                                    ) as e:
                                        # transaction.savepoint_rollback(sid)

                                        _l.error(
                                            f"Catch BookUnhandledException trying to book error_rule_scenario {e}"
                                        )

                                        try:
                                            self.book(item, self.error_rule_scenario, error=e)
                                            # _l.info("Error Handler Savepoint commit for %s" % index)
                                            # transaction.savepoint_commit(sid)

                                        except Exception as e:
                                            # any exception will work on error scenario
                                            _l.error(f"Could not book error scenario {e}")
                                            # _l.info("Error Handler Savepoint rollback for %s" % index)
                                            # transaction.savepoint_rollback(sid)
                        else:
                            """For case when its actually skip mode"""
                            selector_values = rule_scenario.selector_values.all()

                            for selector_value in selector_values:
                                if selector_value.value == rule_value:
                                    found = True

                    if not found:
                        # sid = transaction.savepoint()
                        # _l.info("Create checkpoint for %s" % index)

                        item.status = "skip"
                        item.message = f"Selector {rule_value} does not match anything in scheme"
                        try:
                            self.book(item, self.default_rule_scenario)
                            # transaction.savepoint_commit(sid)

                        except Exception as e:
                            _l.error(f"Could not book default scenario {e}")
                            # transaction.savepoint_rollback(sid)

                else:
                    try:
                        # sid = transaction.savepoint()

                        item.status = "skip"
                        item.message = f"Selector {rule_value} does not match anything in scheme"

                        self.book(item, self.default_rule_scenario)

                        # transaction.savepoint_commit(sid)

                    except Exception as e:
                        _l.error(f"Could not book default scenario {e}")

                        # transaction.savepoint_rollback(sid)

                self.result.processed_rows = self.result.processed_rows + 1

                self.task.update_progress(
                    {
                        "current": self.result.processed_rows,
                        "total": len(self.items),
                        "percent": round(self.result.processed_rows / (len(self.items) / 100)),
                        "description": f"Row {self.result.processed_rows} processed",
                    }
                )

            except Exception as e:
                item.status = "error"
                item.message = f"Error {e}"

                _l.error(
                    f"TransactionImportProcess.Task {self.task}.  ========= process "
                    f"row {str(item.row_number)} ======== Exception {e}"
                    f" Traceback {traceback.format_exc()}"
                )
            finally:
                index = index + 1

        self.result.items = self.items

        _l.info(f"TransactionImportProcess.Task {self.task}. process_items DONE")
        _l.info(
            "TransactionImportProcess: process_items done: %s",
            f"{time.perf_counter() - st:3.3f}",
        )

    def get_verbose_result(self):
        booked_count = 0
        error_count = 0

        for item in self.result.items:
            if item.status == "error":
                error_count = error_count + 1

            if item.booked_transactions:
                booked_count = booked_count + len(item.booked_transactions)

        result = (
            f"Processed {len(self.items)} rows and successfully booked {booked_count} "
            f"transactions. Error rows {error_count}"
        )

        return result

    def process(self):
        try:
            self.process_items()

        except Exception as e:
            _l.error(
                f"TransactionImportProcess.Task {self.task}. process Exception {e} Traceback {traceback.format_exc()}"
            )

            self.result.error_message = f"General Import Error. Exception {e}"

            if self.execution_context and self.execution_context["started_by"] == "procedure":
                send_system_message(
                    master_user=self.master_user,
                    performed_by="System",
                    description=f"Can't process file. Exception {e}",
                )

        finally:
            self.import_result = TransactionImportResultSerializer(instance=self.result, context=self.context).data

            self.task.result_object = self.import_result

            self.result.reports = []

            self.result.reports.append(self.generate_file_report())
            self.result.reports.append(self.generate_json_report())

            if self.items_has_error():
                error_rows_count = 0
                for result_item in self.result.items:
                    if result_item.status == "error":
                        error_rows_count = error_rows_count + 1

                if error_rows_count != 0:
                    send_system_message(
                        master_user=self.master_user,
                        action_status="required",
                        type="warning",
                        title=f"Transaction Import Partially Failed. Task id: {self.task.id}",
                        description=f"Error rows {error_rows_count}/{len(self.result.items)}",
                    )

            system_message_description = (
                f"New transactions created (Import scheme - {str(self.scheme.name)}) - {len(self.items)}"
            )

            import_system_message_title = "Transaction import (finished)"

            system_message_performed_by = self.member.username

            system_message_title = "New transactions (import from file)"
            if self.process_type == ProcessType.JSON and (
                self.execution_context and self.execution_context["started_by"] == "procedure"
            ):
                system_message_title = "New transactions (import from broker)"
                system_message_performed_by = "System"

                import_system_message_title = "Transaction import from broker (finished)"

            send_system_message(
                master_user=self.master_user,
                performed_by=system_message_performed_by,
                section="import",
                type="success",
                title="Import Finished. Prices Recalculation Required",
                description=(
                    "Please, run schedule or execute procedures to calculate portfolio prices and nav history"
                ),
            )

        attachments = []
        if len(self.result.reports):
            attachments = [self.result.reports[0].id, self.result.reports[1].id]

        send_system_message(
            master_user=self.master_user,
            performed_by=system_message_performed_by,
            section="import",
            type="success",
            title=import_system_message_title,
            attachments=attachments,
        )

        send_system_message(
            master_user=self.master_user,
            performed_by=system_message_performed_by,
            section="transactions",
            type="success",
            title=system_message_title,
            description=system_message_description,
        )

        if self.procedure_instance and self.procedure_instance.schedule_instance:
            self.procedure_instance.schedule_instance.run_next_procedure()

        self.task.verbose_result = self.get_verbose_result()

        if len(self.result.reports):
            self.task.add_attachment(self.result.reports[0].id)
            self.task.add_attachment(self.result.reports[1].id)

        self.task.save()

    def whole_file_preprocess(self):
        if self.scheme.data_preprocess_expression:
            names = {"data": self.file_items}

            try:
                # _l.info("whole_file_preprocess  names %s" % names)

                self.file_items = formula.safe_eval(
                    self.scheme.data_preprocess_expression,
                    names=names,
                    context=self.context,
                )

                # _l.info("whole_file_preprocess  self.raw_items %s" % self.raw_items)

            except Exception as e:
                _l.error(f"Could not execute preprocess expression. Error {repr(e)}")

        _l.info(f"whole_file_preprocess.file_items {len(self.file_items)}")

        return self.file_items
